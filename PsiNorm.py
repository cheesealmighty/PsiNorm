# -*- coding: utf-8 -*-
#!/usr/bin/env python3
"""
PsiNorm Persentil Hesaplayıcı - Versiyon 1.9.1
Copyright (C) 2017 Bilal Bahadır Akbulut & Yavuz Ayhan

Bu program özgür yazılımdır: Özgür Yazılım Vakfı tarafından yayımlanan
 GNU Genel Kamu Lisansı’nın sürüm 3 ya da (isteğinize bağlı olarak) 
 daha sonraki sürümlerinin hükümleri altında yeniden dağıtabilir ve/veya
 değiştirebilirsiniz.

Bu program, yararlı olması umuduyla dağıtılmış olup, programın BİR TEMİNATI YOKTUR;
 TİCARETİNİN YAPILABİLİRLİĞİNE VE ÖZEL BİR AMAÇ İÇİN UYGUNLUĞUNA dair bir teminat da vermez.
 Ayrıntılar için GNU Genel Kamu Lisansı’na göz atınız.

Bu programla birlikte GNU Genel Kamu Lisansı’nın bir kopyasını elde etmiş olmanız gerekir. 
Eğer elinize ulaşmadıysa <http://www.gnu.org/licenses/> adresine bakınız.

--------------------------------------------------------------
Dr. Bilal Bahadır Akbulut
İletişim adresi: b.bahadirakbulut@gmail.com
------------------------------------
PsiNorm Percentile Calculator - Version 1.9.1
Copyright (C) 2017 Bilal Bahadır Akbulut & Yavuz Ayhan

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <http://www.gnu.org/licenses/>.

-------------------------------------------------------------

Dr. Bilal Bahadır Akbulut
contact adress: b.bahadirakbulut@gmail.com

"""
def resetGlobals():
    globals()["patientAdmin"] = None
                
    globals()["patientID"] = None
    
    globals()["patientName"] = None
    
    globals()["patientAge"] = None
    
    globals()["patientSex"] = None
    
    globals()["patientEdu"] = None
    
    globals()["whatToDo"] = None
    
    globals()["testNumToDo"] = None
    
    globals()["patientNotes"] = None
    
def guiSettings():
    import tkinter as tk
    titleText = "PsiNorm Persentil Hesaplayıcı - 1.9.1 - Seçenekler "

    class MyFirstGUI:
        def __init__(self, master):
            self.master = master
            master.title("PsiNorm Persentil Hesaplayıcı")
            
            self.label = tk.Label(master, text=titleText, relief=tk.GROOVE, font= (None, 16))
            self.label.grid(row=0, column=0, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=5)    


            self.label = tk.Label(master, text="Kaydedilecek klasör: ", relief=tk.GROOVE)
            self.label.grid(row=1, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    
         
            self.labelFolderName = tk.Label(master, relief=tk.GROOVE, 
                                  text= "Results (Varsayılan)" if (settings("folder_name")=="default") else settings("folder_name"))
            self.labelFolderName.grid(row=1, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)   
            
            self.entryFolderName = tk.Entry(master, width=50)
            self.entryFolderName.grid(row=1, column=2, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)

            self.buttonFolderNameSave = tk.Button(master, text="Kaydet", command=self.folder_name)
            self.buttonFolderNameSave.grid(row=1, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 

            self.buttonFolderNameDefault = tk.Button(master, text="Varsayılan", command=self.buttonFolderNameDefaultFunc)
            self.buttonFolderNameDefault.grid(row=1, column=4, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
    
    
            self.label = tk.Label(master, text="CSV raporu çıkar: ", relief=tk.GROOVE)
            self.label.grid(row=2, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    
         
            self.button1T = tk.Button(master, text="Evet", command=self.button1, state = (tk.DISABLED if settings("output_csv") else tk.NORMAL))
            self.button1T.grid(row=2, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
            
            self.button1F = tk.Button(master, text="Hayır", command=self.button1, state = (tk.NORMAL if settings("output_csv") else tk.DISABLED))
            self.button1F.grid(row=2, column=2, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)
            
            
            self.label = tk.Label(master, text="Text raporu çıkar: ", relief=tk.GROOVE)
            self.label.grid(row=3, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    
         
            self.button2T = tk.Button(master, text="Evet", command=self.button2, state = (tk.DISABLED if settings("output_txt") else tk.NORMAL))
            self.button2T.grid(row=3, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
            
            self.button2F = tk.Button(master, text="Hayır", command=self.button2, state = (tk.NORMAL if settings("output_txt") else tk.DISABLED))
            self.button2F.grid(row=3, column=2, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)


            self.label = tk.Label(master, text="Excel raporu çıkar: ", relief=tk.GROOVE)
            self.label.grid(row=4, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    
         
            self.button3T = tk.Button(master, text="Evet", command=self.button3, state = (tk.DISABLED if settings("output_excel") else tk.NORMAL))
            self.button3T.grid(row=4, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
            
            self.button3F = tk.Button(master, text="Hayır", command=self.button3, state = (tk.NORMAL if settings("output_excel") else tk.DISABLED))
            self.button3F.grid(row=4, column=2, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)


            self.label = tk.Label(master, text="Excel dosyasının ismi: ", relief=tk.GROOVE)
            self.label.grid(row=5, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    
            
            self.labelExcelName = tk.Label(master, relief=tk.GROOVE, 
                                  text= "Results (Varsayılan)" if (settings("excel_name")=="default") else settings("excel_name"))
            self.labelExcelName.grid(row=5, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)   
            
            self.entryExcelName = tk.Entry(master, width=50)
            self.entryExcelName.grid(row=5, column=2, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)

            self.buttonExcelNameSave = tk.Button(master, text="Kaydet", command=self.excel_name)
            self.buttonExcelNameSave.grid(row=5, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 

            self.buttonExcelNameDefault = tk.Button(master, text="Varsayılan", command=self.buttonExcelNameDefaultFunc)
            self.buttonExcelNameDefault.grid(row=5, column=4, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)
            
            
            self.label = tk.Label(master, text="Rapor dosyası isim formatı: ", relief=tk.GROOVE)
            self.label.grid(row=6, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    
         
            self.button4Date = tk.Button(master, text="Önce tarih", command=self.button4DateFunc, state = (tk.DISABLED if settings("textFileNameFormat")=="firstDate" else tk.NORMAL))
            self.button4Date.grid(row=6, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
            
            self.button4Admin = tk.Button(master, text="Önce uygulayıcı", command=self.button4AdminFunc, state = (tk.DISABLED if settings("textFileNameFormat")=="firstAdmin" else tk.NORMAL))
            self.button4Admin.grid(row=6, column=2, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)

            self.button4ID = tk.Button(master, text="Önce hasta kodu", command=self.button4IDFunc, state = (tk.DISABLED if settings("textFileNameFormat")=="firstID" else tk.NORMAL))
            self.button4ID.grid(row=6, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)
            
            
            self.label = tk.Label(master, text="Z skoru yorumlanması: ", relief=tk.GROOVE)
            self.label.grid(row=7, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    

            self.button5ZNorm = tk.Button(master, text="0-1 Normal, 1-2 Hafif, 2-3 Orta, 3+ Ağır(Varsayılan)", command=self.button5ZNorm, state = (tk.DISABLED if settings("zInterval")=="0-1" else tk.NORMAL))
            self.button5ZNorm.grid(row=7, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)

            self.button5ZAlt = tk.Button(master, text="0-1.5 Normal, 1.5-2 Hafif, 2-3 Orta, 3+ Ağır", command=self.button5ZAlt, state = (tk.DISABLED if settings("zInterval")=="0-1.5" else tk.NORMAL))
            self.button5ZAlt.grid(row=7, column=2, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)
            
            
            self.buttonexit = tk.Button(master, text="Çıkış", command=self.exitFunc)
            self.buttonexit.grid(row=8, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=2)
 
        def exitFunc(self):
            root.destroy()
            
        def folder_name(self):
            new_setting = self.entryFolderName.get()
            if checkGuiInputString(new_setting):
                new_setting = new_setting.replace(" ", "_")
                changeSettingsString("folder_name", new_setting)
                self.labelFolderName.config(text= "Results (Varsayılan)" if (settings("folder_name")=="default") else settings("folder_name"))
            else:
                tk.messagebox.showerror("Hata!", "Lütfen harf, sayı, '-', '_', '(' ve ')' dışında giriş yapmayınız.")
            
        def buttonFolderNameDefaultFunc(self):
            changeSettingsString("folder_name", "default")
            self.labelFolderName.config(text= "Results (Varsayılan)" if (settings("folder_name")=="default") else settings("folder_name"))
            
        def button1(self):
            changeSettingsBool("output_csv")    
            self.button1T.config(state= (tk.DISABLED if settings("output_csv") else tk.NORMAL))
            self.button1F.config(state= (tk.NORMAL if settings("output_csv") else tk.DISABLED))
            
        def button2(self):
            changeSettingsBool("output_txt")     
            self.button2T.config(state= (tk.DISABLED if settings("output_txt") else tk.NORMAL))
            self.button2F.config(state= (tk.NORMAL if settings("output_txt") else tk.DISABLED))
            
        def button3(self):
            changeSettingsBool("output_excel")    
            self.button3T.config(state= (tk.DISABLED if settings("output_excel") else tk.NORMAL))
            self.button3F.config(state= (tk.NORMAL if settings("output_excel") else tk.DISABLED))   
            
        def excel_name(self):
            new_setting = self.entryExcelName.get()
            if checkGuiInputString(new_setting):
                new_setting = new_setting.replace(" ", "_")
                changeSettingsString("excel_name", new_setting)
                self.labelExcelName.config(text= "Results (Varsayılan)" if (settings("excel_name")=="default") else settings("excel_name"))
            else:
                tk.messagebox.showerror("Hata!", "Lütfen harf, sayı, '-', '_', '(' ve ')' dışında giriş yapmayınız.")

        def buttonExcelNameDefaultFunc(self):
            changeSettingsString("excel_name", "default")
            self.labelExcelName.config(text= "Results (Varsayılan)" if (settings("excel_name")=="default") else settings("excel_name"))
            
        def button4DateFunc(self):
            changeSettingsString("textFileNameFormat", "firstDate")
            self.button4Date.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstDate" else tk.NORMAL))
            self.button4Admin.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstAdmin" else tk.NORMAL))
            self.button4ID.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstID" else tk.NORMAL))
            
        def button4AdminFunc(self):
            changeSettingsString("textFileNameFormat", "firstAdmin")
            self.button4Date.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstDate" else tk.NORMAL))
            self.button4Admin.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstAdmin" else tk.NORMAL))
            self.button4ID.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstID" else tk.NORMAL))
            
        def button4IDFunc(self):
            changeSettingsString("textFileNameFormat", "firstID")
            self.button4Date.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstDate" else tk.NORMAL))
            self.button4Admin.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstAdmin" else tk.NORMAL))
            self.button4ID.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstID" else tk.NORMAL))
            
        def button5ZNorm(self):
            changeSettingsString("zInterval", "0-1")
            self.button5ZNorm.config(state = (tk.DISABLED if settings("zInterval")=="0-1" else tk.NORMAL))
            self.button5ZAlt.config(state = (tk.DISABLED if settings("zInterval")=="0-1.5" else tk.NORMAL))
        def button5ZAlt(self):
            changeSettingsString("zInterval", "0-1.5")
            self.button5ZNorm.config(state = (tk.DISABLED if settings("zInterval")=="0-1" else tk.NORMAL))
            self.button5ZAlt.config(state = (tk.DISABLED if settings("zInterval")=="0-1.5" else tk.NORMAL))
            
                
            
    root = tk.Tk()
    tk.Grid.rowconfigure(root, [0, 1, 2, 3, 4, 5, 6, 7, 8], weight=1)
    tk.Grid.columnconfigure(root, [0, 1, 2, 3, 4], weight=1)
    root.config(borderwidth=10, relief=tk.GROOVE)
    
    my_gui = MyFirstGUI(root)
    root.mainloop() 

def guiPatientInfo():
    import tkinter as tk
    titleText = "PsiNorm Persentil Hesaplayıcı - 1.9.1 - Hasta Bilgileri" 
    patientSex = None

    class patientInfo:
        def __init__(self, master):
            self.master = master
            master.title("PsiNorm Persentil Hesaplayıcı")
            
            root.bind("<Return>", self.pressEnter)
            
            self.label = tk.Label(master, text=titleText, relief=tk.GROOVE, font= (None, 16))
            self.label.grid(row=0, column=0, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=5)    

            self.label = tk.Label(master, text="Testi uygulayan kişi: ", relief=tk.GROOVE)
            self.label.grid(row=1, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    

            self.entryPatientAdmin = tk.Entry(master, width=50)
            self.entryPatientAdmin.grid(row=1, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=2)
            self.entryPatientAdmin.focus_force()
            self.label1 = tk.Label(master, text="*", relief=tk.GROOVE)
            self.label1.grid(row=1, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    


            self.label = tk.Label(master, text="Hastanın kodu: ", relief=tk.GROOVE)
            self.label.grid(row=2, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    

            self.entryPatientID = tk.Entry(master, width=50)
            self.entryPatientID.grid(row=2, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=2)

            self.label2 = tk.Label(master, text="*", relief=tk.GROOVE)
            self.label2.grid(row=2, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    


            self.label = tk.Label(master, text="Hastanın ismi: ", relief=tk.GROOVE)
            self.label.grid(row=3, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    

            self.entryPatientName = tk.Entry(master, width=50)
            self.entryPatientName.grid(row=3, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=2)

            self.label3 = tk.Label(master, text="*", relief=tk.GROOVE)
            self.label3.grid(row=3, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    


            self.label = tk.Label(master, text="Hastanın yaşı: ", relief=tk.GROOVE)
            self.label.grid(row=4, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    

            self.entryPatientAge = tk.Entry(master, width=50)
            self.entryPatientAge.grid(row=4, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=2)

            self.label4 = tk.Label(master, text="*", relief=tk.GROOVE)
            self.label4.grid(row=4, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    


            self.label = tk.Label(master, text="Hastanın cinsiyeti: ", relief=tk.GROOVE)
            self.label.grid(row=5, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    
         
            self.buttonFemale = tk.Button(master, text="Kadın", command=self.buttonFemaleFunc, state = (tk.DISABLED if patientSex == "Female" else tk.NORMAL))
            self.buttonFemale.grid(row=5, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
            
            self.buttonMale = tk.Button(master, text="Erkek", command=self.buttonMaleFunc, state = (tk.DISABLED if patientSex == "Male" else tk.NORMAL))
            self.buttonMale.grid(row=5, column=2, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)
                 
            self.label5 = tk.Label(master, text="*", relief=tk.GROOVE)
            self.label5.grid(row=5, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    
            
            
            self.label = tk.Label(master, text="Hastanın toplam eğitim yılı: ", relief=tk.GROOVE)
            self.label.grid(row=6, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    

            self.entryPatientEdu = tk.Entry(master, width=50)
            self.entryPatientEdu.grid(row=6, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=2)

            self.label6 = tk.Label(master, text="*", relief=tk.GROOVE)
            self.label6.grid(row=6, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    


            self.label7 = tk.Label(master, text="Verilerin kaydedileceği yer \n(Seçeneklerden değiştirebilirsiniz.): ", relief=tk.GROOVE)
            self.label7.grid(row=7, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)

            self.label7a = tk.Label(master, text=(settings("excel_path") + settings("excel_name")), relief=tk.GROOVE)
            self.label7a.grid(row=7, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=3)
            #informs user of the designated data location before proceeding

            self.buttonSave = tk.Button(master, text="Kaydet", command=self.saveFunc)
            self.buttonSave.grid(row=8, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
 
            self.buttonAbort = tk.Button(master, text="İptal", command=self.abortFunc)
            self.buttonAbort.grid(row=8, column=2, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
        
        def pressEnter(self, event):
            try:
                root.focus_get().invoke()
            except:
                self.saveFunc()
                pass   
        
        def buttonFemaleFunc(self):
            self.buttonFemale.config(state= tk.DISABLED)
            self.buttonMale.config(state= tk.NORMAL)
            
        def buttonMaleFunc(self):
            self.buttonFemale.config(state= tk.NORMAL)
            self.buttonMale.config(state= tk.DISABLED)
            
        def saveFunc(self):
            patientAdmin = self.entryPatientAdmin.get()
            patientID = self.entryPatientID.get()
            patientName = self.entryPatientName.get()
            patientAge = self.entryPatientAge.get()
            patientEdu = self.entryPatientEdu.get()
            inputError = False

            
            if not checkGuiInputString(patientAdmin):
                self.label1.config(text = "Hata! Lütfen harf, sayı, '-', '_', '(' ve ')' dışında giriş yapmayınız.", bg="red")
                inputError = True
            else:
                self.label1.config(text = "*", bg="green")
                
            if not checkGuiInputString(patientID):
                self.label2.config(text = "Hata! Lütfen harf, sayı, '-', '_', '(' ve ')' dışında giriş yapmayınız.", bg="red")
                inputError = True
            else:
                self.label2.config(text = "*", bg="green")
                
            if not checkGuiInputString(patientName):
                self.label3.config(text = "Hata! Lütfen harf, sayı, '-', '_', '(' ve ')' dışında giriş yapmayınız.", bg="red")
                inputError = True
            else:
                self.label3.config(text = "*", bg="green")
                
            if not checkGuiInputInt(patientAge):
                self.label4.config(text = "Hata! Lütfen sayı dışında giriş yapmayınız.", bg="red")
                inputError = True
            else:
                patientAge = int(patientAge)
                if 150 > patientAge > 0:
                    self.label4.config(text = "*", bg="green")
                else:
                    self.label4.config(text = "Hata! Lütfen geçerli bir sayı giriniz.", bg="red")
                    inputError = True
                
            if str(self.buttonFemale["state"]) == "disabled" or str(self.buttonMale["state"]) == "disabled":
                self.label5.config(text = "*", bg="green")
                if str(self.buttonFemale["state"]) == "disabled":
                    patientSex = "Female"
                if str(self.buttonMale["state"]) == "disabled":
                    patientSex = "Male"       
            else:
                self.label5.config(text = "Hata! Lütfen bir seçim yapınız.", bg="red")
                inputError = True
                
            if not checkGuiInputInt(patientEdu):
                self.label6.config(text = "Hata! Lütfen sayı dışında giriş yapmayınız.", bg="red")
                inputError = True
            else:
                patientEdu = int(patientEdu)
                if 150 > patientEdu >= 0:
                    self.label6.config(text = "*", bg="green")
                else:
                    self.label6.config(text = "Hata! Lütfen geçerli bir sayı giriniz.", bg="red")
                    inputError = True

            if inputError:
                return
            else:
                self.on_saving(patientAdmin, patientID, patientName, patientAge, patientSex, patientEdu)
                                
        def abortFunc(self):
            patientInfo.on_closing()
            
        def on_saving(self, temppatientAdmin, temppatientID, temppatientName, temppatientAge, temppatientSex, temppatientEdu):            
            text_dump = ("Testi uygulayan kişi: " + temppatientAdmin + "\nHastanın kodu: " + temppatientID +
            "\nHastanın ismi: " + temppatientName + "\nHastanın yaşı: " + str(temppatientAge) +
            "\nHastanın cinsiyeti: " + funcLangLocal(temppatientSex) + "\nHastanın toplam eğitim yılı: " + str(temppatientEdu)+
            "\nBu kaydı onaylıyor musunuz?")
            if tk.messagebox.askokcancel("Çıkış", text_dump):
                
                globals()["patientAdmin"] = temppatientAdmin
                
                globals()["patientID"] = temppatientID
                
                globals()["patientName"] = temppatientName
                
                globals()["patientAge"] = temppatientAge
                
                globals()["patientSex"] = temppatientSex
                
                globals()["patientEdu"] = temppatientEdu

                root.destroy()
            
        def on_closing():
            if tk.messagebox.askokcancel("Çıkış", "Bu kaydı iptal etmek istediğinizden emin misiniz?"):
                root.destroy()
    
    root = tk.Tk()
    tk.Grid.rowconfigure(root, [0, 1, 2, 3, 4, 5, 6, 7, 8], weight=1)
    tk.Grid.columnconfigure(root, [0, 1, 2, 3], weight=1)
    root.config(borderwidth=10, relief=tk.GROOVE)
    root.protocol("WM_DELETE_WINDOW", patientInfo.on_closing)
    my_gui = patientInfo(root)
    root.mainloop() 


def getUserDocumentsPsiPath():
    import ctypes.wintypes
    CSIDL_PERSONAL= 5       # My Documents
    SHGFP_TYPE_CURRENT= 0   # Want current, not default value
    
    buf= ctypes.create_unicode_buffer(ctypes.wintypes.MAX_PATH)
    ctypes.windll.shell32.SHGetFolderPathW(0, CSIDL_PERSONAL, 0, SHGFP_TYPE_CURRENT, buf)
    
    documentsPath = buf.value
    psiPath = documentsPath + "/PsiNorm/"
    ensure_dir(psiPath)
    return psiPath


def guiAgreeTerms():
    import tkinter as tk
    from tkinter import messagebox
    from tkinter.scrolledtext import ScrolledText as ScrolledText
    import sys
    
    class MyFirstGUI:
        def __init__(self, master):
            self.master = master
            master.title("Kullanım Koşulları")
            text_dump = jsonLoader("info_data")["agreeTerms_data"]
            self.label = ScrolledText(master, wrap = tk.WORD, width = 80, height = 30)
            self.label.grid(row=0, column=0, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=4) 
            self.label.insert(tk.INSERT, text_dump)
            self.label.config(state=tk.DISABLED)
            self.label.insert(tk.END, " in ScrolledText")
 
            self.greet_button = tk.Button(master, width = 10, height = 1, text="Evet", command=self.greet, relief=tk.GROOVE)
            self.greet_button.grid(row=1, column=1, padx=10, pady=10, sticky=tk.S)

            self.close_button = tk.Button(master, width = 10, height = 1, text="Hayır", command=self.areyousure, relief=tk.GROOVE)
            self.close_button.grid(row=1, column=2, padx=10, pady=10, sticky=tk.S)
    
    
        def greet(self):
            root.destroy()
            
        def areyousure(self):
            answer = messagebox.askyesno("Uyarı!", "Programı kullanabilmek için kabul etmelisiniz. \nBir önceki ekrana dön?")
            if answer:
                return
            else:
                sys.exit()
                
        def on_closing():
            answer = messagebox.askyesno("Uyarı!", "Programı kullanabilmek için kabul etmelisiniz. \nBir önceki ekrana dön?")
            if answer:
                return
            else:
                sys.exit()
    
    root = tk.Tk()
    
    tk.Grid.rowconfigure(root, [0,1], weight=1)
    tk.Grid.columnconfigure(root, [0,1,2,3], weight=1)
    root.config(borderwidth=10, relief=tk.GROOVE)
    
    my_gui = MyFirstGUI(root)
    root.protocol("WM_DELETE_WINDOW", MyFirstGUI.on_closing)
    root.mainloop()
    return True

def guiSimpleTextDump(gui_title, text_dump):
    import tkinter as tk
    from tkinter.scrolledtext import ScrolledText as ScrolledText
    
    class MyFirstGUI:
        def __init__(self, master):
            self.master = master
            master.title(gui_title)
            
            root.bind("<Return>", self.pressEnter)
            
            self.label = ScrolledText(master, wrap = tk.WORD,width  = 80, height = 30)
            self.label.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
            self.label.insert(tk.INSERT, text_dump)
            self.label.config(state=tk.DISABLED)
            self.label.insert(tk.END, " in ScrolledText")
       
            self.close_button = tk.Button(master, text="Tamam.", command=self.close)
            self.close_button.pack(side=tk.BOTTOM)
            self.close_button.focus_force()
            
        def pressEnter(self, event):
            try:
                root.focus_get().invoke()
            except:
                pass   
            
        def close(self):
            root.destroy()
    
    root = tk.Tk()
    my_gui = MyFirstGUI(root)
    root.mainloop()

def guiPatientNotes(gui_title, text_dump):
    import tkinter as tk
    from tkinter.scrolledtext import ScrolledText as ScrolledText
    
    class patientNoteGUI:
        def __init__(self, master):
            self.master = master
            master.title(gui_title)
            
            self.noteBox = ScrolledText(master, wrap = tk.WORD,width  = 80, height = 30)
            self.noteBox.grid(row=0, column=0, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=4) 
            self.noteBox.insert(tk.INSERT, text_dump)
            self.noteBox.focus_force()
            
            self.noteBox.bind("<Key>", self.resetSaveStatus)
 
            self.save_button = tk.Button(master, height=2, width=10, text="Kaydet", command=self.save)
            self.save_button.grid(row=1, column=1, padx=5, pady=5, sticky=tk.S)  
      
            self.close_button = tk.Button(master, height=2, width=10, text="Çıkış", command=self.close)
            self.close_button.grid(row=1, column=2, padx=5, pady=5, sticky=tk.S) 
            
            self.saveStatus = tk.Label(master, height=1, width=40, text="")
            self.saveStatus.grid(row=2, column=1, padx=5, pady=5, sticky=tk.S, columnspan=2) 
            
        def resetSaveStatus(self, event):
            self.saveStatus.config(text="")
            
        def save(self):
            globals()["patientNotes"] = self.noteBox.get("1.0", tk.END)
            
            self.saveStatus.config(text="Test kaydedildi.")
            
        def close(self):
            tempPatientNotes = self.noteBox.get("1.0", tk.END)
            if tempPatientNotes != patientNotes:
                if tk.messagebox.askokcancel("Çıkış", "Değişiklikleri kaydetmeden çıkmak istediğinize emin misiniz?"):
                    root.destroy()
                else: 
                    return
            else:
                root.destroy()
    
    root = tk.Tk()
    
    tk.Grid.rowconfigure(root, [0,1,2], weight=1)
    tk.Grid.columnconfigure(root, [0,1,2,3], weight=1)
    root.config(borderwidth=10, relief=tk.GROOVE)
    
    my_gui = patientNoteGUI(root)
    root.mainloop()


def guiSimplePopup(gui_title, text_dump):
    import tkinter as tk
    
    class MyFirstGUI:
        def __init__(self, master):
            self.master = master
            master.title(gui_title)
            
            self.label = tk.Label(master, text=text_dump)
            self.label.grid(row=1, column=1, padx=10, pady=10, sticky=tk.N)    
       
            self.greet_button = tk.Button(master, text="Tamam", command=self.close)
            self.greet_button.grid(row=2, column=1, padx=10, pady=10, sticky=tk.S) 
            
        def close(self):
            root.destroy()
    
    root = tk.Tk()
    my_gui = MyFirstGUI(root)
    root.mainloop()

def guiFAQ():
    text_dump = jsonLoader("info_data")["FAQ_data"]
    gui_title = "Sık Sorulan Sorular"
    guiSimpleTextDump(gui_title, text_dump)


def guiAbout():
    text_dump = jsonLoader("info_data")["about_data"]
    gui_title = "PsiNorm Hakkında"
    guiSimpleTextDump(gui_title, text_dump)


def guiReferences():
    text_dump = jsonLoader("info_data")["references_data"]
    gui_title = "Kaynakça"
    guiSimpleTextDump(gui_title, text_dump)

def guiStartupMenu():
    
    resetGlobals()
    
    import tkinter as tk
    from tkinter import messagebox
    import sys
    titleText = " PsiNorm Persentil Hesaplayıcı - 1.9.1"
    
    titleSubText= """
 Copyright (C) 2017 Bilal Bahadır Akbulut & Yavuz Ayhan
 
 Bu programın KESİNLİKLE HİÇBİR TEMİNATI YOKTUR; ayrıntılar için programın içinde bulunduğu klasördeki lisans dosyalarına başvurunuz.

 Bu bir özgür yazılımdır ve bazı koşullar altında yeniden dağıtmakta serbestsiniz; ayrıntılar için programın içinde bulunduğu klasördeki lisans dosyalarına başvurunuz.
 """
    titleWarning = "HESAPLAMALAR BU EKRANIN YANINDA AÇILAN KONSOLDAN YAPILMAKTADIR, LÜTFEN KAPATMAYINIZ."

    class startupGUI:
        def __init__(self, master):
            self.master = master
            master.title("PsiNorm Persentil Hesaplayıcı")
            
            psiNormIcon = tk.PhotoImage(file = "psiNormIcon.gif")
            root.bind("<Return>", self.pressEnter)

            self.label = tk.Label(master,image=psiNormIcon)
            self.label.image = psiNormIcon
            self.label.grid(row=0, column=0, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=6)    
            
            self.label = tk.Label(master, text=titleText, bg="white", relief=tk.GROOVE, font=(None, 16))
            self.label.grid(row=1, column=0, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=6)    
            
            self.label = tk.Label(master, text=titleSubText, bg="white", relief=tk.GROOVE, font=(None, 10))
            self.label.grid(row=2, column=0, padx=5, pady=5, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=6)   
            
            self.label = tk.Label(master, text=titleWarning, bg="white", relief=tk.GROOVE, font=(None, 12))
            self.label.grid(row=3, column=0, padx=5, pady=5, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=6)   
            
            self.buttonStartup = tk.Button(master, text="Programı çalıştır", command=self.startup)
            self.buttonStartup.grid(row=4, column=0, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
            self.buttonStartup.focus_force()
    
            self.buttonguiFAQ = tk.Button(master, text="Sık Sorulan Sorular", command=self.guiFAQ)
            self.buttonguiFAQ.grid(row=4, column=1, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
            
            self.buttonguiAbout = tk.Button(master, text="Hakkında", command=self.guiAbout)
            self.buttonguiAbout.grid(row=4, column=2, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
            
            self.buttonguiReferences = tk.Button(master, text="Kaynakça", command=self.guiReferences)
            self.buttonguiReferences.grid(row=4, column=3, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)
            
            self.buttonMenuSettings = tk.Button(master, text="Ayarlar", command=self.menuSettings)
            self.buttonMenuSettings.grid(row=4, column=4, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
            
            self.buttonProgramExit = tk.Button(master, text="Çıkış Yap", command=self.programExit)
            self.buttonProgramExit.grid(row=4, column=5, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
    
        def pressEnter(self, event):
            try:
                root.focus_get().invoke()
            except:
                pass   
            
        def guiFAQ(self):
            guiFAQ()
            
        def guiAbout(self):
            guiAbout()
            
        def guiReferences(self):
            guiReferences()
            
        def menuSettings(self):
            guiSettings()
        
        def programExit(self):
            startupGUI.on_closing()
            
        def on_closing():
            if messagebox.askokcancel("Çıkış", "Programı kapatmak istediğinizden emin misiniz?"):
                sys.exit()
                
        def startup(self):
            root.destroy()          
    
    root = tk.Tk()
    tk.Grid.rowconfigure(root, [0,1,2,3,4], weight=1)
    tk.Grid.columnconfigure(root, [0,1,2,3,4,5], weight=1)
    root.config(borderwidth=10, relief=tk.GROOVE)
    my_gui = startupGUI(root)
    root.protocol("WM_DELETE_WINDOW", startupGUI.on_closing)
    root.mainloop() 

def menuAutoCreate(menuDict):
    """
    Returns the template for the guiTestChoose.
    doneTests = Dict formatting must be done before printing.
    
    menuDict = 
        {"1":
            {"testDataDict": "testDataDict1",
            "testName": "testName1",
            "testGroupName":"testGroupName1"}
            } 
    """
    testGroupDict = {}
    for testNum in menuDict.keys():
        try:
            testGroupDict[menuDict[testNum]["testGroupName"]].append(testNum)
        except:
            testGroupDict[menuDict[testNum]["testGroupName"]] = [testNum]
     #Groups the tests under the group names       
    
    decorator = """|>"""
    for i in range(62):
        decorator += "="
    decorator += "<|>"
    for i in range(63):
        decorator += "="
    decorator += "<|\n"
    
    textDump = """"""
    textDump += decorator
    
    from natsort import natsorted as nt
    
    testGroupNamesList = nt(testGroupDict.keys(), key=lambda y: y.lower())
    #Naturally sorts group names so they look orderly
    
    i = 0
    while i < len(testGroupNamesList):
        howManyLinesEven = 0
        evenList = []
        howManyLinesOdd = 0
        oddList = []
        
        testGroupNameEven = testGroupNamesList[i]
        textDump += "| {:54}         |".format(testGroupNameEven)
        howManyLinesEven = len(testGroupDict[testGroupNameEven])
        
        evenList = nt(testGroupDict[testGroupNameEven], key=lambda y: y.lower())
        #If it's an even number, it puts the menu template on the left side of the screen
        
        
        if i != len(testGroupNamesList)-1:
            testGroupNameOdd = testGroupNamesList[i+1]
            textDump += " {:55}         |".format(testGroupNameOdd) + "\n"
            howManyLinesOdd = len(testGroupDict[testGroupNameOdd])   
            oddList = nt(testGroupDict[testGroupNameOdd], key=lambda y: y.lower())
        #If it's odd, on the right side.  
            
        if i == len(testGroupNamesList)-1:
            textDump += "{:56}         |".format("") + "\n"
            #Ensures everything is correctly whitespaced
            
        howManyLines = max(howManyLinesEven, howManyLinesOdd)
        #Checks how many lines there are, so if a group has less tests, it will have extra whitespaces
        
        for line in range(howManyLines):
            if line < howManyLinesEven:
                data = {"testNum": evenList[line], "testName": menuDict[evenList[line]]["testName"]}
                textDump += "|{d[testNum]:>3}) {d[testName]:55}{{doneTests[{d[testNum]}]:^4}}|".format(d=data)
            else:
                textDump += "|{:64}|".format("")
            
            if line < howManyLinesOdd:
                data = {"testNum": oddList[line], "testName": menuDict[oddList[line]]["testName"]}
                textDump += "{d[testNum]:>3}) {d[testName]:56}{{doneTests[{d[testNum]}]:^4}}|".format(d=data) + "\n"
            else:
                textDump += "{:65}|".format("") + "\n"
            #Automatically creates a menu
            
        textDump += decorator
        
        i += 2
   
    return textDump


def guiTestChoose(title, menuDict, textDump):
    import tkinter as tk
    from tkinter.scrolledtext import ScrolledText as ScrolledText
    from tkinter import messagebox
    import sys
    
    class testChooseGUI:
        def __init__(self, master, title):
            self.master = master
            master.title(title)
            
            root.bind("<Return>", self.pressEnter)
            root.protocol("WM_DELETE_WINDOW", self.close)
            
            
            self.label = ScrolledText(master, wrap = tk.WORD,width  = 80, height = 30)
            self.label.grid(row=0, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=2)
            self.label.insert(tk.INSERT, textDump)
            self.label.config(state=tk.DISABLED)
            self.label.insert(tk.END, " in ScrolledText")
            
            self.status = tk.Label(master, text="", relief=tk.GROOVE)
            self.status.grid(row=1, column=0, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=2)  
            
            self.label = tk.Label(master, text="Yapılacak Test No (Bir sayı girip ENTER tuşuna basınız.): ", relief=tk.GROOVE)
            self.label.grid(row=2, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)    
            
            self.entryUserInput = tk.Entry(master)
            self.entryUserInput.bind("<Key>", self.resetStatus)
            self.entryUserInput.focus_force()
            self.entryUserInput.grid(row=2, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)
      
            self.patientNotes = tk.Button(master, text="Hasta Notları", command=self.getPatientNotes)
            self.patientNotes.grid(row=3, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)
            
            self.patientInfoForms = tk.Button(master, text="Hasta Veri Formları", command=self.getPatientInfoForms)
            self.patientInfoForms.grid(row=3, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)
            
            self.save_button = tk.Button(master, text="Kaydet ve Çık", command=self.save)
            self.save_button.grid(row=4, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)

            self.close_button = tk.Button(master, text="Kaydetmeden Çıkış", command=self.close)
            self.close_button.grid(row=4, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)   
      
        def resetStatus(self, event):
            self.status.config(text="")
        
        def pressEnter(self, event):
            try:
                if root.focus_get() == self.entryUserInput:
                    self.entry()
                else:
                    root.focus_get().invoke()
            except:
                self.status.config(text="Lütfen uygun bir seçim yapınız.")
                
        def getPatientNotes(self):
            globals()["whatToDo"] = "getPatientNotes"
            root.destroy()
        
        def getPatientInfoForms(self):
            globals()["whatToDo"] = "getPatientInfoForms"
            root.destroy()
        
        def entry(self):
            userInput = self.entryUserInput.get()

            if checkGuiInputInt(userInput):
                if str(userInput) in menuDict.keys():  
                    globals()["whatToDo"] = "doTest"
                    globals()["testNumToDo"] = userInput
                    root.destroy()
                else:
                    self.status.config(text="Lütfen üstteki listede belirtilen sayılardan birini seçiniz.")

            else:
                self.status.config(text="Uygunsuz test numarası, lütfen tekrar deneyiniz.")
                
        def save(self):
            text = "Verileri kaydetip çıkış yapılacak, emin misiniz?"
            answer = messagebox.askyesno("Uyarı!", text)
            if answer:
                globals()["whatToDo"] = "saveProgram"
                root.destroy()
            else:
                return
            
        def close(self):
            text = "KAYDETMEDEN çıkılacak, emin misiniz?"
            answer = messagebox.askyesno("Uyarı!", text)
            if answer:
                globals()["whatToDo"] = "exitPatient"
                root.destroy()           
            else:
                return
            
    
    root = tk.Tk()
    tk.Grid.rowconfigure(root, [0, 1, 2, 3, 4], weight=1)
    tk.Grid.columnconfigure(root, [0, 1], weight=1)
    root.config(borderwidth=10, relief=tk.GROOVE)
    
    root.state('zoomed')
    
    my_gui = testChooseGUI(root, title)
    
    root.mainloop()

def guiTest(testDataDict):
    import tkinter as tk
    from tkinter import messagebox as messagebox
    
    totalParaNum = int(testDataDict["paraNum"])
    paraTypeDict = testDataDict["paraTypeDict"]
    testName = testDataDict["testName"]
    
    globals()["resultDict"] = None
    
    class testGUI:
        def __init__(self, master):
            self.master = master
            master.title(testName)
            
            root.bind("<Return>", self.pressEnter)
            root.bind("<Key>", self.autoCalculate)
            
            self.entriesList = []
            self.entriesSanity = []
            
            for paraNum in range(totalParaNum):
                # create left side info labels
                tk.Label(text=testDataDict[str(paraNum)]).grid(row=paraNum, column=0, padx=1, pady=1)
                # create entries list
                self.entriesList.append(tk.Entry(width=40))
                # grid layout the entries
                self.entriesList[paraNum].grid(row=paraNum, column=1, padx=1, pady=1)
                
                if str(paraNum) in testDataDict["mathOper"].keys():
                    self.entriesList[paraNum].insert(tk.END, "Otomatik hesaplanır.")
                    self.entriesList[paraNum].config(state=tk.DISABLED)
                    
                self.entriesSanity.append(tk.Label(text = "*", bg='grey', width=40))
                
                self.entriesSanity[paraNum].grid(row=paraNum, column=2, padx=1, pady=1)
    
            self.entriesList[0].focus_force()

            self.status = tk.Label(master, height=1, width=120, text="")
            self.status.grid(row=totalParaNum, column=0, padx=5, pady=5, sticky=tk.S, columnspan=3) 
    
            self.save_button = tk.Button(master, height=2, width=10, text="Kaydet ve Çık", state=tk.DISABLED, command=self.save)
            self.save_button.grid(row=totalParaNum+1, column=1, padx=5, pady=5, sticky=tk.S)  
            self.save_button.bind("<Enter>", self.saveStatus)
            self.save_button.bind("<Leave>", self.resetStatus)
        
        
            self.close_button = tk.Button(master, height=2, width=10, text="Çıkış", command=self.close)
            self.close_button.grid(row=totalParaNum+1, column=2, padx=5, pady=5, sticky=tk.S) 
            
        def pressEnter(self, event):
            try:
                root.focus_get().invoke()
            except:
                if self.save_button["state"] != "disabled":
                    self.save()
                else:
                    pass   
          
        def resetStatus(self, event):
            self.status.config(text = "")
            
        def saveStatus(self, event):
            if self.save_button["state"] == "disabled":
                text = "Kaydetmek için lütfen yapılmayan veya yapılamayan testlere 999 giriniz."
            else:
                text = ""
            
            self.status.config(text = text)
                
            
        def autoCalculate(self, event):
            tempResultDict = {}
            self.save_button.config(state=tk.NORMAL)
            for paraNum in range(totalParaNum):
                paraNum = str(paraNum)
                tempResultDict[paraNum] = 999
            try:
                for paraNum in range(totalParaNum):
                    paraNum = str(paraNum)
                    result = self.entriesList[int(paraNum)].get()
                    if result == "":
                        self.save_button.config(state=tk.DISABLED)
                    inputSane = True
                       
                    if paraNum not in testDataDict["mathOper"].keys():
                        if paraTypeDict[paraNum] == "intType":
                            inputSane = checkGuiInput(result, "flo")
                            if not inputSane:
                                self.entriesSanity[int(paraNum)].config(text= "Lütfen sadece sayı giriniz.")
                            else:
                                tempResultDict[paraNum] = float(result)
                        elif paraTypeDict[paraNum] == "strType":
                            tempResultDict[paraNum] = str(result)
                        
                        if not inputSane:
                            self.entriesSanity[int(paraNum)].config(bg = "red")
                            self.save_button.config(state=tk.DISABLED)
                        else:
                            self.entriesSanity[int(paraNum)].config(text= "*", bg = "green")
                            
                            
                    else:
                        tempResultDict[paraNum] = 999
                        
                        try:
                            mathList = testDataDict["mathOper"][paraNum]
                            
                            firstVal = tempResultDict[str(mathList[0])]
                            secondVal = tempResultDict[str(mathList[2])]
                            operator = mathList[1]
                            
                            bothResultsExist = 999 not in [firstVal, secondVal]
                            
                            if bothResultsExist:
                                if operator == "+":
                                    tempResultDict[paraNum] = firstVal + secondVal
                                elif operator == "-":
                                    tempResultDict[paraNum] = firstVal - secondVal
                                elif operator == "*":
                                    tempResultDict[paraNum] = firstVal * secondVal
                                elif operator == "/":
                                    if secondVal == 0:
                                        secondVal = 0.00000001 #If secondVal is equal to 0, protects program from failing.
                                    tempResultDict[paraNum] = firstVal / secondVal
                                else:
                                    print("Yanlış matematik operatörü, kullanılabilir seçenekler: +,-,*,/")
                                    
                                    text = "Yanlış matematik operatörü, kullanılabilir seçenekler: +,-,*,/"
                                    self.entriesSanity[int(paraNum)].config(text= text)
                                    
                                    tempResultDict[paraNum] = 999
                                    
                            else:
                                text = "Bir ya da birden fazla test parametresi eksik."
                                self.entriesSanity[int(paraNum)].config(text= text, bg="red")
                                
                                self.entriesList[int(paraNum)].config(state=tk.NORMAL)
                                self.entriesList[int(paraNum)].delete(0, tk.END)
                                
                                self.entriesList[int(paraNum)].insert(tk.END, "Otomatik hesaplanır.")
                                self.entriesList[int(paraNum)].config(state=tk.DISABLED)

                            if bothResultsExist and tempResultDict[paraNum] != 999:
                                self.entriesList[int(paraNum)].config(state=tk.NORMAL)
                                self.entriesList[int(paraNum)].delete(0, tk.END)

                                self.entriesList[int(paraNum)].insert(tk.END, str(tempResultDict[paraNum]))
                                self.entriesList[int(paraNum)].config(state=tk.DISABLED)
                                
                                text = "*"
                                self.entriesSanity[int(paraNum)].config(text= text, bg="green")
                                
                        except SystemExit:
                            raise
                        except:
                            if settings("debug"):
                                raise
                            print("KRİTİK TEST HATASI(mathOper), EĞER DATA DOSYALARINDA DEĞİŞİKLİK YAPTI İSENİZ KONTROL EDİNİZ.")
                            tempResultDict[paraNum] = 999
                            pass
                    
                    if result == "":
                        self.entriesSanity[int(paraNum)].config(text= "*", bg = "grey")
                        
                globals()["resultDict"] = tempResultDict
                
                        
            except:
                raise
            
        def save(self):
            root.destroy()
            
        def close(self):
            if tk.messagebox.askokcancel("Çıkış", "Değişiklikleri kaydetmeden çıkmak istediğinize emin misiniz?"):
                globals()["resultDict"] = None
                root.destroy()
                
            else: 
                return

    
    root = tk.Tk()
    
    rowList = []
    i = 0
    while i < testDataDict["paraNum"] + 2:
        rowList.append(i)
        i += 1
        
    tk.Grid.rowconfigure(root, rowList, weight=1)
    tk.Grid.columnconfigure(root, [0,1,2], weight=1)
    root.config(borderwidth=10, relief=tk.GROOVE)
    
    my_gui = testGUI(root)
    root.mainloop()
    return resultDict


def criticalError(errorTitle, errorMessage, shouldIBeep):
    """
    errorTitle = Title of the popup, string, if None, prints a console message instead.
    errorMessage = string
    shouldIbeep = Should it make a beep sound, boolean
    """
    
    if shouldIBeep:
        import winsound
        try:
            winsound.Beep(440, 50)
        except:
            pass
    
    if errorTitle:
        guiSimplePopup(errorTitle, errorMessage)
        
    else:
        print(errorMessage)
           
        
def goBackToDefault(fileName):
    print("Bir şeyler yanlış gitti, program varsayılan seçeneklere dönüyor.")
    
    import os
    cwd = os.getcwd()
    
    import shutil
    
    """
    import shutil
    shutil.copy2('/src/dir/file.ext', '/dst/dir/newname.ext') # complete target filename given
    shutil.copy2('/src/file.ext', '/dst/dir') # target filename is /dst/dir/file.ext
    """
    missingFile = fileName
    missingFileList = missingFile.split("/Data")
    missingFileR = missingFileList[1]
    
    defaultFile = cwd + "/Data/Default" + missingFileR   
    
    dirToEnsure = ""
    i = 0
    while i < len(missingFile.split("/"))-1:
        dirToEnsure += missingFile.split("/")[i] + "/"
        i += 1
        
    ensure_dir(dirToEnsure)
    shutil.copy(defaultFile, missingFile)


def funcLangLocal(item):
    if item == "Female":
        localItem = "Kadın"
    elif item == "Male":
        localItem = "Erkek"
        
    return localItem


def jsonLoader(jsonFileName):
    # Currently: "form_data", "info_data"["agreeTerms_data", "FAQ_data", "about_data", "references_data"]
    documentsPath = getUserDocumentsPsiPath()
    
    import configparser
    parser = configparser.ConfigParser()
    jsonAddressList = documentsPath + "/Data/Cogs/" + "jsonAddressList.ini"
    
    parser.read(jsonAddressList, encoding = 'utf-8-sig')
        
    jsonFileExists = False
    
    for section in parser.sections():
        if jsonFileName in parser[section]:
            jsonFileExists = True
            sectionToUse = section
        
    if not jsonFileExists:
        goBackToDefault(jsonAddressList)
        parser.read(jsonAddressList, encoding = 'utf-8-sig')
        
        for section in parser.sections():
            if jsonFileName in parser[section]:
                jsonFileExists = True
                sectionToUse = section
    
    if not jsonFileExists:
        class jsonFileLooped(Exception):
            """What did  y'all do?!? """

        raise jsonFileLooped(
"""
jsonAddressList.ini içerisinde talep edilen JSON dosyası mevcut değil.
Eğer kişiselleştirilmiş test eklediyseniz, lütfen doğru adresi kaydettiğinizden emin olunuz.
Eğer herhangi bir değişiklik yapmadınız ve buna rağmen bu uyarıyı görüyorsanız lütfen programcıya ulaşınız.
""")    
      
    jsonFileAddress = parser[sectionToUse][jsonFileName]
    
    if sectionToUse == 'DefaultTests':
        jsonFileAddress = documentsPath + "/Data/Test/" + jsonFileAddress 
        
    elif sectionToUse == 'CustomTests':
        jsonFileAddress = documentsPath + "/Data/Test/Custom/" + jsonFileAddress
    
    elif sectionToUse == 'InternalFiles':
        jsonFileAddress = documentsPath + "/Data/" + jsonFileAddress
    
    try:
        import json
        with open(jsonFileAddress, 'r', encoding="utf8") as fp:
            mainDict = json.load(fp)
        
        fp.close()
        return mainDict
    
    except:
        try:           
            print("JSON file missing. Restoring the default.")
            goBackToDefault(jsonFileAddress)
            
            import json
            with open(jsonFileAddress, 'r', encoding="utf8") as fp:
                mainDict = json.load(fp)
            
            fp.close()
            return mainDict
        
        
        except:
            print("Default JSON file missing. User error?")
            raise

def timeGlobalization():
    from time import strftime
    if "date" in globals():
        globals()["date"] = strftime("%Y-%m-%d")
    else:
        global date
        date = strftime("%Y-%m-%d")
        
    if "time" in globals():
        globals()["time"] = strftime("%H:%M:%S")
    else:    
        global time
        time = strftime("%H:%M:%S")
    
    if "stime" in globals():
        globals()["stime"] = strftime("%H%M")
    else:
        global stime
        stime = strftime("%H%M")
    

def wait(t):
    import time
    while t:
        print(str(t))
        time.sleep(1)
        t -= 1

def exitable_input(interface):
    #An interface system to exit at any time
    while True:
        thing = input(interface)
        if thing == "exit":
            sure = input("Programı kapatmak istediğinizden emin misiniz? (e)vet/(h)ayır: ")
            #https://youtu.be/BLikP6BDH5w
            if sure in ["e", "E", "evet", "Evet"]: 
                import sys
                sys.exit()
            else:
                continue               
        else:
            return thing

def checkGuiInput(userInput, inputType):
    import re
    if inputType == "str":
        if re.match("^[0-9a-zA-Z\-\_\(\)ığĞüÜşŞİöÖçÇ ]+$", userInput):
            return True
        else:
            return False
        
    if inputType == "flo":
        try:
            userInput = float(userInput)
            return True
        except:
            return False
    
    if inputType == "int":
        try:
            userInput = int(userInput)
            return True
        except:
            return False

def checkGuiInputString(userInput):
    import re
    if re.match("^[0-9a-zA-Z\-\_\(\)ığĞüÜşŞİöÖçÇ ]+$", userInput):
        return True
    else:
        return False
    
def checkGuiInputFloat(userInput):
    try:
        userInput = float(userInput)
        return True
    except:
        return False
    
def checkGuiInputInt(userInput):
    try:
        userInput = int(userInput)
        return True
    except:
        return False
    
def amILooping():
    if not "crashNum" in globals():
        global crashNum
        crashNum = 1
    else:
        globals()["crashNum"] = crashNum + 1
        
    if crashNum > 10:
        print("Houston we have a problem here.")
        raise
    

def hexaInput(questionString):
    import re
    while True:
        result = exitable_input(questionString)
        if not re.match("^[0-9a-zA-Z\-\_\(\)\/ığĞüÜşŞİöÖçÇ ]+$", result):
            print("Lütfen sadece sayı, harf ve/ya \"-\" giriniz.")
            continue
        else:
            break
    return result

def freeInput(questionString):
    result = exitable_input(questionString)
    return result

def numInput(questionString):
    import re
    while True:
        result = exitable_input(questionString)
        if not re.match("^[0-9]+$", result):
            print("Lütfen sadece sayı giriniz.")
            continue
        else:
            break
    return result
    
def floInput(questionString):
    while True:
        result = exitable_input(questionString)
        try:
            result = float(result)
            break
        
        except SystemExit:
            raise
        
        except:
            print("Lütfen sadece sayı giriniz.")
            continue
    return result

    
def ensure_dir(file_path):
    import os
    try:
        os.makedirs(file_path, exist_ok = True)
    
    except:
        print("Houston we have a problem here")
        
    return

def changeSettingsBool(setting_name):
    #Enables the user to change settings within the program
    configFilePath = getUserDocumentsPsiPath() + "settings.ini"
    try:
        import configparser
        parser = configparser.ConfigParser()
        parser.read(configFilePath, encoding = 'utf-8-sig')
        
        from configobj import ConfigObj
        config = ConfigObj(configFilePath, encoding='UTF8')
         
        state = parser.get('General', setting_name)
        
        if state == "True":
            state = "False"
            
        elif state == "False":
            state = "True"    
        
            
        config['General'][setting_name] = state   
        config.write()
                      
    except SystemExit:
        raise
    except:
        raise
        print("Lütfen yalnızca sayı giriniz")
    
    return
    
def changeSettingsString(setting_name, new_setting):
    #Enables the user to change settings within the program
    configFilePath = getUserDocumentsPsiPath() + "settings.ini"
    try:
        import configparser
        parser = configparser.ConfigParser()
        parser.read(configFilePath, encoding = 'utf-8-sig')
        
        from configobj import ConfigObj
        config = ConfigObj(configFilePath, encoding='UTF8')
                
        state = parser.get('General', setting_name)
        
        state = new_setting
                  
        config['General'][setting_name] = state   
        config.write()
                      
    except SystemExit:
        raise
    except:
        raise
    
    return

def settings(which_setting):
    #returns a boolean whether if it should create a file or not. Also a string for data path
    import os
    cwd = os.getcwd()
    configFilePath = getUserDocumentsPsiPath() + "settings.ini"
    x = True
    
    documentsPath = getUserDocumentsPsiPath()
    
    if which_setting == "cogsFolder":
        setting_to_return = cwd + "/Data/Cogs/"
        return setting_to_return
    
    
    if which_setting in ["csv_path", "txt_path", "excel_path"]:
        if (which_setting == "csv_path"):
            setting_to_return =  settings("folder_name") + ("csv/")
        if (which_setting == "txt_path"):
            setting_to_return =  settings("folder_name") + ("txt/")
        if (which_setting == "excel_path"):
            setting_to_return =  settings("folder_name") + ("excel/")
        ensure_dir(setting_to_return)
        return setting_to_return
    
    while x:
        try:
            import configparser
            parser = configparser.ConfigParser()
            parser.read(configFilePath, encoding = 'utf-8-sig')
            
            setting_to_return = parser.get('General', which_setting)
            
            if (which_setting == "excel_name"):
                if setting_to_return == "default":
                    setting_to_return = "PsiNorm"
                                
                setting_to_return = setting_to_return + ".xlsx"
                x = False
                return setting_to_return
            
            if (which_setting == "folder_name"):
                if setting_to_return == "default":
                    setting_to_return =  documentsPath + ("Results/")
                else:
                    setting_to_return = documentsPath + (setting_to_return) + "/"
                    
            if which_setting == "textFileNameFormat":
                return setting_to_return
            
            if which_setting == "zInterval":
                return setting_to_return

            if setting_to_return == "True":
                setting_to_return = True
                
            if setting_to_return == "False":
                setting_to_return = False
                
            if setting_to_return != True and setting_to_return != False:
                ensure_dir(setting_to_return)
                #Ensures the file path is applicable and available
                
            x = False
            return setting_to_return
        
        except:
            amILooping()
            print("Ayarlar dosyası bulunamadı veya hatalı, tekrar oluşturuluyor.")
            with open(configFilePath, 'w', encoding='utf-8-sig') as file:
                file.write(
"""[General]
folder_name = default

output_csv = False

output_txt = True

output_excel = True
excel_name = PsiNorm

textFileNameFormat = firstDate

jsonaddresslist = jsonaddresslist.ini

debug = False

zInterval = 0-1

first_ever_run = True
"""
)
            continue

def excelWriter(excel_path, data_num, printable_list, which_data, excelColNameDict):
    while True:
        try:
            patientName_local = patientName      
                
            demographic_data = [patientID, patientName_local, patientAdmin, date, time, patientAge, funcLangLocal(patientSex), patientEdu]
            
            
            excelWriter_data = jsonLoader("excelWriter_data")
            
            excelWriter_data["test_name_list"]
            
            test_name_list = excelWriter_data["test_name_list"]

            masterDataList = [excelWriter_data["userNotesHandle"]] + test_name_list
                         
            testExcelColNameDict = { "masterData": masterDataList, **excelWriter_data["testExcelColNameDict"]}
            

            excel_sheet_names_list = []
            
            mainDict = jsonLoader("form_data")
                
            for i in mainDict["formNames"].keys():
                excel_sheet_names_list.append(mainDict[i]["excelSheetName"])
            
            
            excel_sheet_names_list = ["Ana Veri"] + excel_sheet_names_list + test_name_list
            
            from openpyxl import Workbook
            from openpyxl import load_workbook
            
            wb = Workbook(write_only=True)
            
            while True:
                try:
                    for sheet_name in excel_sheet_names_list:
                        data_workbook = load_workbook(filename = excel_path + settings("excel_name"), read_only=False)
                        active_sheet = data_workbook[sheet_name]
                        data_workbook.close()
                    break
                
                except PermissionError:
                    guiSimpleTextDump("Uyarı!", "Excel dosyası arka planda açık görünüyor, lütfen kapatınız. ")
                    continue
                
                except:
                    userInput = numInput("""
Excel dosyası bulunamadı veya hatalı. 
> Dosyayı tekrar kontrol edip, tekrar denemek için (1)
> Excel kaydını iptal etmek için (2) DİKKAT, BU SEÇENEK PROGRAMI KAPATIR.
> Dosyayı tekrar oluşturmak için (3)
giriniz. 
Cevap: """)
                    if userInput == "1":
                        continue
                    elif userInput == "2":
                        print("Excel kaydı iptal edildi. ")
                        import sys
                        sys.exit()
                    elif userInput == "3":
                        userInput = hexaInput("""
Bu işlem eğer eski bir excel dosyası var ise silecektir.
Eğer kayıtlı veriniz var ise, lütfen yedekleme yapınız.
Bu işlemi yapmak istediğinizden emin misiniz? (e)vet/(h)ayır: """)
                        if userInput in ["e", "E", "evet", "Evet"]: 
                            for sheet_name in excel_sheet_names_list:
                                wb.create_sheet(title = sheet_name)
                                                   
                            wb.save(filename = excel_path + settings("excel_name"))
                            data_workbook = load_workbook(filename = excel_path + settings("excel_name"), read_only=False)
                            
                            
                            for key in testExcelColNameDict.keys():
                                if key != "masterData":
                                    active_sheet = data_workbook[test_name_list[int(key)-1]]
                                else:
                                    active_sheet = data_workbook["Ana Veri"]
                                test_col_list = ["Hasta Kodu","Hasta isim", "Uygulayıcı", "Tarih", "Zaman", "Yaş", "Cinsiyet","Eğitim Yılı"]
                                test_col_list = test_col_list + testExcelColNameDict[key]
                                for col, val in enumerate(test_col_list, start=1):
                                    active_sheet.cell(row=1, column=col).value = val
                                    
                            
                            data_workbook.save(filename = excel_path + settings("excel_name"))    
                            data_workbook.close()
                                
                            break
                        else:
                            continue
                    else:
                        print("Lütfen varolan seçeneklerden seçiniz. ")
                        continue
        
            
            
            if which_data == "testData":
                try:
                    while True:
                        try:
                            data_workbook = load_workbook(filename = excel_path + settings("excel_name"), read_only=False)
                            active_sheet = data_workbook[test_name_list[data_num-1]]
        
                            active_sheet.append(demographic_data + printable_list)
            
                            data_workbook.save(filename = excel_path + settings("excel_name"))
                            print(test_name_list[data_num-1] + " Excel'e kaydediliyor. ")
                            data_workbook.close()
                            break
                        
                        except PermissionError:
                            print("Dosya arka planda açık görünüyor, lütfen kapatınız. Program 5 saniye sonra tekrar deneyecek.")
                            wait(5)
                            continue
                        
                        except:
                            print("Excel sayfası bulunamadı. " + test_name_list[data_num-1] + " oluşturuluyor. ")
                            for i in range(len(test_name_list)):
                                wb.create_sheet(title = test_name_list[i])
                
                            wb.save(filename = excel_path + settings("excel_name"))
                            data_workbook = load_workbook(filename = excel_path + settings("excel_name"), read_only=False)
                            active_sheet = data_workbook["Sheet"]
                            data_workbook.remove_sheet(active_sheet)
                            data_workbook.save(filename = excel_path + settings("excel_name"))
                            data_workbook.close()
                            continue
                        
                except:
                    raise
            
            elif which_data == "masterData":
                try:
                    while True:
                        try:
                            data_workbook = load_workbook(filename = excel_path + settings("excel_name"), read_only=False)
                            active_sheet = data_workbook["Ana Veri"]
                            #sets the correct sheet name
                            
                            doneTests = printable_list
                            
                            masterBinaryList = []
                            
                            for num in range(len(test_name_list)):
                                if doneTests[num+1] == "(+)":
                                    masterBinaryList.append("1")
                                else:
                                    masterBinaryList.append("0")
                            
                            #converts whichever tests are done to a binary value list
                            
                            active_sheet.append(demographic_data + [patientNotes] + masterBinaryList)
            
                            data_workbook.save(filename = excel_path + settings("excel_name"))
                            print("Ana veri sayfası Excel'e kaydediliyor. ")
                            data_workbook.close()
                            break
                        
                        except PermissionError:
                            print("Dosya arka planda açık görünüyor, lütfen kapatınız. Program 5 saniye sonra tekrar deneyecek.")
                            wait(5)
                            continue
                        
                        except:
                            print("Excel sayfası bulunamadı. " + "Ana Veri" + " oluşturuluyor. ")
                            wb.create_sheet(title = "Ana Veri")
                
                            wb.save(filename = excel_path + settings("excel_name"))
                            data_workbook = load_workbook(filename = excel_path + settings("excel_name"), read_only=False)
                            active_sheet = data_workbook["Sheet"]
                            data_workbook.remove_sheet(active_sheet)
                            data_workbook.save(filename = excel_path + settings("excel_name"))
                            data_workbook.close()
                            continue
                        
                except:
                    raise
            
                
                    
            else:
                final_list = []
                colNum_list = ["Hasta Kodu","Hasta isim", "Uygulayıcı", "Tarih", "Zaman", "Yaş", "Cinsiyet","Eğitim Yılı"]
                for i in range(len(printable_list)):
                    colNum_list.append(str(printable_list[i][0]) + ") " + str(excelColNameDict[printable_list[i][0]]))
                    final_list.append(printable_list[i][1])
                                
        
                while True:
                    try:
                        data_workbook = load_workbook(filename = excel_path + settings("excel_name"), read_only=False)
                        active_sheet = data_workbook[which_data]
                        
                        for col, val in enumerate(colNum_list, start=1):
                            active_sheet.cell(row=1, column=col).value = val
                        
                        active_sheet.append(demographic_data + final_list)
        
                        data_workbook.save(filename = excel_path + settings("excel_name"))
                        print(which_data + " Excel'e kaydediliyor. ")
                        data_workbook.close()
                        break
                    
                    except:
                        print("Excel sayfası bulunamadı. " + which_data + " oluşturuluyor. ")
                            
                        wb.create_sheet(title = which_data)
                        
                        wb.save(filename = excel_path + settings("excel_name"))
                        data_workbook = load_workbook(filename = excel_path + settings("excel_name"), read_only=False)
                        active_sheet = data_workbook["Sheet"]
                        data_workbook.remove_sheet(active_sheet)
                        data_workbook.save(filename = excel_path + settings("excel_name"))
                        data_workbook.close()
                        continue

            break
     
        except PermissionError:
            print("Dosya arka planda açık görünüyor, lütfen kapatınız. Program 5 saniye sonra tekrar deneyecek.")
            wait(5)
            continue
        except:
            raise
            

def csvWriter(path, test_name, printable_list):
    while True:
        try:
            import csv
            
            test_name = path + test_name
            
            with open(test_name, 'a', encoding='UTF-8', newline='') as csvfile:
                data_writer = csv.writer(csvfile, delimiter='~', quotechar='|', quoting=csv.QUOTE_MINIMAL)
                data = [patientAdmin, date, time, patientID, patientAge, funcLangLocal(patientSex), patientEdu] + printable_list
                data_writer.writerow(data)
                
            break
        except PermissionError:
            print("Dosya arka planda açık görünüyor, lütfen kapatınız. Program 5 saniye sonra tekrar deneyecek.")
            wait(5)
            continue
        except:
            raise
        
def txtWrite(path, console_results):
    while True:
        try:
            
            fileNameFormat = settings("textFileNameFormat")
            
            if fileNameFormat == "firstID":
                file_name = path + str(patientID.replace(" ", "_")) + "-(" + date + "-" + stime + ")-" + patientAdmin.replace(" ", "_") + ".txt"
            
            if fileNameFormat == "firstDate":
                file_name = path + "(" + date + "-" + stime + ")-" + str(patientID.replace(" ", "_")) + "-" + patientAdmin.replace(" ", "_") + ".txt"
                
            if fileNameFormat == "firstAdmin":
                file_name = path + patientAdmin.replace(" ", "_") + "-(" + date + "-" + stime + ")-" + str(patientID.replace(" ", "_")) + ".txt"
            
            import os.path
            if os.path.exists(file_name):
                with open(file_name, 'a', encoding='UTF-8') as file:
                    file.write("\n" + console_results)
                    print("Rapora ek yapılıyor. ")
                file.close()
                    
            else:
                with open(file_name, 'w', encoding='UTF-8') as file:
                    file.write("======= PsiNorm Persentil Hesaplayıcı =======" +
                               "\n============ Copyright (C) 2017 =============" +
                               "\n==== Bilal Bahadır Akbulut & Yavuz Ayhan ====" +
                    "\n=============================================\n" + 
                    "Testi uygulayan: " + patientAdmin + "\nGünün tarihi: " + date + "\nSaat: " + time +
                    "\nHastanın ismi: " +  patientName + "\nHastanın kodu: " + str(patientID) +
                    "\nHastanın yaşı: " + str(patientAge) + "\nHastanın cinsiyeti: " + str(funcLangLocal(patientSex)) +
                    "\nHastanın toplam eğitim yılı: " + str(patientEdu) +
                    "\n=============================================\n")
                    
                    file.write(console_results)
                    print("Rapor oluşturuluyor. ")
                file.close()
            break
            
        except PermissionError:
            print("Dosya arka planda açık görünüyor, lütfen kapatınız. Program 5 saniye sonra tekrar deneyecek.")
            wait(5)
            continue
        except:
            raise


def inputpatientAge():
#a simple loop to get patient's age  
    while True:   
        patientAge = exitable_input("Hastanın yaşı: ")
        
        if patientAge.isnumeric():
            patientAge = int(patientAge)
            
            if patientAge < 0:
                #ensures the number is bigger than 0
                print("Lütfen 0'dan büyük bir sayı giriniz.")
                continue
            else:
                break

        else:
            print("Lütfen sadece sayı giriniz.")
            continue
        
    return patientAge
    
def inputpatientSex():   
#a simple loop to get the sex of the patient, provides two options, female or male
    while True:    
        patientSex_user_input = exitable_input("Hastanın cinsiyeti: (1) Kadın - (2) Erkek: ")
        if patientSex_user_input == "1":
            patientSex = "Female"
            break
        elif patientSex_user_input == "2":
            patientSex = "Male"
            break
        else:
            print("Lütfen sadece 1 veya 2 giriniz.")
            continue
    return patientSex
        
def inputpatientEdu(): 
#a simple loop to get the number of years of education
    while True:
        patientEdu = exitable_input("Hastanın toplam eğitim yılı: ")
        if patientEdu.isnumeric():
            patientEdu = int(patientEdu)
            break
        else:
            print("Lütfen sadece sayı giriniz.")
            continue
        
    return patientEdu

def calculate_age(patientAge):
    #gets day of birth and ensures it's in correct format
    from datetime import datetime, date
    today = date.today()
    global birthday
    
    while True:
        try: 
            print("Lütfen doğum tarihini giriniz.")
            global bornDay
            bornDay = numInput("Gün: ")
            global bornMonth
            bornMonth = numInput("Ay: ")
            global bornYear
            bornYear = numInput("Yıl: ")
            
            birthday = (bornDay + "/" + bornMonth + "/" + bornYear)
            
            if "999" in [bornDay, bornMonth, bornYear]:
                if bornYear == "999":
                    print("Yıl bilinmiyor, yaş hesaplanamaz.")
                    formPatientAge = "999"
                    
                elif bornYear != "999" and bornMonth == "999":
                    birthday = datetime.strptime(bornYear, "%Y")                    
                    print("Sadece yıl üzerinden yaş hesaplanıyor.")
                    formPatientAge = today.year - birthday.year
                    
                elif bornYear != "999" and bornMonth != "999":
                    birthday = datetime.strptime(bornMonth+"/"+bornYear, "%m/%Y")
                    print("Sadece ay ve yıl üzerinden yaş hesaplanıyor.")   
                    formPatientAge = today.year - birthday.year - ((today.month) < (birthday.month))
                    
            else:              
                birthday = datetime.strptime(birthday, "%d/%m/%Y")
                formPatientAge = today.year - birthday.year - ((today.month, today.day) < (birthday.month, birthday.day))
            
            if formPatientAge < 0 or formPatientAge > 150:
                print("Lütfen girdiğiniz tarihi kontrol edip tekrar giriniz.")
                continue
                
        except ValueError:
            print("Hatalı tarih, lütfen tekrar giriniz.\n")
            continue
        except SystemExit:
            raise
        except:
            raise
          
        birthday = (bornDay + "/" + bornMonth + "/" + bornYear)
        
        if patientAge != formPatientAge:
            #if there's discrepeancy between the calculated age and age entered, forces user to choose
            print("""
Girişte girilen yaş ile forma girilen yaş arasında fark bulundu.
Hangi sayıyı kullanmak istersiniz?
1) Girişte girilen: """ + str(patientAge) + """
2) Formda hesaplanan: """ + str(formPatientAge))
            
            while True:
                userInput = int(numInput("\n"))
                if 0 < userInput < 3:
                    if userInput == 1:
                        formPatientAge = patientAge
                    else:
                        globals()['patientAge'] = formPatientAge
                    
                    break
                
                else:
                    print("Lütfen '1' veya '2' giriniz.")
                    continue
                
        return birthday

def formMain(mainDict): #main function for the dataforms
    def eQuest(qDict, qNum, templateDict): #expandable questions
        
        ansDict = {}
        
        if qNum in templateDict.keys():
            ansDict = templateDict[qNum]
        
        def greet(qDict, ansDict):
            #prints out the data structure
            print("\n" + qNum + ") " + qDict["qText"])  
            
            toprint = ""
            toprintList = []
            for x in range(len(ansDict.keys())):
                toprint = str(x+1) + ") " +  str(ansDict[str(x+1)])
                
                toprintList.append(toprint)
                        
                        
            for i in range(len(toprintList)):
                print(toprintList[i])
        
            print("--------------------------------------------")    
               
        def addMoreFunc(qDict, ansDict): #add one item
            tempList = []
            for i in range(qDict["colNum"]):
                tempList.append(freeInput(qDict[str(i + 1)]))
                
            tempString = ""
            for i in range(len(tempList)):
                tempString = tempString + tempList[i]
                if len(tempList) > 1 and (i+1) < len(tempList):
                    tempString = tempString + " / "
                    
            ansDict[str(len(ansDict.keys()) + 1)] = tempString
            
            return ansDict
            
        def changeFunc(qDict, ansDict): #change an item
            while True:
                userInput = numInput("Lütfen değiştirmek istediğiniz maddeyi giriniz: ")
                if 0 < int(userInput) < len(ansDict.keys()) + 1:
                    tempList = []
                    for i in range(qDict["colNum"]):
                        tempList.append(freeInput(qDict[str(i + 1)]))
                        
                    tempString = ""
                    for i in range(len(tempList)):
                        tempString = tempString + tempList[i]
                        if len(tempList) > 1 and (i+1) < len(tempList):
                            tempString = tempString + " / "
                            
                    ansDict[userInput] = tempString
                    
                    return ansDict
                    
                else:
                    print("Lütfen varolan maddelerden birini seçiniz. ")
                    continue
                    
        
        while True:
            greet(qDict, ansDict)
            try:
                print("""
Seçenekler:
1) Yeni madde ekle.
2) Varolan maddeyi değiştir.
3) Varolan maddeyi sil.
4) Bir sonraki soruya geç.                              
                """)
                
                userInput = int(numInput("Lütfen seçiniz: "))
                if userInput == 1:
                    ansDict = addMoreFunc(qDict, ansDict)
                elif userInput == 2:
                    if len(ansDict.keys()) < 1:
                        print("Değiştirilecek girdi mevcut değildir, lütfen başka bir seçim yapınız.")
                        continue
                    else:
                        ansDict = changeFunc(qDict, ansDict)
                        continue
                
                elif userInput == 3:
                    if len(ansDict.keys()) < 1:
                        print("Silinecek girdi mevcut değildir, lütfen başka bir seçim yapınız.")
                        continue
                    else:
                        userInput = numInput("Silincecek madde numarası: ")
                        if 0 < int(userInput) <= len(ansDict.keys()):
                            to_shorten = len(ansDict.keys()) - int(userInput) 
                            if to_shorten > 0:
                                for i in range(to_shorten):
                                    ansDict[str(int(userInput) + i)] = ansDict[str(int(userInput) + i + 1)]
                                
                                del ansDict[str(len(ansDict.keys()))]
                            
                            if to_shorten == 0:
                                del ansDict[userInput]
                            
                            continue
                        else:
                            print("Lütfen varolan maddelerden seçiniz. ")
                            continue
                
                elif userInput == 4:
                    return ansDict
                

            except SystemError:
                raise
            except:
                raise
            
        
    #eQuest(qDict)
    
    
    def mQuest(qDict, qNum): #multiple choice question
        while True:
            try:
                print("\n" + qNum + ") " + qDict["qText"])
                for i in range(qDict["optNum"]):
                    print(str(i+1) + ") " + qDict[str(i+1)])
                userInput = numInput("Cevap: ")            
            except SystemError:
                raise
            except:
                raise
                
            if 0 < int(userInput) <= qDict["optNum"]:
                break
            else:
                print("Lütfen varolan seçeneklerden birini seçiniz.")
                continue
            
        if str(userInput) in qDict["extraMark"]:
            return qDict[userInput], (qNum + "-" + userInput)
        else:
            return qDict[userInput], None
                
    
    def cQuest(qDict, qNum): #classic question
        while True:
            try:
                userInput = freeInput(qNum + ") " + qDict["qText"])
            except SystemError:
                raise
            except:
                print("Bir hata oluştu, lütfen bildiriniz.")
                raise
            return userInput
          
    def decideFunc(qDict, qNum, templateDict): #checks which form of question it s
        extraMark = None
        
        if qDict["qType"] == "eQuest":
            answer = eQuest(qDict, qNum, templateDict)
            
        elif qDict["qType"] == "mQuest":
            answer, extraMark = mQuest(qDict, qNum)
            
        elif qDict["qType"] == "cQuest":
            answer = cQuest(qDict, qNum)
            
        elif qDict["qType"] == "runFunc":
            arg = eval(qDict["arg"])
            answer = eval(qDict["runFunc"])(arg)
            
        elif qDict["qType"] == "pullData":
            print("\n" + qNum + ") " + qDict["qText"] + str(eval(qDict["data"])))       
            answer = eval(qDict["data"])
        
        else:
            print("Beklenmeyen bir hata oluştu, program kapatılacak.")
            raise
            
        return answer, extraMark
            
    
    
    def formFunc(mainDict, excelColNameDict): #director of the questions
        endAnsDict = {}
        
        for i in range(len(list(mainDict["qDict"].keys()))):
            if str(i+1) in mainDict["titleDict"].keys():
                print(mainDict["titleDict"][str(i+1)])
                        
            qNum = str(i+1)
            qDict = mainDict["qDict"][qNum]
            xtraQDict = mainDict["xtraQDict"]
            templateDict = mainDict["templateDict"]
            
            endAnsDict[qNum], extraMark = decideFunc(qDict, qNum, templateDict)
            
            extraMarkList = [extraMark]
            
                
            while True:
                while None in extraMarkList:   
                    extraMarkList.remove(None)
                        
                if len(extraMarkList) > 0:
                    extraMarkToUse = extraMarkList[0]
                    extraMarkList.remove(extraMarkList[0])
                    
                    for i in range(len(list(xtraQDict[extraMarkToUse].keys()))):
                        i = str(i + 1)
                        ansKey = extraMarkToUse + "-" + i
                        endAnsDict[ansKey], extraMark = decideFunc(xtraQDict[extraMarkToUse][i], i, templateDict)
                        if extraMark != None:
                            extraMarkList.append(extraMarkToUse + "-" + extraMark)                 
                        
                        print("--------------------------------------------")
                        
                else:
                    break
                            
            print("--------------------------------------------")
      
        while True:
            userInput = numInput("""
Form bitmiştir.
> Değiştirmek istediğiniz soru var ise (1)
> Kaydetmek için (2)
giriniz: """)
            if userInput == "1":
                while True:
                    userInput = numInput("Değiştirmek istediğiniz sorunun numarasını giriniz: ")
                    if userInput in endAnsDict.keys():
                        for i in endAnsDict.keys():
                            key = i
                            if key.startswith(userInput + "-"):
                                del endAnsDict[key]
                        
                        qNum = userInput
                        qDict = mainDict["qDict"][qNum]
                        xtraQDict = mainDict["xtraQDict"]
                        templateDict = mainDict["templateDict"]
                        
                        endAnsDict[qNum], extraMark = decideFunc(qDict, qNum, templateDict)
                
                        extraMarkList = [extraMark]
                        
                        while True:
                            while None in extraMarkList:   
                                extraMarkList.remove(None)
                                    
                            if len(extraMarkList) > 0:
                                extraMarkToUse = extraMarkList[0]
                                extraMarkList.remove(extraMarkList[0])
                                
                                for i in range(len(list(xtraQDict[extraMarkToUse].keys()))):
                                    i = str(i + 1)
                                    ansKey = extraMarkToUse + "-" + i
                                    endAnsDict[ansKey], extraMark = decideFunc(xtraQDict[extraMarkToUse][i], i, templateDict)
                                    if extraMark != None:
                                        extraMarkList.append(extraMarkToUse + "-" + extraMark)                 
                                    
                                    print("--------------------------------------------")
                                    
                            else:
                                break
                                        
                        print("--------------------------------------------")
                        
                        break
                        
                    else:
                        print("Lütfen varolan sorulardan seçiniz.")
                        continue
                
                
            elif userInput == "2":
                break
            
            else:
                print("Lütfen varolan seçeneklerden birini seçiniz.")
                continue
            
        
        endAnsList, console_results = ansListDecoder(mainDict, endAnsDict, excelColNameDict)
        return endAnsList, console_results
    
      
    def ansListDecoder(mainDict, endAnsDict, excelColNameDict): #translates the answer list into excel & txt format
        endConsAnsList = []
        endExcAnsList = []
        endConsAnsDict = endAnsDict.copy() #readies up a console dictionary
        import natsort
        
        for key in natsort.natsorted(endAnsDict.keys(), key=lambda y: y.lower()):
            endExcAnsList.append([key,endAnsDict[key]])
            endConsAnsList.append([key,endAnsDict[key]])
            #sorts answer dictionary to a list for accounting expanding questions
    
        for i in endExcAnsList:
            qNum = i[0]
            ans = i[1]
            if isinstance(ans, dict):
                expList = []
                
                for key in natsort.natsorted(ans.keys(), key=lambda y: y.lower()):
                    expList.append(ans[key])
                    
                for i in range(len(expList)): #makes expQ answers compatible with the rest of the answers
                    endConsAnsDict[qNum + "-" + str(i+1)] = expList[i]                
                del endConsAnsDict[qNum]
                
                if len(expList) <= 20: #ensures there are 20 columns in the excel file for expanding Q
                    tempString = ""
                    for i in range(mainDict["qDict"][qNum]["colNum"]):
                        tempString = tempString + '999'
                        if mainDict["qDict"][qNum]["colNum"] > 1 and (i+1) < mainDict["qDict"][qNum]["colNum"]:
                            tempString = tempString + " / "
                    for i in range(20-len(expList)):                    
                        expList.append(tempString)
                        
                for i in range(len(expList)): #makes expQ answers compatible with the rest of the answers
                    endAnsDict[qNum + "-" + str(i+1)] = expList[i]
                del endAnsDict[qNum]
                 
                
        excelColNamesList = keyCounter(mainDict)
        for i in excelColNamesList:
            if i not in endAnsDict.keys():
                endAnsDict[i] = "-"
        
                
        endExcAnsList = []
        for key in natsort.natsorted(endAnsDict.keys(), key=lambda y: y.lower()):
            endExcAnsList.append([key,endAnsDict[key]])
            #finalizes the list
            
        endConsAnsList = []
        for key in natsort.natsorted(endConsAnsDict.keys(), key=lambda y: y.lower()):
            endConsAnsList.append([key,endConsAnsDict[key]])
            #finalizes the list
            
        console_results = "\n========================================================\n"
        for i in range(len(endConsAnsList)):
            console_results = console_results + str(endConsAnsList[i][0]) + ") "
            #if str(endExcAnsList[i][0]) in mainDict["qDict"].keys():
             #   console_results = console_results + mainDict["qDict"][str(endExcAnsList[i][0])]["qText"]
            #if str(endExcAnsList[i][0]) in mainDict["xtraQDict"].keys():
            #    console_results = console_results + mainDict["xtraQDict"][str(endExcAnsList[i][0])]["qText"]               
             
            console_results = console_results + excelColNameDict[str(endConsAnsList[i][0])]
            
            console_results = console_results + str(endConsAnsList[i][1]) + "\n"
                    
        console_results = console_results + "========================================================"

        return endExcAnsList, console_results
        
            
    def keyCounter(mainDict): #ensures all the proper columns exists within the excel file
        excelColNameDict = {}
        
        for i in mainDict["qDict"].keys():       
            key = i
            if mainDict["qDict"][key]["qType"] == "eQuest":
                for i in range(20):  
                    excelColNameDict[key + "-" + str(i+1)] = mainDict["qDict"][key]["qText"] + "-" + str(i+1)
                
            elif mainDict["qDict"][key]["qType"] == "mQuest":
                excelColNameDict[key] = mainDict["qDict"][key]["qText"]
                for i in mainDict["qDict"][key]["extraMark"]:
                    if i != None:
                        xtraKey = i
                        for i in mainDict["xtraQDict"][key + "-" + xtraKey].keys():
                            excelColNameDict[key + "-" + xtraKey + "-" + i] = mainDict["xtraQDict"][key + "-" + xtraKey][i]["qText"]
                            doubleXtraKey = i
                            if "extraMark" in mainDict["xtraQDict"][key + "-" + xtraKey][doubleXtraKey].keys():
                                excelColNameDict[key + "-" + xtraKey + "-" + doubleXtraKey] = mainDict["xtraQDict"][key + "-" + xtraKey][doubleXtraKey]["qText"]
    
    
            elif mainDict["qDict"][key]["qType"] == "cQuest":
                excelColNameDict[key] = mainDict["qDict"][key]["qText"]
                
            elif mainDict["qDict"][key]["qType"] == "runFunc":
                excelColNameDict[key] = mainDict["qDict"][key]["qText"]
                
            elif mainDict["qDict"][key]["qType"] == "pullData":
                excelColNameDict[key] = mainDict["qDict"][key]["qText"]
                
            else:
                print("This shouldn't happen.")
                
        return excelColNameDict

    excelColNameDict = keyCounter(mainDict)
    printable_list, console_results = formFunc(mainDict, excelColNameDict)
    return printable_list, console_results, excelColNameDict
    

def mainMenu():
    global patientNotes
    patientNotes = """ """
    
    dataDict = {}
    
    menuDict = jsonLoader("menuData")
    
    doneTests = {}
    for testNum in menuDict.keys():
        doneTests[int(testNum)] = "(-)"

    while True:
        try:          
            title = "PsiNorm Testler"
            
            textDump = menuAutoCreate(menuDict)
            
            textDump = textDump.format(doneTests = doneTests)
            
            guiTestChoose(title, menuDict, textDump)
            
            if whatToDo == "doTest":
                
                print("\nLÜTFEN DİKKAT: Test içerisinde uygulamadığınız veya olmayan değerleri 999 olarak giriniz.")
                
                testDataDict = menuDict[testNumToDo]["testDataDict"]
                
                if testDataDict == "testWechslerDataDict":
                    data = testWechsler()
                
                elif testDataDict == "testCctDataDict":
                    data = testCct()
                    
                else: 
                    data = funcTestTemplate(testDataDict)
                
                doneTests[int(testNumToDo)] = "(+)"
                
                dataDict.update({int(testNumToDo) : data})
                continue
                #using the number, calls the function user wanted, adds that to the done tests
                
            elif whatToDo == "getPatientNotes":
                guiPatientNotes("Hasta notları", patientNotes)
            
            elif whatToDo == "saveProgram":
                try: 
                    for testNum in doneTests.keys():
                        if doneTests[int(testNum)] != "(-)":
                                            
                            test_name = dataDict[testNum][0]
                            printable_list = dataDict[testNum][1]
                            console_results = dataDict[testNum][2]
                            #using the numbers from the used tests list, it creates data to feed into the writers
                    
                            if settings("output_excel"):
                                excelWriter(settings("excel_path"), testNum, printable_list, "testData", None)
                                
                            if settings("output_csv"):
                                csvWriter(settings("csv_path"), test_name, printable_list)
                                #writes the printable_list in a CSV file
                            if settings("output_txt"):
                                txtWrite(settings("txt_path"), console_results)
                                #writes the console results on the txt file
                            
                    if settings("output_excel"):
                        excelWriter(settings("excel_path"), None, doneTests, "masterData", None)
                    
                    if settings("output_txt"):
                        done_tests_string = "\n----------------------------------------\nYapılan testler listesi: "
                        
                        for testNum in doneTests.keys():
                            if doneTests[int(testNum)] != "(-)":
                                done_tests_string += "\n" + str(testNum)+ ") "
                                done_tests_string += menuDict[str(testNum)]["testName"]
                
                        done_tests_string += "\n----------------------------------------\n"
                        
                        done_tests_string += "Hastaya dair uygulayıcı notları: \n"
                        
                        done_tests_string += patientNotes
                        
                        done_tests_string += "\n----------------------------------------\n"
                        
                        txtWrite(settings("txt_path"), done_tests_string)

                    if settings("output_txt"):
                        references = jsonLoader("info_data")["references_data"]
                        txtWrite(settings("txt_path"), references)
                        
                    print("\nSONUÇLAR KULLANICIYA AİT 'BELGELERİM' VEYA 'DOCUMENTS' KLASÖRÜ İÇERİSİNDE 'PsiNorm' İÇİNE KAYDEDİLMEKTEDİR.")
                    print("\nDOSYA İSİMLERİNDEKİ BOŞLUKLAR TEKNİK SEBEPLERDEN ÖTÜRÜ OTOMATİK OLARAK \"_\" İLE DEĞİŞTİRİLMEKTEDİR.")
                    print("\nŞu ana kadar yapılanlar kaydedildi.\nProgram baştan başlatılıyor.")    
                    break
                    #this is the exit command
                except SystemExit:
                    raise
                except:
                    print("Kaydedilirken bir hata oluştu.")
                    wait(2)
                    raise
                    
            elif whatToDo == "getPatientInfoForms":
                mainDict = jsonLoader("form_data")
                
                form_list = []
                for i in mainDict["formNames"].keys():
                    form_list.append(i)
                    
                to_save_dict = {}
                                    
                while True:
                    print("|>===================================================================================================<|\n")
                    for i in mainDict["formNames"].keys():    
                        print(i + ") " + mainDict["formNames"][i])
                    print("\n|>===================================================================================================<|\n")
                    print("Tüm girilenleri kaydetip nöropsikolojik testlere dönmek için '888' giriniz. ")
                    userInput = numInput("Lütfen bir seçim yapınız: ")
                    if userInput in form_list:
                        printable_list, console_results, excelColNameDict = formMain(mainDict[userInput])
                        to_save_dict[userInput] = [printable_list, console_results, mainDict[userInput]["excelSheetName"], excelColNameDict]
                        continue
                    elif userInput == "888":
                        print("Nöropsikolojik testlere dönülüyor.\n")
                        if len(to_save_dict.keys()) > 0:
                            for i in to_save_dict.keys():
                                printable_list = to_save_dict[i][0]
                                console_results = to_save_dict[i][1]
                                excelSheetName = to_save_dict[i][2]
                                excelColNameDict = to_save_dict[i][3]
                                
                                if settings("output_excel"):
                                    excelWriter(settings("excel_path"), None, printable_list, excelSheetName, excelColNameDict)
                                   
                                if settings("output_csv"):
                                    csvWriter(settings("csv_path"), test_name, printable_list)
                                    #writes the printable_list in a CSV file
                                if settings("output_txt"):
                                    console_results = "\n===========================================\n" + console_results
                                    console_results = excelSheetName + "\n" + console_results
                                    console_results = console_results + "\n===========================================\n"
                                    txtWrite(settings("txt_path"), console_results)
                            #writes the console results on the txt file
                                
                        break
                        
                    else:
                        print("Lütfen varolan seçeneklerden birini seçiniz. ")
                        continue
                continue
                    
            elif whatToDo == "exitPatient":
                break
            
            else:
                criticalError("Kritik hata!", "mainMenu() - Bir şeyler çok yanlış gitti, lütfen yazılımcıya ulaşınız.", True)
        
        except SystemExit:
            raise
        except:
            raise
            print("\nLütfen sadece sayı giriniz.")
            continue
            #DEATH PROTECTION TOME

def mainStartup(): #the mainStartup function
    print("""
================================================
 PsiNorm Persentil Hesaplayıcı - 1.9.1
================================================
""")
    timeGlobalization()
    guiPatientInfo()
    
    try:
        if None in [patientAdmin, patientID, patientName, patientAge, patientSex, patientEdu]:
            print("Kayıt iptal edilmiş. Geri dönülüyor. ")
            return
        else:
            mainMenu()
    except NameError:
        if settings("debug"):
            raise
        print("Kayıt iptal edilmiş. Geri dönülüyor. ")
        return
    except:
        raise
        

def calcZscore(result, mean, sd):
    #finds out which SD interval patient result is in
    if mean != None and sd != None:
        if sd == 0:
            sd = 0.00000001 #protects the program from failing
        zScore = ((result - mean) / sd)
        
        zScore = float("%.2f" % zScore)
    else:
        zScore = None
    return zScore

def calcPercentile(zScore):
    #gets the Z score and calculate the percentile
    import math

    def percentile(zScore):
        return .5 * (math.erf(zScore / 2 ** .5) + 1)

    if zScore != None:
        try:
            perc = str("%.2f" % (100 * float(percentile(zScore))))
            
        except:
            print("Persentil hesaplanırken bir hata oluştu.")
            perc = None
            pass
    else:
        perc = None

    return perc
    
def outputPrintlist(result, zScore, zScoreVerbalResult, perc):
    #puts all the lists in their proper, more manageable order to print in CSV
    printableList = []
    
    if not result in ["999", 999, 999.0, "999.0"]:
        printableList.append(result)
        if zScore != None:
            printableList.append(zScore)
            printableList.append(perc)
            printableList.append(zScoreVerbalResult)
        else:
            printableList.append("N/A")
            printableList.append("N/A")
            printableList.append("N/A")
    else:
        printableList.append("N/A")
        printableList.append("N/A")
        printableList.append("N/A")
        printableList.append("N/A")
        
    return printableList

def outputConsoleResults(result, zScore, zScoreVerbalResult, perc):
    #gets the results ready to print onto the screen    
    
    if (result != 999.00) and (zScore != None):
        consoleResult = ("Hastanın puanı: " + str(result) + " - " + 
        str(zScoreVerbalResult) + " Z skoru: " + str(zScore) + " - Persentil: " + str(perc))
        
    elif (result != 999.00) and (zScore == None):
        consoleResult = ("Hastanın puanı: " + str(result) + " - Bu parametreye ait norm verisi yoktur.")
    
    else:
        consoleResult = "Bu basamak uygulanmamış veya uygulanamamıştır."

    return consoleResult


def zScoreToVerbal(zScore): 
    #assumes Z scores get better as it goes up, multiply with "-1" if otherwise before using this function
    zScoreVerbalResult = None
    
    if settings("zInterval") == "0-1":
        cutOffList = [-1, -2, -3]    
    if settings("zInterval") == "0-1.5":
        cutOffList = [-1.5, -2, -3] 
        
    if zScore != None:
        if cutOffList[0] <= zScore:
            zScoreVerbalResult = "Normal."
        elif cutOffList[1] <= zScore < cutOffList[0]:
            zScoreVerbalResult = "Hafif derecede bozulma."
        elif cutOffList[2] <= zScore < cutOffList[1]:
            zScoreVerbalResult = "Orta derecede bozulma."
        elif  zScore < cutOffList[2]:
            zScoreVerbalResult = "Ağır derecede bozulma." 
        else:
            zScoreVerbalResult = "KRİTİK HATA, LÜTFEN YAZILIMCI İLE İLETİŞİME GEÇİNİZ."
    else:
        zScoreVerbalResult = "Bu grup için norm değeri bulunmamaktadır."
        
    return zScoreVerbalResult


def zScoreInterpreter(paraNum, zScore, zScoreLegend):
    #z_score_legend = {"all":"less"} or {"all":"more"} means the all the Z scores
    # get better as they go lower or higher
    #z_score_legend = {"0":"less", "1":"more", "2":"less"} means 0 and 2nd Z scores
    # get better as they go low and 1 gets better when high
    try:
        if zScoreLegend["all"] == "more":          
            tempZScore = zScore
                    
        elif zScoreLegend["all"] == "less":
            if zScore != None:
                tempZScore = -1 * zScore
            else:
                tempZScore = zScore
            
    except:
            if zScoreLegend[paraNum] == "more":
                tempZScore = zScore
                
            elif zScoreLegend[paraNum] == "less":
                if zScore != None:
                    tempZScore = -1 * zScore
                else:
                    tempZScore = zScore
                
    
    perc = calcPercentile(tempZScore)
            
    zScoreVerbalResult = zScoreToVerbal(tempZScore)
    
    
    return perc, zScoreVerbalResult


def testWechsler():
    testDataDict = jsonLoader("testWechslerDataDict")
    
    resultDict = guiTest(testDataDict)
    
    tempResultDict = resultDict
    
    del globals()["resultDict"]
    
    resultDict = tempResultDict

    resultList = []
    
    
    from natsort import natsorted as nt
    
    resultDictKeys = nt(resultDict.keys(), key=lambda y: y.lower())
    
    for key in resultDictKeys:
        resultList.append(resultDict[key])

    try:
        printableList = []
        for y in range(len(resultList)):
            printableList.append(resultList[y])
            
            
            valid_items_dict = {
                    0: 0, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0
                    }
               
            #For every scaled score, list of raw scores, min and max values. 
        raw_to_scaled_dict = testDataDict["raw_to_scaled_dict"] 
        
        # On the left are the scaled values, and on the right, are the raw
        # While inefficient, program goes through every scaled dict entry and checks
        # if the patient score is equal/between given value(s).
        
        scaled_dict = {}
        
        for raw_score_index in range(len(printableList)):
            raw_score = printableList[raw_score_index]
            for dictKey in range(20):
                to_check_list = raw_to_scaled_dict[str(dictKey)][raw_score_index] 
                if len(to_check_list) == 2:
                    if to_check_list[0] <= raw_score <= to_check_list[1]:
                        scaled_dict[raw_score_index] = dictKey
                        valid_items_dict[raw_score_index] = 1
                        
                if len(to_check_list) == 1:
                    if to_check_list[0] != None:
                        if to_check_list[0] == raw_score:
                            scaled_dict[raw_score_index] = dictKey
                            valid_items_dict[raw_score_index] = 1
                            
        
        
        for i in scaled_dict.keys():
            printableList[i] = scaled_dict[i]
            
        verb_items_list = [0, 1, 2, 3, 4, 5]
        perf_items_list = [6, 7, 8, 9, 10]
        verb_score = 0
        perf_score = 0
        verb_items = 0
        perf_items = 0
                
        for i in verb_items_list:
            if valid_items_dict[i] == 1:
                verb_score += printableList[i]
                verb_items += 1
        
        if verb_items != 0:
            verb_score = (verb_score) * (6 / verb_items)
        else:
            verb_score = 0
        
        for i in perf_items_list:
            if valid_items_dict[i] == 1:
                perf_score += printableList[i]
                perf_items += 1
                
        if perf_items != 0:
            perf_score = (perf_score) * (5 / perf_items)
        else:
            perf_score = 0
        
        total_score = verb_score + perf_score
        
        IQ_dict_list =  testDataDict["IQ_dict_list"] 
            
        # 1. Check if age is between two values
        # 2. If true, one by one, check if verb/perf/total is between given values
        # 3. For each item, find and return where they lie in the population.
        
        try:
            result_values = ["N/A", "N/A", "N/A"]
            testing_values = [verb_score, perf_score, total_score]
            #result_values = [None, None, None]
            for IQ_dict in IQ_dict_list:
                if IQ_dict["age"][0] <= patientAge <= IQ_dict["age"][1]:
                    for i in range(3):
                        if IQ_dict["120"][i] <= testing_values[i]:
                            result_values[i] = "Üstün zeka - (120 ve üzeri IQ) "
                        elif IQ_dict["110"][i] <= testing_values[i]:
                            result_values[i] = "Parlak zeka - (110-119 IQ) "
                        elif IQ_dict["90"][i] <= testing_values[i]:
                            result_values[i] = "Normal zeka - (90-109 IQ) "
                        elif IQ_dict["80"][i] <= testing_values[i]:
                            result_values[i] = "Donuk normal zeka - (80-89 IQ)"
                        elif IQ_dict["70"][i] <= testing_values[i]:
                            result_values[i] = "Sınırda zeka geriliği - (70-79 IQ)"
                        elif testing_values[i] < IQ_dict["70"][i]:
                            result_values[i] = "Mental retardasyon - (69 ve altı IQ)"
                            
                        else:
                            print("Bir şeyler yanlış gitti. Lütfen programcı ile bağlantıya geçiniz. ")
                            result_values = ["N/A", "N/A", "N/A"]
                            
    
        except:
             raise
                
                
        for i in range(3):
            printableList.append(testing_values[i])
            printableList.append(result_values[i])
            
        
        testName = testDataDict["testName"] + ".csv" #declares name of the CSV file to save the data in
        
        outputconsoleResults = []
        for i in range(len(resultList)):
            outputconsoleResults.append("Hastanın standart puanı: " + str(printableList[i]))
       
        consoleResults = "==================================\nWechsler zeka testinin sonuçları: "
        
        
        for i in range(testDataDict["paraNum"]):
            consoleResults += ("\n" + testDataDict[str(i)] + str(outputconsoleResults[i]))
            
        consoleResults += "\nSözel standart puan: " + str(verb_score) + " - " + str(result_values[0])
        consoleResults += "\nPerformans standart puan: " + str(perf_score) + " - " + str(result_values[1])
        consoleResults += "\nToplam standart puan: " + str(total_score) + " - " + str(result_values[2])
        

        consoleResults = consoleResults + ("\n==================================")
        
        guiSimpleTextDump(testDataDict["testName"], consoleResults)
        return [testName, printableList, consoleResults]
        
    except:
        print("Wechsler zeka testini değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return
          

def testCct(): #Color trails joiner
    while True:
        try:
            dataCct = ['CCT_data.csv', "", ""]
            
            dataCct1 = funcTestTemplate("testCct1DataDict")
            dataCct2 = funcTestTemplate("testCct2DataDict")
            
            dataCct[1] = dataCct1[1] + dataCct2[1]
            dataCct[2] = dataCct1[2] + dataCct2[2]
            
        except:
            print("CCT'yi değerlendirirken bir hata oluştu, program kapatılacak.")
            continue
        
        else:
            break
    
    return dataCct

def cutOffInterpreter(result, paraNum, testDataDict):
    normExists = False
    for i in testDataDict["cutOffList"]:
        if i["sex"] == patientSex and (i["ageLow"] <= patientAge <= i["ageHigh"]) and (i["eduLow"] <= patientEdu <= i["eduHigh"]):
            correctNorm = i
            normExists = True
            break
        
    #Find the correct cutoff values by iterating through every entry in JSON
    if normExists: 
        if correctNorm[paraNum]["parameterNormExists"]:
            if not result in [999, 999.0, "999", "999.0"]:
                cutOffGroupCutoffList = correctNorm[paraNum]["cutOffGroupCutoffList"]
                #[cutOff1, cutOff2, cutOff3...]
                cutOffGroupNameList = correctNorm[paraNum]["cutOffGroupNameList"]
                #[groupName1, groupName2, groupName3...]
    
                
                """
                Example data structure:
                {
                "sex": "Male",
                "eduLow": 0,
                "eduHigh": 4,
                "ageLow": 0,
                "ageHigh": 69,
                "0": {
                "parameterNormExists": true,
                "cutOffGroupCutoffList": [1, 2, 3],
                "cutOffGroupNameList": ["Low", "Medium", "High", "Very High"]
                },
                "1": {
                "parameterNormExists": true,
                "cutOffGroupCutoffList": [5, 10, 20, 50],
                "cutOffGroupNameList": ["Little", "Medium", "Big", "Very Big", "Huge"]
                }
                "2": {
                "parameterNormExists": false
                }
                }
                """
                
                # patientResult <= cutOff1 = groupName1 
                # cutOff1 < patientResult <= cutOff2 = groupName2
                # cutOff2 < patientResult <= cutOff3 = groupName3
                # cutOff3 < patientResult = groupName4                        
        
                for cutOffGroupNum in range(len(cutOffGroupNameList)):
                    
                    if cutOffGroupNum == 0:
                        if result <= cutOffGroupCutoffList[cutOffGroupNum]:
                            verbalResult = cutOffGroupNameList[cutOffGroupNum]
                        
                    elif cutOffGroupNum < len(cutOffGroupCutoffList):
                        if cutOffGroupCutoffList[cutOffGroupNum-1] < result <= cutOffGroupCutoffList[cutOffGroupNum]:
                            verbalResult = cutOffGroupNameList[cutOffGroupNum]
                        
                    elif cutOffGroupNum == len(cutOffGroupCutoffList):
                        if cutOffGroupCutoffList[cutOffGroupNum-1] < result:
                            verbalResult = cutOffGroupNameList[cutOffGroupNum]
                #iterates through all the cutoff-group couples, adding appropriate verbal results when available
            else:
                verbalResult = "Bu basamak uygulanmamış veya uygulanamamıştır."
        else:
            verbalResult = "Bu parametreye ait norm verisi yoktur."
            #Adds a N/A entry for every missing norm value 
    
    else:            
        verbalResult = "Bu parametreye ait norm verisi yoktur."
    
    if result in [999, 999.0, "999", "999.0"]:
        result = "N/A"
        printable = [result, verbalResult]
        console = "Bu basamak uygulanmamış veya uygulanamamıştır."
    
    else:
        printable = [result, verbalResult]
        console = str(result) + ", " + verbalResult
    
                    
    return printable, console
                    

def funcTestTemplate(JSONname):#Test Name
    testDataDict = jsonLoader(JSONname) 
    #Load test data from JSON file
    
    paraTypeDict = testDataDict["paraTypeDict"]
    
    resultDict = guiTest(testDataDict)
    
    tempResultDict = resultDict
    
    del globals()["resultDict"]
    
    resultDict = tempResultDict

    try:
        if testDataDict["testType"] == "zScore":
            #if test type is zScore calculating type      
            
            normExists = False
            for norm in testDataDict["normList"]:
                if norm["sex"] == patientSex and (norm["ageLow"] <= patientAge <= norm["ageHigh"]) and (norm["eduLow"] <= patientEdu <= norm["eduHigh"]):
                    correctNorm = norm
                    normExists = True
                    break
            #Find the correct norm values by iterating through every entry in JSON    
            
            
            meanDict = {}
            sdDict = {}
            
            if normExists:
                for paraNum in range(testDataDict["paraNum"]):
                    paraNum = str(paraNum)
                    meanDict[paraNum] = (correctNorm[paraNum][0])
                    sdDict[paraNum] = (correctNorm[paraNum][1])
            else:
                for paraNum in range(testDataDict["paraNum"]):
                    paraNum = str(paraNum)
                    meanDict[paraNum] = None
                    sdDict[paraNum] = None
                
            #Create appropriate mean and standard deviation lists for further calculations
                    
            
            printableDict = {}
            consoleDict = {}
            for paraNum in range(testDataDict["paraNum"]):
                paraNum = str(paraNum)
                
                result = resultDict[paraNum]
                mean = meanDict[paraNum]
                sd = sdDict[paraNum]
                
                if paraTypeDict[paraNum] == "intType":
                    zScore= calcZscore(result, mean, sd)
                    
                    perc, zScoreVerbalResult = zScoreInterpreter(paraNum, zScore, testDataDict["zScoreLegend"])
                    
                    printableDict[paraNum] = outputPrintlist(result, zScore, zScoreVerbalResult, perc)
                    consoleDict[paraNum] = outputConsoleResults(result, zScore, zScoreVerbalResult, perc)
            
                else: #strType
                    if result == None:
                        result = " "
                    printableDict[paraNum] = [result]
                    consoleDict[paraNum] = "Hastanın değerlendirme sonucu: " + str(result)
                    
        elif testDataDict["testType"] == "cutOff":
            #if the test is a simple cutoff type
      
            printableDict = {}
            consoleDict = {}
            
            for paraNum in range(testDataDict["paraNum"]):
                paraNum = str(paraNum)
                
                if paraTypeDict[paraNum] == "intType":
                    result = resultDict[paraNum]
                    printableDict[paraNum], consoleDict[paraNum] = cutOffInterpreter(result, paraNum, testDataDict)
                
                else: #strType
                    if result == None:
                        result = " "
                    printableDict[paraNum] = [result]
                    consoleDict[paraNum] = "Hastanın değerlendirme sonucu: " + str(result)


        printableList = []
        for paraNum in printableDict.keys():
            printableList += printableDict[paraNum]
        #Creates a dump for excel/CSV
        
        consoleResults = "==================================\n" + testDataDict["testName"]
        for paraNum in consoleDict.keys():
            consoleResults += "\n" + testDataDict[paraNum] + consoleDict[paraNum]      
        consoleResults += ("\n==================================")
        #Creates a dump for console/txt file

        testName = testDataDict["testName"] + ".csv" #declares name of the CSV file to save the data in
        
        guiSimpleTextDump(testDataDict["testName"], consoleResults)
        return [testName, printableList, consoleResults]
            
    except:
        print(testDataDict["testName"] + " değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return
        #saves the program from fiery death
        

def progStructure():
    first_run = True
    if first_run:
        import os
        import sys
        os.system("mode con: cols=160 lines=50")
        print("""
================================================
 PsiNorm Persentil Hesaplayıcı - 1.9.1
================================================
 Copyright (C) 2017 Bilal Bahadır Akbulut & Yavuz Ayhan
 
 Bu programın KESİNLİKLE HİÇBİR TEMİNATI YOKTUR; 
 ayrıntılar için programın içinde bulunduğu klasördeki lisans dosyalarına başvurunuz.

 Bu bir özgür yazılımdır, ve bazı koşullar altında yeniden dağıtmakta serbestsiniz;
 ayrıntılar için programın içinde bulunduğu klasördeki lisans dosyalarına başvurunuz.
================================================
          BU EKRANI KAPATMAYINIZ
================================================
""")
        if settings('first_ever_run'):
            if guiAgreeTerms():
                configFilePath = getUserDocumentsPsiPath() + "settings.ini"
                from configobj import ConfigObj
                config = ConfigObj(configFilePath, encoding='UTF8')
                config['General']['first_ever_run'] = "False"
                config.write()
                
                
        guiStartupMenu()
        mainStartup()
        first_run = False
    
    while True:
        guiStartupMenu()
        mainStartup()
        
    sys.exit()
        #informs the user data has been saved then restarts
 
    
    
import logging
try:
    timeGlobalization()
    progStructure()
    
except SystemExit:
    raise
    
except Exception as e:
    crashLogFolder = getUserDocumentsPsiPath() + "/Crash Logs/"
    ensure_dir(crashLogFolder)
    
    
    crashlogPath = crashLogFolder + ("psinormcrash-" + date + "-" + stime + ".log")
    
    guiSimplePopup("Kritik hata!", "Bir hata oluştu ve program kapatılacak. \n\n'Documents' veya 'Belgelerim' içerisinde, 'PsiNorm/Crash Logs' klasörünün içerisinde bulunan 'psinormcrash" + date + "-" + stime + ".log' dosyasını b.bahadirakbulut@gmail.com'a yönlendiriniz.")
    logging.basicConfig(level=logging.DEBUG, filename= crashlogPath, filemode='w')
    
    howtosenderrorlog = """
Bu dosyayı email ile b.bahadirakbulut@gmail.com adresine gönderiniz.

--------------------------------------------------------------------
PsiNorm - Versiyon 1.9.1
--------------------------------------------------------------------

    """
    
    logging.exception(howtosenderrorlog)   
    logging.shutdown()
