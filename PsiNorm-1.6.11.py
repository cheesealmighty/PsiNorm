# -*- coding: utf-8 -*-
#!/usr/bin/env python3
"""
PsiNorm Persentil Hesaplayıcı - Versiyon 1.6.11
Copyright (C) 2017 Bilal Bahadır Akbulut & Yavuz Ayhan

Bu program özgür yazılımdır: Özgür Yazılım Vakfı tarafından yayımlanan
 GNU Genel Kamu Lisansı’nın sürüm 3 ya da (isteğinize bağlı olarak) 
 daha sonraki sürümlerinin hükümleri altında yeniden dağıtabilir ve/veya
 değiştirebilirsiniz.

Bu program, yararlı olması umuduyla dağıtılmış olup, programın BİR TEMİNATI YOKTUR;
 TİCARETİNİN YAPILABİLİRLİĞİNE VE ÖZEL BİR AMAÇ İÇİN UYGUNLUĞUNA dair bir teminat da vermez.
 Ayrıntılar için GNU Genel Kamu Lisansı’na göz atınız.

Bu programla birlikte GNU Genel Kamu Lisansı’nın bir kopyasını elde etmiş olmanız gerekir. 
Eğer elinize ulaşmadıysa <http://www.gnu.org/licenses/> adresine bakınız.

--------------------------------------------------------------
Dr. Bilal Bahadır Akbulut
İletişim adresi: b.bahadirakbulut@gmail.com
------------------------------------
PsiNorm Percentile Calculator - Version 1.6.11
Copyright (C) 2017 Bilal Bahadır Akbulut & Yavuz Ayhan

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <http://www.gnu.org/licenses/>.

-------------------------------------------------------------

Dr. Bilal Bahadır Akbulut
contact adress: b.bahadirakbulut@gmail.com

"""

def guiSettings():
    import tkinter as tk
    titleText = "PsiNorm Persentil Hesaplayıcı - 1.6.11 - Seçenekler "

    class MyFirstGUI:
        def __init__(self, master):
            self.master = master
            master.title("PsiNorm Persentil Hesaplayıcı")
            
            self.label = tk.Label(master, text=titleText, relief=tk.GROOVE, font= (None, 16))
            self.label.grid(row=0, column=0, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=5)    

            self.label = tk.Label(master, text="Kaydedilecek klasör: ", relief=tk.GROOVE)
            self.label.grid(row=1, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    
         
            self.labelFolderName = tk.Label(master, relief=tk.GROOVE, 
                                  text= "Results (Varsayılan)" if (settings("folder_name")=="default") else settings("folder_name"))
            self.labelFolderName.grid(row=1, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)   
            
            self.entryFolderName = tk.Entry(master, width=50)
            self.entryFolderName.grid(row=1, column=2, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)

            self.buttonFolderNameSave = tk.Button(master, text="Kaydet", command=self.folder_name)
            self.buttonFolderNameSave.grid(row=1, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 

            self.buttonFolderNameDefault = tk.Button(master, text="Varsayılan", command=self.buttonFolderNameDefaultFunc)
            self.buttonFolderNameDefault.grid(row=1, column=4, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
    
            self.label = tk.Label(master, text="CSV raporu çıkar: ", relief=tk.GROOVE)
            self.label.grid(row=2, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    
         
            self.button1T = tk.Button(master, text="Evet", command=self.button1, state = (tk.DISABLED if settings("output_csv") else tk.NORMAL))
            self.button1T.grid(row=2, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
            
            self.button1F = tk.Button(master, text="Hayır", command=self.button1, state = (tk.NORMAL if settings("output_csv") else tk.DISABLED))
            self.button1F.grid(row=2, column=2, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)
            
            self.label = tk.Label(master, text="Text raporu çıkar: ", relief=tk.GROOVE)
            self.label.grid(row=3, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    
         
            self.button2T = tk.Button(master, text="Evet", command=self.button2, state = (tk.DISABLED if settings("output_txt") else tk.NORMAL))
            self.button2T.grid(row=3, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
            
            self.button2F = tk.Button(master, text="Hayır", command=self.button2, state = (tk.NORMAL if settings("output_txt") else tk.DISABLED))
            self.button2F.grid(row=3, column=2, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)

            self.label = tk.Label(master, text="Excel raporu çıkar: ", relief=tk.GROOVE)
            self.label.grid(row=4, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    
         
            self.button3T = tk.Button(master, text="Evet", command=self.button3, state = (tk.DISABLED if settings("output_excel") else tk.NORMAL))
            self.button3T.grid(row=4, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
            
            self.button3F = tk.Button(master, text="Hayır", command=self.button3, state = (tk.NORMAL if settings("output_excel") else tk.DISABLED))
            self.button3F.grid(row=4, column=2, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)

            self.label = tk.Label(master, text="Excel dosyasının ismi: ", relief=tk.GROOVE)
            self.label.grid(row=5, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    
            
            self.labelExcelName = tk.Label(master, relief=tk.GROOVE, 
                                  text= "Results (Varsayılan)" if (settings("excel_name")=="default") else settings("excel_name"))
            self.labelExcelName.grid(row=5, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)   
            
            self.entryExcelName = tk.Entry(master, width=50)
            self.entryExcelName.grid(row=5, column=2, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)

            self.buttonExcelNameSave = tk.Button(master, text="Kaydet", command=self.excel_name)
            self.buttonExcelNameSave.grid(row=5, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 

            self.buttonExcelNameDefault = tk.Button(master, text="Varsayılan", command=self.buttonExcelNameDefaultFunc)
            self.buttonExcelNameDefault.grid(row=5, column=4, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)
            
            self.label = tk.Label(master, text="Rapor dosyası isim formatı: ", relief=tk.GROOVE)
            self.label.grid(row=6, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    
         
            self.button4Date = tk.Button(master, text="Önce tarih", command=self.button4DateFunc, state = (tk.DISABLED if settings("textFileNameFormat")=="firstDate" else tk.NORMAL))
            self.button4Date.grid(row=6, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
            
            self.button4Admin = tk.Button(master, text="Önce uygulayıcı", command=self.button4AdminFunc, state = (tk.DISABLED if settings("textFileNameFormat")=="firstAdmin" else tk.NORMAL))
            self.button4Admin.grid(row=6, column=2, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)

            self.button4ID = tk.Button(master, text="Önce hasta kodu", command=self.button4IDFunc, state = (tk.DISABLED if settings("textFileNameFormat")=="firstID" else tk.NORMAL))
            self.button4ID.grid(row=6, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)
            
            self.buttonexit = tk.Button(master, text="Çıkış", command=self.exitFunc)
            self.buttonexit.grid(row=7, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=2)
 
        def exitFunc(self):
            root.destroy()
            
        def folder_name(self):
            new_setting = self.entryFolderName.get()
            if checkGuiInputString(new_setting):
                new_setting = new_setting.replace(" ", "_")
                changeSettingsString("folder_name", new_setting)
                self.labelFolderName.config(text= "Results (Varsayılan)" if (settings("folder_name")=="default") else settings("folder_name"))
            else:
                tk.messagebox.showerror("Hata!", "Lütfen harf, sayı, '-', '_', '(' ve ')' dışında giriş yapmayınız.")
            
        def buttonFolderNameDefaultFunc(self):
            changeSettingsString("folder_name", "default")
            self.labelFolderName.config(text= "Results (Varsayılan)" if (settings("folder_name")=="default") else settings("folder_name"))
            
        def button1(self):
            changeSettingsBool("output_csv")    
            self.button1T.config(state= (tk.DISABLED if settings("output_csv") else tk.NORMAL))
            self.button1F.config(state= (tk.NORMAL if settings("output_csv") else tk.DISABLED))
            
        def button2(self):
            changeSettingsBool("output_txt")     
            self.button2T.config(state= (tk.DISABLED if settings("output_txt") else tk.NORMAL))
            self.button2F.config(state= (tk.NORMAL if settings("output_txt") else tk.DISABLED))
            
        def button3(self):
            changeSettingsBool("output_excel")    
            self.button3T.config(state= (tk.DISABLED if settings("output_excel") else tk.NORMAL))
            self.button3F.config(state= (tk.NORMAL if settings("output_excel") else tk.DISABLED))   
            
        def excel_name(self):
            new_setting = self.entryExcelName.get()
            if checkGuiInputString(new_setting):
                new_setting = new_setting.replace(" ", "_")
                changeSettingsString("excel_name", new_setting)
                self.labelExcelName.config(text= "Results (Varsayılan)" if (settings("excel_name")=="default") else settings("excel_name"))
            else:
                tk.messagebox.showerror("Hata!", "Lütfen harf, sayı, '-', '_', '(' ve ')' dışında giriş yapmayınız.")

        def buttonExcelNameDefaultFunc(self):
            changeSettingsString("excel_name", "default")
            self.labelExcelName.config(text= "Results (Varsayılan)" if (settings("excel_name")=="default") else settings("excel_name"))
            
        def button4DateFunc(self):
            changeSettingsString("textFileNameFormat", "firstDate")
            self.button4Date.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstDate" else tk.NORMAL))
            self.button4Admin.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstAdmin" else tk.NORMAL))
            self.button4ID.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstID" else tk.NORMAL))
            
        def button4AdminFunc(self):
            changeSettingsString("textFileNameFormat", "firstAdmin")
            self.button4Date.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstDate" else tk.NORMAL))
            self.button4Admin.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstAdmin" else tk.NORMAL))
            self.button4ID.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstID" else tk.NORMAL))
            
        def button4IDFunc(self):
            changeSettingsString("textFileNameFormat", "firstID")
            self.button4Date.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstDate" else tk.NORMAL))
            self.button4Admin.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstAdmin" else tk.NORMAL))
            self.button4ID.config(state = (tk.DISABLED if settings("textFileNameFormat")=="firstID" else tk.NORMAL))
            
            
            
    root = tk.Tk()
    tk.Grid.rowconfigure(root, [0, 1, 2, 3, 4, 5, 6, 7], weight=1)
    tk.Grid.columnconfigure(root, [0, 1, 2, 3, 4], weight=1)
    root.config(borderwidth=10, relief=tk.GROOVE)
    
    my_gui = MyFirstGUI(root)
    root.mainloop() 

def guiPatientInfo():
    import tkinter as tk
    titleText = "PsiNorm Persentil Hesaplayıcı - 1.6.11 - Hasta Bilgileri" 
    patient_sex = None

    class patientInfo:
        def __init__(self, master):
            self.master = master
            master.title("PsiNorm Persentil Hesaplayıcı")
            
            self.label = tk.Label(master, text=titleText, relief=tk.GROOVE, font= (None, 16))
            self.label.grid(row=0, column=0, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=5)    

            self.label = tk.Label(master, text="Testi uygulayan kişi: ", relief=tk.GROOVE)
            self.label.grid(row=1, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    

            self.entryPatientAdmin = tk.Entry(master, width=50)
            self.entryPatientAdmin.grid(row=1, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=2)

            self.label1 = tk.Label(master, text="*", relief=tk.GROOVE)
            self.label1.grid(row=1, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    


            self.label = tk.Label(master, text="Hastanın kodu: ", relief=tk.GROOVE)
            self.label.grid(row=2, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    

            self.entryPatientID = tk.Entry(master, width=50)
            self.entryPatientID.grid(row=2, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=2)

            self.label2 = tk.Label(master, text="*", relief=tk.GROOVE)
            self.label2.grid(row=2, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    


            self.label = tk.Label(master, text="Hastanın ismi: ", relief=tk.GROOVE)
            self.label.grid(row=3, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    

            self.entryPatientName = tk.Entry(master, width=50)
            self.entryPatientName.grid(row=3, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=2)

            self.label3 = tk.Label(master, text="*", relief=tk.GROOVE)
            self.label3.grid(row=3, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    


            self.label = tk.Label(master, text="Hastanın yaşı: ", relief=tk.GROOVE)
            self.label.grid(row=4, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    

            self.entryPatientAge = tk.Entry(master, width=50)
            self.entryPatientAge.grid(row=4, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=2)

            self.label4 = tk.Label(master, text="*", relief=tk.GROOVE)
            self.label4.grid(row=4, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    


            self.label = tk.Label(master, text="Hastanın cinsiyeti: ", relief=tk.GROOVE)
            self.label.grid(row=5, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    
         
            self.buttonFemale = tk.Button(master, text="Kadın", command=self.buttonFemaleFunc, state = (tk.DISABLED if patient_sex == "Kadın" else tk.NORMAL))
            self.buttonFemale.grid(row=5, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
            
            self.buttonMale = tk.Button(master, text="Erkek", command=self.buttonMaleFunc, state = (tk.DISABLED if patient_sex == "Erkek" else tk.NORMAL))
            self.buttonMale.grid(row=5, column=2, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)
                 
            self.label5 = tk.Label(master, text="*", relief=tk.GROOVE)
            self.label5.grid(row=5, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    
            
            
            self.label = tk.Label(master, text="Hastanın toplam eğitim yılı: ", relief=tk.GROOVE)
            self.label.grid(row=6, column=0, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    

            self.entryPatientEdu = tk.Entry(master, width=50)
            self.entryPatientEdu.grid(row=6, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=2)

            self.label6 = tk.Label(master, text="*", relief=tk.GROOVE)
            self.label6.grid(row=6, column=3, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=1)    


            self.buttonSave = tk.Button(master, text="Kaydet", command=self.saveFunc)
            self.buttonSave.grid(row=7, column=1, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
 
            self.buttonAbort = tk.Button(master, text="İptal", command=self.abortFunc)
            self.buttonAbort.grid(row=7, column=2, ipadx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
        
        def buttonFemaleFunc(self):
            self.buttonFemale.config(state= tk.DISABLED)
            self.buttonMale.config(state= tk.NORMAL)
            
        def buttonMaleFunc(self):
            self.buttonFemale.config(state= tk.NORMAL)
            self.buttonMale.config(state= tk.DISABLED)
            
        def saveFunc(self):
            patient_admin = self.entryPatientAdmin.get()
            patient_ID = self.entryPatientID.get()
            patient_name = self.entryPatientName.get()
            patient_age = self.entryPatientAge.get()
            patient_edu = self.entryPatientEdu.get()
            inputError = False

            
            if not checkGuiInputString(patient_admin):
                self.label1.config(text = "Hata! Lütfen harf, sayı, '-', '_', '(' ve ')' dışında giriş yapmayınız.")
                inputError = True
            else:
                self.label1.config(text = "*")
                
            if not checkGuiInputString(patient_ID):
                self.label2.config(text = "Hata! Lütfen harf, sayı, '-', '_', '(' ve ')' dışında giriş yapmayınız.")
                inputError = True
            else:
                self.label2.config(text = "*")
                
            if not checkGuiInputString(patient_name):
                self.label3.config(text = "Hata! Lütfen harf, sayı, '-', '_', '(' ve ')' dışında giriş yapmayınız.")
                inputError = True
            else:
                self.label3.config(text = "*")
                
            if not checkGuiInputInt(patient_age):
                self.label4.config(text = "Hata! Lütfen sayı dışında giriş yapmayınız.")
                inputError = True
            else:
                patient_age = int(patient_age)
                if 150 > patient_age > 0:
                    self.label4.config(text = "*")
                else:
                    self.label4.config(text = "Hata! Lütfen geçerli bir sayı giriniz.")
                    inputError = True
                
            if str(self.buttonFemale["state"]) == "disabled" or str(self.buttonMale["state"]) == "disabled":
                self.label5.config(text = "*")
                if str(self.buttonFemale["state"]) == "disabled":
                    patient_sex = "Kadın"
                if str(self.buttonMale["state"]) == "disabled":
                    patient_sex = "Erkek"       
            else:
                self.label5.config(text = "Hata! Lütfen bir seçim yapınız.")
                inputError = True
                
            if not checkGuiInputInt(patient_edu):
                self.label6.config(text = "Hata! Lütfen sayı dışında giriş yapmayınız.")
                inputError = True
            else:
                patient_edu = int(patient_edu)
                if 150 > patient_edu >= 0:
                    self.label6.config(text = "*")
                else:
                    self.label6.config(text = "Hata! Lütfen geçerli bir sayı giriniz.")
                    inputError = True

            if inputError:
                return
            else:
                self.on_saving(patient_admin, patient_ID, patient_name, patient_age, patient_sex, patient_edu)
                                
        def abortFunc(self):
            patientInfo.on_closing()
            
        def on_saving(self, temppatient_admin, temppatient_ID, temppatient_name, temppatient_age, temppatient_sex, temppatient_edu):            
            text_dump = ("Testi uygulayan kişi: " + temppatient_admin + "\nHastanın kodu: " + temppatient_ID +
            "\nHastanın ismi: " + temppatient_name + "\nHastanın yaşı: " + str(temppatient_age) +
            "\nHastanın cinsiyeti: " + temppatient_sex + "\nHastanın toplam eğitim yılı: " + str(temppatient_edu)+
            "\nBu kaydı onaylıyor musunuz?")
            if tk.messagebox.askokcancel("Çıkış", text_dump):
                global patient_admin
                patient_admin = temppatient_admin
                global patient_ID
                patient_ID = temppatient_ID
                global patient_name
                patient_name = temppatient_name
                global patient_age
                patient_age = temppatient_age
                global patient_sex
                patient_sex = temppatient_sex
                global patient_edu
                patient_edu = temppatient_edu
                root.destroy()
            
        def on_closing():
            if tk.messagebox.askokcancel("Çıkış", "Bu kaydı iptal etmek istediğinizden emin misiniz?"):
                root.destroy()
    
    root = tk.Tk()
    tk.Grid.rowconfigure(root, [0, 1, 2, 3, 4, 5, 6, 7], weight=1)
    tk.Grid.columnconfigure(root, [0, 1, 2, 3], weight=1)
    root.config(borderwidth=10, relief=tk.GROOVE)
    root.protocol("WM_DELETE_WINDOW", patientInfo.on_closing)
    my_gui = patientInfo(root)
    root.mainloop() 


def getUserDocumentsPsiPath():
    import ctypes.wintypes
    CSIDL_PERSONAL= 5       # My Documents
    SHGFP_TYPE_CURRENT= 0   # Want current, not default value
    
    buf= ctypes.create_unicode_buffer(ctypes.wintypes.MAX_PATH)
    ctypes.windll.shell32.SHGetFolderPathW(0, CSIDL_PERSONAL, 0, SHGFP_TYPE_CURRENT, buf)
    
    documentsPath = buf.value
    psiPath = documentsPath + "/PsiNorm/"
    ensure_dir(psiPath)
    return psiPath


def guiAgreeTerms():
    import tkinter as tk
    from tkinter import messagebox
    from tkinter.scrolledtext import ScrolledText as ScrolledText
    import sys
    
    class MyFirstGUI:
        def __init__(self, master):
            self.master = master
            master.title("Kullanım Koşulları")
            text_dump = jsonLoader("info_data")["agreeTerms_data"]
            self.label = ScrolledText(master, wrap = tk.WORD,width  = 80, height = 30)
            self.label.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
            self.label.insert(tk.INSERT, text_dump)
            self.label.config(state=tk.DISABLED)
            self.label.insert(tk.END, " in ScrolledText")
    
            self.close_button = tk.Button(master, text="Hayır", command=self.areyousure, relief=tk.GROOVE)
            self.close_button.pack(padx=3, pady=3, side=tk.BOTTOM, expand=True, fill=tk.BOTH)
    
            self.greet_button = tk.Button(master, text="Evet", command=self.greet, relief=tk.GROOVE)
            self.greet_button.pack(padx=3, pady=3, side=tk.BOTTOM, expand=True, fill=tk.BOTH)
    
        def greet(self):
            root.destroy()
            
        def areyousure(self):
            answer = messagebox.askyesno("Uyarı!", "Programı kullanabilmek için kabul etmelisiniz. \nBir önceki ekrana dön?")
            if answer:
                return
            else:
                sys.exit()
                
        def on_closing():
            answer = messagebox.askyesno("Uyarı!", "Programı kullanabilmek için kabul etmelisiniz. \nBir önceki ekrana dön?")
            if answer:
                return
            else:
                sys.exit()
    
    root = tk.Tk()
    my_gui = MyFirstGUI(root)
    root.protocol("WM_DELETE_WINDOW", MyFirstGUI.on_closing)
    root.mainloop()
    return True

def guiSimpleTextDump(gui_title, text_dump):
    import tkinter as tk
    from tkinter.scrolledtext import ScrolledText as ScrolledText
    
    class MyFirstGUI:
        def __init__(self, master):
            self.master = master
            master.title(gui_title)
            
            self.label = ScrolledText(master, wrap = tk.WORD,width  = 80, height = 30)
            self.label.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
            self.label.insert(tk.INSERT, text_dump)
            self.label.config(state=tk.DISABLED)
            self.label.insert(tk.END, " in ScrolledText")
       
            self.close_button = tk.Button(master, text="Tamam.", command=self.close)
            self.close_button.pack(side=tk.BOTTOM)    
            
        def close(self):
            root.destroy()
    
    root = tk.Tk()
    my_gui = MyFirstGUI(root)
    root.mainloop()

def guiSimplePopup(gui_title, text_dump):
    import tkinter as tk
    
    class MyFirstGUI:
        def __init__(self, master):
            self.master = master
            master.title(gui_title)
            
            self.label = tk.Label(master, text=text_dump)
            self.label.grid(row=1, column=1, padx=10, pady=10, sticky=tk.N)    
       
            self.greet_button = tk.Button(master, text="Tamam", command=self.close)
            self.greet_button.grid(row=2, column=1, padx=10, pady=10, sticky=tk.S) 
            
        def close(self):
            root.destroy()
    
    root = tk.Tk()
    my_gui = MyFirstGUI(root)
    root.mainloop()

def guiFAQ():
    text_dump = jsonLoader("info_data")["FAQ_data"]
    gui_title = "Sık Sorulan Sorular"
    guiSimpleTextDump(gui_title, text_dump)


def guiAbout():
    text_dump = jsonLoader("info_data")["about_data"]
    gui_title = "PsiNorm Hakkında"
    guiSimpleTextDump(gui_title, text_dump)


def guiReferences():
    text_dump = jsonLoader("info_data")["references_data"]
    gui_title = "Kaynakça"
    guiSimpleTextDump(gui_title, text_dump)

def guiStartupMenu():
    import tkinter as tk
    from tkinter import messagebox
    import sys
    titleText = " PsiNorm Persentil Hesaplayıcı - 1.6.11"
    
    titleSubText= """
 Copyright (C) 2017 Bilal Bahadır Akbulut & Yavuz Ayhan
 
 Bu programın KESİNLİKLE HİÇBİR TEMİNATI YOKTUR; ayrıntılar için programın içinde bulunduğu klasördeki lisans dosyalarına başvurunuz.

 Bu bir özgür yazılımdır ve bazı koşullar altında yeniden dağıtmakta serbestsiniz; ayrıntılar için programın içinde bulunduğu klasördeki lisans dosyalarına başvurunuz.
 """
    titleWarning = "HESAPLAMALAR BU EKRANIN YANINDA AÇILAN KONSOLDAN YAPILMAKTADIR, LÜTFEN KAPATMAYINIZ."

    class MyFirstGUI:
        def __init__(self, master):
            self.master = master
            master.title("PsiNorm Persentil Hesaplayıcı")
            
            psiNormIcon = tk.PhotoImage(file = "psiNormIcon.gif")

            self.label = tk.Label(master,image=psiNormIcon)
            self.label.image = psiNormIcon
            self.label.grid(row=0, column=0, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=6)    
            
            self.label = tk.Label(master, text=titleText, bg="white", relief=tk.GROOVE, font=(None, 16))
            self.label.grid(row=1, column=0, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=6)    
            
            self.label = tk.Label(master, text=titleSubText, bg="white", relief=tk.GROOVE, font=(None, 10))
            self.label.grid(row=2, column=0, padx=5, pady=5, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=6)   
            
            self.label = tk.Label(master, text=titleWarning, bg="white", relief=tk.GROOVE, font=(None, 12))
            self.label.grid(row=3, column=0, padx=5, pady=5, sticky=tk.N+tk.S+tk.E+tk.W, columnspan=6)   
            
            self.greet_button = tk.Button(master, text="Programı çalıştır", command=self.startup)
            self.greet_button.grid(row=4, column=0, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
    
            self.greet_button = tk.Button(master, text="Sık Sorulan Sorular", command=self.guiFAQ)
            self.greet_button.grid(row=4, column=1, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
            
            self.greet_button = tk.Button(master, text="Hakkında", command=self.guiAbout)
            self.greet_button.grid(row=4, column=2, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
            
            self.greet_button = tk.Button(master, text="Kaynakça", command=self.guiReferences)
            self.greet_button.grid(row=4, column=3, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W)
            
            self.greet_button = tk.Button(master, text="Ayarlar", command=self.menuSettings)
            self.greet_button.grid(row=4, column=4, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
            
            self.greet_button = tk.Button(master, text="Çıkış Yap", command=self.programExit)
            self.greet_button.grid(row=4, column=5, padx=10, pady=10, sticky=tk.N+tk.S+tk.E+tk.W) 
    
        
            
        def guiFAQ(self):
            guiFAQ()
            
        def guiAbout(self):
            guiAbout()
            
        def guiReferences(self):
            guiReferences()
            
        def menuSettings(self):
            guiSettings()
        
        def programExit(self):
            MyFirstGUI.on_closing()
            
        def on_closing():
            if messagebox.askokcancel("Çıkış", "Programı kapatmak istediğinizden emin misiniz?"):
                sys.exit()
                
        def startup(self):
            root.destroy()          
    
    root = tk.Tk()
    tk.Grid.rowconfigure(root, [0,1,2,3,4], weight=1)
    tk.Grid.columnconfigure(root, [0,1,2,3,4,5], weight=1)
    root.config(borderwidth=10, relief=tk.GROOVE)
    my_gui = MyFirstGUI(root)
    root.protocol("WM_DELETE_WINDOW", MyFirstGUI.on_closing)
    root.mainloop() 

def jsonLoader(jsonFileName):
    # Currently: "form_data", "info_data"["agreeTerms_data", "FAQ_data", "about_data", "references_data"]
    try:
        import json
        with open(settings(jsonFileName), 'r', encoding="utf8") as fp:
            mainDict = json.load(fp)
        
        fp.close()
        return mainDict
    
    except:
        print("Json file missing.")
        raise

def timeGlobalization():
    from time import strftime
    if "date" in globals():
        globals()["date"] = strftime("%Y-%m-%d")
    else:
        global date
        date = strftime("%Y-%m-%d")
        
    if "time" in globals():
        globals()["time"] = strftime("%H:%M:%S")
    else:    
        global time
        time = strftime("%H:%M:%S")
    
    if "stime" in globals():
        globals()["stime"] = strftime("%H%M")
    else:
        global stime
        stime = strftime("%H%M")
    

def wait(t):
    import time
    while t:
        print(str(t))
        time.sleep(1)
        t -= 1

def exitable_input(interface):
    #An interface system to exit at any time
    while True:
        thing = input(interface)
        if thing == "exit":
            sure = input("Programı kapatmak istediğinizden emin misiniz? (e)vet/(h)ayır: ")
            #https://youtu.be/BLikP6BDH5w
            if sure in ["e", "E", "evet", "Evet"]: 
                import sys
                sys.exit()
            else:
                continue               
        else:
            return thing

def checkGuiInputString(userInput):
    import re
    if re.match("^[0-9a-zA-Z\-\_\(\)ığĞüÜşŞİöÖçÇ ]+$", userInput):
        return True
    else:
        return False
    
def checkGuiInputFloat(userInput):
    try:
        userInput = float(userInput)
        return True
    except:
        return False
    
def checkGuiInputInt(userInput):
    try:
        userInput = int(userInput)
        return True
    except:
        return False

def hexaInput(questionString):
    import re
    while True:
        result = exitable_input(questionString)
        if not re.match("^[0-9a-zA-Z\-\_\(\)\/ığĞüÜşŞİöÖçÇ ]+$", result):
            print("Lütfen sadece sayı, harf ve/ya \"-\" giriniz.")
            continue
        else:
            break
    return result

def freeInput(questionString):
    result = exitable_input(questionString)
    return result

def numInput(questionString):
    import re
    while True:
        result = exitable_input(questionString)
        if not re.match("^[0-9]+$", result):
            print("Lütfen sadece sayı giriniz.")
            continue
        else:
            break
    return result
    
def floInput(questionString):
    while True:
        result = exitable_input(questionString)
        try:
            result = float(result)
            break
        
        except SystemExit:
            raise
        
        except:
            print("Lütfen sadece sayı giriniz.")
            continue
    return result

    
def ensure_dir(file_path):
    import os
    try:
        os.makedirs(file_path, exist_ok = True)
    
    except:
        print("Houston we have a problem here")
        
    return

def changeSettingsBool(setting_name):
    #Enables the user to change settings within the program
    configFilePath = getUserDocumentsPsiPath() + "settings.ini"
    try:
        import configparser
        parser = configparser.ConfigParser()
        parser.read(configFilePath, encoding = 'utf-8-sig')
        
        from configobj import ConfigObj
        config = ConfigObj(configFilePath, encoding='UTF8')
         
        state = parser.get('General', setting_name)
        
        if state == "True":
            state = "False"
            
        elif state == "False":
            state = "True"    
        
            
        config['General'][setting_name] = state   
        config.write()
                      
    except SystemExit:
        raise
    except:
        raise
        print("Lütfen yalnızca sayı giriniz")
    
    return
    
def changeSettingsString(setting_name, new_setting):
    #Enables the user to change settings within the program
    configFilePath = getUserDocumentsPsiPath() + "settings.ini"
    try:
        import configparser
        parser = configparser.ConfigParser()
        parser.read(configFilePath, encoding = 'utf-8-sig')
        
        from configobj import ConfigObj
        config = ConfigObj(configFilePath, encoding='UTF8')
                
        state = parser.get('General', setting_name)
        
        state = new_setting
                  
        config['General'][setting_name] = state   
        config.write()
                      
    except SystemExit:
        raise
    except:
        raise
    
    return

def settings(which_setting):
    #returns a boolean whether if it should create a file or not. Also a string for data path
    import os
    cwd = os.getcwd()
    configFilePath = getUserDocumentsPsiPath() + "settings.ini"
    x = True
    
    documentsPath = getUserDocumentsPsiPath()
    
    if which_setting in ["csv_path", "txt_path", "excel_path"]:
        if (which_setting == "csv_path"):
            setting_to_return =  settings("folder_name") + ("csv/")
        if (which_setting == "txt_path"):
            setting_to_return =  settings("folder_name") + ("txt/")
        if (which_setting == "excel_path"):
            setting_to_return =  settings("folder_name") + ("excel/")
        ensure_dir(setting_to_return)
        return setting_to_return
    
    while x:
        try:
            import configparser
            parser = configparser.ConfigParser()
            parser.read(configFilePath, encoding = 'utf-8-sig')
            
            setting_to_return = parser.get('General', which_setting)
            
            if (which_setting == "excel_name"):
                if setting_to_return == "default":
                    setting_to_return = "PsiNorm"
                                
                setting_to_return = setting_to_return + ".xlsx"
                x = False
                return setting_to_return
            
            if (which_setting == "form_data"):
                setting_to_return = cwd + "/Data/" + setting_to_return
                x = False
                return setting_to_return
            
            if (which_setting == "info_data"):
                setting_to_return = cwd + "/Data/" + setting_to_return
                x = False
                return setting_to_return
            
            if (which_setting == "folder_name"):
                if setting_to_return == "default":
                    setting_to_return =  documentsPath + ("Results/")
                else:
                    setting_to_return = documentsPath + (setting_to_return) + "/"
                    
            if which_setting == "textFileNameFormat":
                return setting_to_return
            
            if setting_to_return == "True":
                setting_to_return = True
                
            if setting_to_return == "False":
                setting_to_return = False
                
            if setting_to_return != True and setting_to_return != False:
                ensure_dir(setting_to_return)
                #Ensures the file path is applicable and available
                
            x = False
            return setting_to_return
                
        except:
            print("Ayarlar dosyası bulunamadı veya hatalı, tekrar oluşturuluyor.")
            with open(configFilePath, 'w', encoding='utf-8-sig') as file:
                file.write(
"""[General]
folder_name = default

output_csv = False

output_txt = True

output_excel = True
excel_name = PsiNorm

textFileNameFormat = firstDate

form_data = form_data.json
info_data = info_data.json

first_ever_run = True
"""
)
            continue

def excelWriter(excel_path, data_num, printable_list, which_data, excelColNameDict):
    while True:
        try:
            patient_name_local = patient_name      
                
            demographic_data = [patient_ID, patient_name_local, patient_admin, date, time, patient_age, patient_sex, patient_edu]
            
            test_name_list = [
                                "(1)MMT", "(2)MOCA", "(3)3MS", "(4)GISD", "(5)ECR",
                                "(6)Öktem Sözel Bellek Süreçleri", "(7)Rey Karmaşık Figür", "(8)İz Sürme", "(9)Stroop",
                                "(10)Wisconsin", "(11)Görsel Sözel Test", "(12)Renkli İz Sürme",
                                "(13)Wechsler", "(14)Wechsler-Sayı Dizisi", "(15)Sözel Akıcılık", 
                                "(16)Semantik Akıcılık", "(17)Saat Çizme", "(18)SDOT", "(19)Ayları İleri-Geri Sayma",
                                "(20)İşaretleme", "(21)Boston Adlandırma Testi"
                                ]
            
            testExcelColNameDict = {
        "1": ["Puan", "Sözel sonuç"],
        "2": ["Puan", "Sözel sonuç"],
        "3": ["Puan"],
        "4": ["Puan", "Z Skoru", "Persentil", "Sözel sonuç"],
        "5": ["Puan", "Sözel sonuç"],
        "6": ["Kendiliğinden hatırlama", "Z Skoru", "Persentil", "Sözel sonuç",
        "Tanıma boyutu", "Z Skoru", "Persentil", "Sözel sonuç",
        "Toplam hatırlama boyutu", "Z Skoru", "Persentil", "Sözel sonuç",
        "Hatırlama yanlışı boyutu", "Z Skoru", "Persentil", "Sözel sonuç",
        "Yanlış tanıma boyutu", "Z Skoru", "Persentil", "Sözel sonuç",
        "A listesi anlık bellek boyutu", "Z Skoru", "Persentil", "Sözel sonuç",
        "A listesi toplam öğrenme boyutu", "Z Skoru", "Persentil", "Sözel sonuç"],
              
        "7": ["Kopyalama", "Z Skoru", "Persentil", "Sözel sonuç",
        "Anlık hatırlama", "Z Skoru", "Persentil", "Sözel sonuç",
        "Gecikmeli hatırlama", "Z Skoru", "Persentil", "Sözel sonuç",
        "Tanıma doğru pozitif", "Z Skoru", "Persentil", "Sözel sonuç",
        "Tanıma yanlış pozitif", "Z Skoru", "Persentil", "Sözel sonuç"],
              
        "8": ["A skoru", "Z Skoru", "Persentil", "Sözel sonuç",
        "A formu düzeltme sayısı", "Z Skoru", "Persentil", "Sözel sonuç",
        "B skoru", "Z Skoru", "Persentil", "Sözel sonuç",
        "B formu düzeltme sayısı", "Z Skoru", "Persentil", "Sözel sonuç",
        "B+A skoru", "Z Skoru", "Persentil", "Sözel sonuç",
        "B-A skoru", "Z Skoru", "Persentil", "Sözel sonuç"],
        
        "9": ["Bölüm 1", "Z Skoru", "Persentil", "Sözel sonuç",
        "Hata-1", "Z Skoru", "Persentil", "Sözel sonuç",
        "Düzeltme-1", "Z Skoru", "Persentil", "Sözel sonuç",
        "Bölüm 2", "Z Skoru", "Persentil", "Sözel sonuç",
        "Hata-2", "Z Skoru", "Persentil", "Sözel sonuç",
        "Düzeltme-2", "Z Skoru", "Persentil", "Sözel sonuç",
        "Bölüm 3", "Z Skoru", "Persentil", "Sözel sonuç",
        "Hata-3", "Z Skoru", "Persentil", "Sözel sonuç",
        "Düzeltme-3", "Z Skoru", "Persentil", "Sözel sonuç",
        "Bölüm 4", "Z Skoru", "Persentil", "Sözel sonuç",
        "Hata-4", "Z Skoru", "Persentil", "Sözel sonuç",
        "Düzeltme-4", "Z Skoru", "Persentil", "Sözel sonuç",
        "Bölüm 5", "Z Skoru", "Persentil", "Sözel sonuç",
        "Hata-5", "Z Skoru", "Persentil", "Sözel sonuç",
        "Düzeltme-5", "Z Skoru", "Persentil", "Sözel sonuç"],
          
        "10": ["Toplam tepki sayısı puanı",  "Z Skoru", "Persentil", "Sözel sonuç",
        "Tamamlanan kategori sayısı puanı", "Z Skoru", "Persentil", "Sözel sonuç",
        "Toplam perseveretif hata sayısı puanı", "Z Skoru", "Persentil", "Sözel sonuç",
        "Perseveratif hata yüzdesi puanı", "Z Skoru", "Persentil", "Sözel sonuç"],
               
        "11": ["Toplam 'sort' sayısı",  "Z Skoru", "Persentil", "Sözel sonuç",
         "Toplam 'shift' sayısı",  "Z Skoru", "Persentil", "Sözel sonuç"],
         
        "12": ["Time", "Z Skoru", "Persentil", "Sözel sonuç",
        "Errors", "Z Skoru", "Persentil", "Sözel sonuç",
        "Near-Misses", "Z Skoru", "Persentil", "Sözel sonuç",
        "Prompts", "Z Skoru", "Persentil", "Sözel sonuç",
        "Time", "Z Skoru", "Persentil", "Sözel sonuç",
        "Number Errors", "Z Skoru", "Persentil", "Sözel sonuç",
        "Color Errors", "Z Skoru", "Persentil", "Sözel sonuç",
        "Near-Misses", "Z Skoru", "Persentil", "Sözel sonuç",
        "Prompts", "Z Skoru", "Persentil", "Sözel sonuç"],
               
        "13": ["Genel bilgi standart puan ",
        "Yargılama standart puan ",
        "Aritmetik standart puan ",
        "Benzerlik standart puan ",
        "Sayı dizisi standart puan ", 
        "Kelime standart puan ",
        "Şifre standart puan ",
        "Resim tamamlama standart puan ",
        "Küplerle desen standart puan ",
        "Resim düzenleme standart puan ",
        "Parça birleştirme standart puan ",
        "Sözel standart puan ",
        "Sözel IQ",
        "Performans standart puan ",
        "Performans IQ",
        "Toplam standart puan ",
        "Toplam IQ"],
               
        "14": ["Sayı dizisi toplam", "Sözel sonuç",
        "Sayı dizisi ileri", "Sözel sonuç",
        "Sayı dizisi geri", "Sözel sonuç"],
        
        "15": ["'S' harfi için: " , "Z Skoru", "Persentil", "Sözel sonuç",
               "'A' harfi için: ", "Z Skoru", "Persentil", "Sözel sonuç",
               "'Z' harfi için: ", "Z Skoru", "Persentil", "Sözel sonuç"],
               
        "16": ["Hayvan için: ", "Z Skoru", "Persentil", "Sözel sonuç",
               "İnsan için: ", "Z Skoru", "Persentil", "Sözel sonuç",
               "İnsan-Hayvan için: ", "Z Skoru", "Persentil", "Sözel sonuç"],
                
        "17": ["Skor: ", "Z Skoru", "Persentil", "Sözel sonuç"],
        
        "18": ["Puan ", "Z Skoru", "Persentil", "Sözel sonuç"],
        
        "19": ["İleri yorum","Geri yorum"],
        
        "20": ["Düzenli HF - İşaretlenen hedef sayısı puanı: ", "Z Skoru", "Persentil", "Sözel sonuç", 
                   "Düzenli HF - Atlanan hedef sayısı puanı: ", "Z Skoru", "Persentil", "Sözel sonuç", 
                   "Düzenli HF - İşaretlenen yanlış half sayısı puanı: ", "Z Skoru", "Persentil", "Sözel sonuç", 
                   "Düzenli HF - Toplam hata puanı: ", "Z Skoru", "Persentil", "Sözel sonuç", 
                   "Düzenli HF - Tarama süresi puanı: ", "Z Skoru", "Persentil", "Sözel sonuç", 
                   "Düzenli ŞF - İşaretlenen hedef sayısı puanı: ", "Z Skoru", "Persentil", "Sözel sonuç", 
                   "Düzenli ŞF - Atlanan hedef sayısı puanı: ", "Z Skoru", "Persentil", "Sözel sonuç",  
                   "Düzenli ŞF - İşaretlenen yanlış half sayısı puanı: ", "Z Skoru", "Persentil", "Sözel sonuç",  
                   "Düzenli ŞF - Toplam hata puanı: ", "Z Skoru", "Persentil", "Sözel sonuç", 
                   "Düzenli ŞF - Tarama süresi puanı: ", "Z Skoru", "Persentil", "Sözel sonuç", 
                   "Düzensiz HF - İşaretlenen hedef sayısı puanı: ", "Z Skoru", "Persentil", "Sözel sonuç", 
                   "Düzensiz HF - Atlanan hedef sayısı puanı: ", "Z Skoru", "Persentil", "Sözel sonuç", 
                   "Düzensiz HF - İşaretlenen yanlış half sayısı puanı: ", "Z Skoru", "Persentil", "Sözel sonuç", 
                   "Düzensiz HF - Toplam hata puanı: ", "Z Skoru", "Persentil", "Sözel sonuç", 
                   "Düzensiz HF - Tarama süresi puanı: ", "Z Skoru", "Persentil", "Sözel sonuç",  
                   "Düzensiz ŞF - İşaretlenen hedef sayısı puanı: ", "Z Skoru", "Persentil", "Sözel sonuç", 
                   "Düzensiz ŞF - Atlanan hedef sayısı puanı: ", "Z Skoru", "Persentil", "Sözel sonuç",  
                   "Düzensiz ŞF - İşaretlenen yanlış half sayısı puanı: ", "Z Skoru", "Persentil", "Sözel sonuç", 
                   "Düzensiz ŞF - Toplam hata puanı: ", "Z Skoru", "Persentil", "Sözel sonuç", 
                   "Düzensiz ŞF - Tarama süresi puanı: ", "Z Skoru", "Persentil", "Sözel sonuç"],
               
        "21": ["KA puanı: ", "Z Skoru", "Persentil", "Sözel sonuç",
               "KA+AİA puanı: ", "Z Skoru", "Persentil", "Sözel sonuç",
            "KA+AİA+SİA puanı: ", "Z Skoru", "Persentil", "Sözel sonuç"]
        
                    }
            
            
            
            
            excel_sheet_names_list = []
            
            mainDict = jsonLoader("form_data")
                
            for i in mainDict["formNames"].keys():
                excel_sheet_names_list.append(mainDict[i]["excelSheetName"])
            
            
            excel_sheet_names_list = excel_sheet_names_list + test_name_list
            
            from openpyxl import Workbook
            from openpyxl import load_workbook
            
            wb = Workbook(write_only=True)
            #wbwo = Workbook(write_only=True)
            
            while True:
                try:
                    for sheet_name in excel_sheet_names_list:
                        data_workbook = load_workbook(filename = excel_path + settings("excel_name"), read_only=False)
                        active_sheet = data_workbook.get_sheet_by_name(sheet_name)
                        data_workbook.close()
                    break
                
                except PermissionError:
                    guiSimpleTextDump("Uyarı!", "Excel dosyası arka planda açık görünüyor, lütfen kapatınız. ")
                    continue
                
                except:
                    userInput = numInput("""
Excel dosyası bulunamadı veya hatalı. 
> Dosyayı tekrar kontrol edip, tekrar denemek için (1)
> Excel kaydını iptal etmek için (2) DİKKAT, BU SEÇENEK PROGRAMI KAPATIR.
> Dosyayı tekrar oluşturmak için (3)
giriniz. 
Cevap: """)
                    if userInput == "1":
                        continue
                    elif userInput == "2":
                        print("Excel kaydı iptal edildi. ")
                        import sys
                        sys.exit()
                    elif userInput == "3":
                        userInput = hexaInput("""
Bu işlem eğer eski bir excel dosyası var ise silecektir.
Eğer kayıtlı veriniz var ise, lütfen yedekleme yapınız.
Bu işlemi yapmak istediğinizden emin misiniz? (e)vet/(h)ayır: """)
                        if userInput in ["e", "E", "evet", "Evet"]: 
                            for sheet_name in excel_sheet_names_list:
                                wb.create_sheet(title = sheet_name)
                                                   
                            wb.save(filename = excel_path + settings("excel_name"))
                            data_workbook = load_workbook(filename = excel_path + settings("excel_name"), read_only=False)
                            
                            
                            for key in testExcelColNameDict.keys():
                                active_sheet = data_workbook.get_sheet_by_name(test_name_list[int(key)-1])
                                test_col_list = ["Hasta Kodu","Hasta isim", "Uygulayıcı", "Tarih", "Zaman", "Yaş", "Cinsiyet","Eğitim Yılı"]
                                test_col_list = test_col_list + testExcelColNameDict[key]
                                for col, val in enumerate(test_col_list, start=1):
                                    active_sheet.cell(row=1, column=col).value = val
                                    
                            #for key in mainDict["formNames"].keys():
                             #    active_sheet = data_workbook.get_sheet_by_name(mainDict["formNames"][key])
                            
                            data_workbook.save(filename = excel_path + settings("excel_name"))    
                            data_workbook.close()
                                
                                
        #                    wb.save(filename = excel_path + settings("excel_name"))
        #                    data_workbook = load_workbook(filename = excel_path + settings("excel_name"), read_only=False)
        #                    active_sheet = data_workbook.get_sheet_by_name("Sheet")
        #                    data_workbook.remove_sheet(active_sheet)
        #                    data_workbook.save(filename = excel_path + settings("excel_name"))
        #                    data_workbook.close()
                            break
                        else:
                            continue
                    else:
                        print("Lütfen varolan seçeneklerden seçiniz. ")
                        continue
        
            
            
            if which_data == "testData":
                try:
                    while True:
                        try:
                            data_workbook = load_workbook(filename = excel_path + settings("excel_name"), read_only=False)
                            active_sheet = data_workbook.get_sheet_by_name(test_name_list[data_num-1])
        
                            active_sheet.append(demographic_data + printable_list)
            
                            data_workbook.save(filename = excel_path + settings("excel_name"))
                            print(test_name_list[data_num-1] + " Excel'e kaydediliyor. ")
                            data_workbook.close()
                            break
                        
                        except PermissionError:
                            print("Dosya arka planda açık görünüyor, lütfen kapatınız. Program 5 saniye sonra tekrar deneyecek.")
                            wait(5)
                            continue
                        
                        except:
                            print("Excel sayfası bulunamadı. " + test_name_list[data_num-1] + " oluşturuluyor. ")
                            for i in range(len(test_name_list)):
                                wb.create_sheet(title = test_name_list[i])
                
                            wb.save(filename = excel_path + settings("excel_name"))
                            data_workbook = load_workbook(filename = excel_path + settings("excel_name"), read_only=False)
                            active_sheet = data_workbook.get_sheet_by_name("Sheet")
                            data_workbook.remove_sheet(active_sheet)
                            data_workbook.save(filename = excel_path + settings("excel_name"))
                            data_workbook.close()
                            continue
                        
                except:
                    raise
                    
            else:
                final_list = []
                colNum_list = ["Hasta Kodu","Hasta isim", "Uygulayıcı", "Tarih", "Zaman", "Yaş", "Cinsiyet","Eğitim Yılı"]
                for i in range(len(printable_list)):
                    colNum_list.append(str(printable_list[i][0]) + ") " + str(excelColNameDict[printable_list[i][0]]))
                    final_list.append(printable_list[i][1])
                                
        
                while True:
                    try:
                        data_workbook = load_workbook(filename = excel_path + settings("excel_name"), read_only=False)
                        active_sheet = data_workbook.get_sheet_by_name(which_data)
                        
                        for col, val in enumerate(colNum_list, start=1):
                            active_sheet.cell(row=1, column=col).value = val
                        
                        active_sheet.append(demographic_data + final_list)
        
                        data_workbook.save(filename = excel_path + settings("excel_name"))
                        print(which_data + " Excel'e kaydediliyor. ")
                        data_workbook.close()
                        break
                    
                    except:
                        print("Excel sayfası bulunamadı. " + which_data + " oluşturuluyor. ")
                            
                        wb.create_sheet(title = which_data)
                        
                        wb.save(filename = excel_path + settings("excel_name"))
                        data_workbook = load_workbook(filename = excel_path + settings("excel_name"), read_only=False)
                        active_sheet = data_workbook.get_sheet_by_name("Sheet")
                        data_workbook.remove_sheet(active_sheet)
                        data_workbook.save(filename = excel_path + settings("excel_name"))
                        data_workbook.close()
                        continue

            break
     
        except PermissionError:
            print("Dosya arka planda açık görünüyor, lütfen kapatınız. Program 5 saniye sonra tekrar deneyecek.")
            wait(5)
            continue
        except:
            raise
            

def csvWriter(path, test_name, printable_list):
    while True:
        try:
            import csv
            
            test_name = path + test_name
            
            with open(test_name, 'a', encoding='UTF-8', newline='') as csvfile:
                data_writer = csv.writer(csvfile, delimiter='~', quotechar='|', quoting=csv.QUOTE_MINIMAL)
                data = [patient_admin, date, time, patient_ID, patient_age, patient_sex, patient_edu] + printable_list
                data_writer.writerow(data)
                
            break
        except PermissionError:
            print("Dosya arka planda açık görünüyor, lütfen kapatınız. Program 5 saniye sonra tekrar deneyecek.")
            wait(5)
            continue
        except:
            raise
        
def txtWrite(path, console_results):
    while True:
        try:
            
            fileNameFormat = settings("textFileNameFormat")
            
            if fileNameFormat == "firstID":
                file_name = path + str(patient_ID.replace(" ", "_")) + "-(" + date + "-" + stime + ")-" + patient_admin.replace(" ", "_") + ".txt"
            
            if fileNameFormat == "firstDate":
                file_name = path + "(" + date + "-" + stime + ")-" + str(patient_ID.replace(" ", "_")) + "-" + patient_admin.replace(" ", "_") + ".txt"
                
            if fileNameFormat == "firstAdmin":
                file_name = path + patient_admin.replace(" ", "_") + "-(" + date + "-" + stime + ")-" + str(patient_ID.replace(" ", "_")) + ".txt"
            
            import os.path
            if os.path.exists(file_name):
                with open(file_name, 'a', encoding='UTF-8') as file:
                    file.write("\n" + console_results)
                    print("Rapora ek yapılıyor. ")
                file.close()
                    
            else:
                with open(file_name, 'w', encoding='UTF-8') as file:
                    file.write("======= PsiNorm Persentil Hesaplayıcı =======" +
                               "\n============ Copyright (C) 2017 =============" +
                               "\n==== Bilal Bahadır Akbulut & Yavuz Ayhan ====" +
                    "\n=============================================\n" + 
                    "Testi uygulayan: " + patient_admin + "\nGünün tarihi: " + date + "\nSaat: " + time +
                    "\nHastanın ismi: " +  patient_name + "\nHastanın kodu: " + str(patient_ID) +
                    "\nHastanın yaşı: " + str(patient_age) + "\nHastanın cinsiyeti: " + str(patient_sex) +
                    "\nHastanın toplam eğitim yılı: " + str(patient_edu) +
                    "\n=============================================\n")
                    
                    file.write(console_results)
                    print("Rapor oluşturuluyor. ")
                file.close()
            break
            
        except PermissionError:
            print("Dosya arka planda açık görünüyor, lütfen kapatınız. Program 5 saniye sonra tekrar deneyecek.")
            wait(5)
            continue
        except:
            raise


def inputPatient_age():
#a simple loop to get patient's age  
    while True:   
        patient_age = exitable_input("Hastanın yaşı: ")
        
        if patient_age.isnumeric():
            patient_age = int(patient_age)
            
            if patient_age < 0:
                #ensures the number is bigger than 0
                print("Lütfen 0'dan büyük bir sayı giriniz.")
                continue
            else:
                break

        else:
            print("Lütfen sadece sayı giriniz.")
            continue
        
    return patient_age
    
def inputPatient_sex():   
#a simple loop to get the sex of the patient, provides two options, female or male
    while True:    
        patient_sex_user_input = exitable_input("Hastanın cinsiyeti: (1) Kadın - (2) Erkek: ")
        if patient_sex_user_input == "1":
            patient_sex = "Kadın"
            break
        elif patient_sex_user_input == "2":
            patient_sex = "Erkek"
            break
        else:
            print("Lütfen sadece 1 veya 2 giriniz.")
            continue
    return patient_sex
        
def inputPatient_edu(): 
#a simple loop to get the number of years of education
    while True:
        patient_edu = exitable_input("Hastanın toplam eğitim yılı: ")
        if patient_edu.isnumeric():
            patient_edu = int(patient_edu)
            break
        else:
            print("Lütfen sadece sayı giriniz.")
            continue
        
    return patient_edu

def calculate_age(patient_age):
    #gets day of birth and ensures it's in correct format
    from datetime import datetime, date
    today = date.today()
    global birthday
    
    while True:
        try: 
            print("Lütfen doğum tarihini giriniz.")
            global bornDay
            bornDay = numInput("Gün: ")
            global bornMonth
            bornMonth = numInput("Ay: ")
            global bornYear
            bornYear = numInput("Yıl: ")
            
            birthday = (bornDay + "/" + bornMonth + "/" + bornYear)
            
            if "999" in [bornDay, bornMonth, bornYear]:
                if bornYear == "999":
                    print("Yıl bilinmiyor, yaş hesaplanamaz.")
                    formPatientAge = "999"
                    
                elif bornYear != "999" and bornMonth == "999":
                    birthday = datetime.strptime(bornYear, "%Y")                    
                    print("Sadece yıl üzerinden yaş hesaplanıyor.")
                    formPatientAge = today.year - birthday.year
                    
                elif bornYear != "999" and bornMonth != "999":
                    birthday = datetime.strptime(bornMonth+"/"+bornYear, "%m/%Y")
                    print("Sadece ay ve yıl üzerinden yaş hesaplanıyor.")   
                    formPatientAge = today.year - birthday.year - ((today.month) < (birthday.month))
                    
            else:              
                birthday = datetime.strptime(birthday, "%d/%m/%Y")
                formPatientAge = today.year - birthday.year - ((today.month, today.day) < (birthday.month, birthday.day))
            
            if formPatientAge < 0 or formPatientAge > 150:
                print("Lütfen girdiğiniz tarihi kontrol edip tekrar giriniz.")
                continue
                
        except ValueError:
            print("Hatalı tarih, lütfen tekrar giriniz.\n")
            continue
        except SystemExit:
            raise
        except:
            raise
          
        birthday = (bornDay + "/" + bornMonth + "/" + bornYear)
        
        if patient_age != formPatientAge:
            #if there's discrepeancy between the calculated age and age entered, forces user to choose
            print("""
Girişte girilen yaş ile forma girilen yaş arasında fark bulundu.
Hangi sayıyı kullanmak istersiniz?
1) Girişte girilen: """ + str(patient_age) + """
2) Formda hesaplanan: """ + str(formPatientAge))
            
            while True:
                userInput = int(numInput("\n"))
                if 0 < userInput < 3:
                    if userInput == 1:
                        formPatientAge = patient_age
                    else:
                        globals()['patient_age'] = formPatientAge
                    
                    break
                
                else:
                    print("Lütfen '1' veya '2' giriniz.")
                    continue
                
        return birthday



def formMain(mainDict): #main function for the dataforms
    def eQuest(qDict, qNum, templateDict): #expandable questions
        
        ansDict = {}
        
        if qNum in templateDict.keys():
            ansDict = templateDict[qNum]
        
        def greet(qDict, ansDict):
            #prints out the data structure
            print("\n" + qNum + ") " + qDict["qText"])  
            
            toprint = ""
            toprintList = []
            for x in range(len(ansDict.keys())):
                toprint = str(x+1) + ") " +  str(ansDict[str(x+1)])
                
                toprintList.append(toprint)
                        
                        
            for i in range(len(toprintList)):
                print(toprintList[i])
        
            print("--------------------------------------------")    
               
        def addMoreFunc(qDict, ansDict): #add one item
            tempList = []
            for i in range(qDict["colNum"]):
                tempList.append(freeInput(qDict[str(i + 1)]))
                
            tempString = ""
            for i in range(len(tempList)):
                tempString = tempString + tempList[i]
                if len(tempList) > 1 and (i+1) < len(tempList):
                    tempString = tempString + " / "
                    
            ansDict[str(len(ansDict.keys()) + 1)] = tempString
            
            return ansDict
            
        def changeFunc(qDict, ansDict): #change an item
            while True:
                userInput = numInput("Lütfen değiştirmek istediğiniz maddeyi giriniz: ")
                if 0 < int(userInput) < len(ansDict.keys()) + 1:
                    tempList = []
                    for i in range(qDict["colNum"]):
                        tempList.append(freeInput(qDict[str(i + 1)]))
                        
                    tempString = ""
                    for i in range(len(tempList)):
                        tempString = tempString + tempList[i]
                        if len(tempList) > 1 and (i+1) < len(tempList):
                            tempString = tempString + " / "
                            
                    ansDict[userInput] = tempString
                    
                    return ansDict
                    
                else:
                    print("Lütfen varolan maddelerden birini seçiniz. ")
                    continue
                    
        
        while True:
            greet(qDict, ansDict)
            try:
                print("""
Seçenekler:
1) Yeni madde ekle.
2) Varolan maddeyi değiştir.
3) Varolan maddeyi sil.
4) Bir sonraki soruya geç.                              
                """)
                
                userInput = int(numInput("Lütfen seçiniz: "))
                if userInput == 1:
                    ansDict = addMoreFunc(qDict, ansDict)
                elif userInput == 2:
                    if len(ansDict.keys()) < 1:
                        print("Değiştirilecek girdi mevcut değildir, lütfen başka bir seçim yapınız.")
                        continue
                    else:
                        ansDict = changeFunc(qDict, ansDict)
                        continue
                
                elif userInput == 3:
                    if len(ansDict.keys()) < 1:
                        print("Silinecek girdi mevcut değildir, lütfen başka bir seçim yapınız.")
                        continue
                    else:
                        userInput = numInput("Silincecek madde numarası: ")
                        if 0 < int(userInput) <= len(ansDict.keys()):
                            to_shorten = len(ansDict.keys()) - int(userInput) 
                            if to_shorten > 0:
                                for i in range(to_shorten):
                                    ansDict[str(int(userInput) + i)] = ansDict[str(int(userInput) + i + 1)]
                                
                                del ansDict[str(len(ansDict.keys()))]
                            
                            if to_shorten == 0:
                                del ansDict[userInput]
                            
                            continue
                        else:
                            print("Lütfen varolan maddelerden seçiniz. ")
                            continue
                
                elif userInput == 4:
                    return ansDict
                

            except SystemError:
                raise
            except:
                raise
            
        
    #eQuest(qDict)
    
    
    def mQuest(qDict, qNum): #multiple choice question
        while True:
            try:
                print("\n" + qNum + ") " + qDict["qText"])
                for i in range(qDict["optNum"]):
                    print(str(i+1) + ") " + qDict[str(i+1)])
                userInput = numInput("Cevap: ")            
            except SystemError:
                raise
            except:
                raise
                
            if 0 < int(userInput) <= qDict["optNum"]:
                break
            else:
                print("Lütfen varolan seçeneklerden birini seçiniz.")
                continue
            
        if str(userInput) in qDict["extraMark"]:
            return qDict[userInput], (qNum + "-" + userInput)
        else:
            return qDict[userInput], None
                
    
    def cQuest(qDict, qNum): #classic question
        while True:
            try:
                userInput = freeInput(qNum + ") " + qDict["qText"])
            except SystemError:
                raise
            except:
                print("Bir hata oluştu, lütfen bildiriniz.")
                raise
            return userInput
          
    def decideFunc(qDict, qNum, templateDict): #checks which form of question it s
        extraMark = None
        
        if qDict["qType"] == "eQuest":
            answer = eQuest(qDict, qNum, templateDict)
            
        elif qDict["qType"] == "mQuest":
            answer, extraMark = mQuest(qDict, qNum)
            
        elif qDict["qType"] == "cQuest":
            answer = cQuest(qDict, qNum)
            
        elif qDict["qType"] == "runFunc":
            arg = eval(qDict["arg"])
            answer = eval(qDict["runFunc"])(arg)
            
        elif qDict["qType"] == "pullData":
            print("\n" + qNum + ") " + qDict["qText"] + str(eval(qDict["data"])))       
            answer = eval(qDict["data"])
        
        else:
            print("Beklenmeyen bir hata oluştu, program kapatılacak.")
            raise
            
        return answer, extraMark
            
    
    
    def formFunc(mainDict, excelColNameDict): #director of the questions
        endAnsDict = {}
        
        for i in range(len(list(mainDict["qDict"].keys()))):
            if str(i+1) in mainDict["titleDict"].keys():
                print(mainDict["titleDict"][str(i+1)])
                        
            qNum = str(i+1)
            qDict = mainDict["qDict"][qNum]
            xtraQDict = mainDict["xtraQDict"]
            templateDict = mainDict["templateDict"]
            
            endAnsDict[qNum], extraMark = decideFunc(qDict, qNum, templateDict)
            
            extraMarkList = [extraMark]
            
                
            while True:
                while None in extraMarkList:   
                    extraMarkList.remove(None)
                        
                if len(extraMarkList) > 0:
                    extraMarkToUse = extraMarkList[0]
                    extraMarkList.remove(extraMarkList[0])
                    
                    for i in range(len(list(xtraQDict[extraMarkToUse].keys()))):
                        i = str(i + 1)
                        ansKey = extraMarkToUse + "-" + i
                        endAnsDict[ansKey], extraMark = decideFunc(xtraQDict[extraMarkToUse][i], i, templateDict)
                        if extraMark != None:
                            extraMarkList.append(extraMarkToUse + "-" + extraMark)                 
                        
                        print("--------------------------------------------")
                        
                else:
                    break
                            
            print("--------------------------------------------")
      
        while True:
            userInput = numInput("""
Form bitmiştir.
> Değiştirmek istediğiniz soru var ise (1)
> Kaydetmek için (2)
giriniz: """)
            if userInput == "1":
                while True:
                    userInput = numInput("Değiştirmek istediğiniz sorunun numarasını giriniz: ")
                    if userInput in endAnsDict.keys():
                        for i in endAnsDict.keys():
                            key = i
                            if key.startswith(userInput + "-"):
                                del endAnsDict[key]
                        
                        qNum = userInput
                        qDict = mainDict["qDict"][qNum]
                        xtraQDict = mainDict["xtraQDict"]
                        templateDict = mainDict["templateDict"]
                        
                        endAnsDict[qNum], extraMark = decideFunc(qDict, qNum, templateDict)
                
                        extraMarkList = [extraMark]
                        
                        while True:
                            while None in extraMarkList:   
                                extraMarkList.remove(None)
                                    
                            if len(extraMarkList) > 0:
                                extraMarkToUse = extraMarkList[0]
                                extraMarkList.remove(extraMarkList[0])
                                
                                for i in range(len(list(xtraQDict[extraMarkToUse].keys()))):
                                    i = str(i + 1)
                                    ansKey = extraMarkToUse + "-" + i
                                    endAnsDict[ansKey], extraMark = decideFunc(xtraQDict[extraMarkToUse][i], i, templateDict)
                                    if extraMark != None:
                                        extraMarkList.append(extraMarkToUse + "-" + extraMark)                 
                                    
                                    print("--------------------------------------------")
                                    
                            else:
                                break
                                        
                        print("--------------------------------------------")
                        
                        break
                        
                    else:
                        print("Lütfen varolan sorulardan seçiniz.")
                        continue
                
                
            elif userInput == "2":
                break
            
            else:
                print("Lütfen varolan seçeneklerden birini seçiniz.")
                continue
            
        
        endAnsList, console_results = ansListDecoder(mainDict, endAnsDict, excelColNameDict)
        return endAnsList, console_results
    
      
    def ansListDecoder(mainDict, endAnsDict, excelColNameDict): #translates the answer list into excel & txt format
        endConsAnsList = []
        endExcAnsList = []
        endConsAnsDict = endAnsDict.copy() #readies up a console dictionary
        import natsort
        
        for key in natsort.natsorted(endAnsDict.keys(), key=lambda y: y.lower()):
            endExcAnsList.append([key,endAnsDict[key]])
            endConsAnsList.append([key,endAnsDict[key]])
            #sorts answer dictionary to a list for accounting expanding questions
    
        for i in endExcAnsList:
            qNum = i[0]
            ans = i[1]
            if isinstance(ans, dict):
                expList = []
                
                for key in natsort.natsorted(ans.keys(), key=lambda y: y.lower()):
                    expList.append(ans[key])
                    
                for i in range(len(expList)): #makes expQ answers compatible with the rest of the answers
                    endConsAnsDict[qNum + "-" + str(i+1)] = expList[i]                
                del endConsAnsDict[qNum]
                
                if len(expList) <= 20: #ensures there are 20 columns in the excel file for expanding Q
                    tempString = ""
                    for i in range(mainDict["qDict"][qNum]["colNum"]):
                        tempString = tempString + '999'
                        if mainDict["qDict"][qNum]["colNum"] > 1 and (i+1) < mainDict["qDict"][qNum]["colNum"]:
                            tempString = tempString + " / "
                    for i in range(20-len(expList)):                    
                        expList.append(tempString)
                        
                for i in range(len(expList)): #makes expQ answers compatible with the rest of the answers
                    endAnsDict[qNum + "-" + str(i+1)] = expList[i]
                del endAnsDict[qNum]
                 
                
        excelColNamesList = keyCounter(mainDict)
        for i in excelColNamesList:
            if i not in endAnsDict.keys():
                endAnsDict[i] = "-"
        
                
        endExcAnsList = []
        for key in natsort.natsorted(endAnsDict.keys(), key=lambda y: y.lower()):
            endExcAnsList.append([key,endAnsDict[key]])
            #finalizes the list
            
        endConsAnsList = []
        for key in natsort.natsorted(endConsAnsDict.keys(), key=lambda y: y.lower()):
            endConsAnsList.append([key,endConsAnsDict[key]])
            #finalizes the list
            
        console_results = "\n========================================================\n"
        for i in range(len(endConsAnsList)):
            console_results = console_results + str(endConsAnsList[i][0]) + ") "
            #if str(endExcAnsList[i][0]) in mainDict["qDict"].keys():
             #   console_results = console_results + mainDict["qDict"][str(endExcAnsList[i][0])]["qText"]
            #if str(endExcAnsList[i][0]) in mainDict["xtraQDict"].keys():
            #    console_results = console_results + mainDict["xtraQDict"][str(endExcAnsList[i][0])]["qText"]               
             
            console_results = console_results + excelColNameDict[str(endConsAnsList[i][0])]
            
            console_results = console_results + str(endConsAnsList[i][1]) + "\n"
                    
        console_results = console_results + "========================================================"

        return endExcAnsList, console_results
        
            
    def keyCounter(mainDict): #ensures all the proper columns exists within the excel file
        excelColNameDict = {}
        
        for i in mainDict["qDict"].keys():       
            key = i
            if mainDict["qDict"][key]["qType"] == "eQuest":
                for i in range(20):  
                    excelColNameDict[key + "-" + str(i+1)] = mainDict["qDict"][key]["qText"] + "-" + str(i+1)
                
            elif mainDict["qDict"][key]["qType"] == "mQuest":
                excelColNameDict[key] = mainDict["qDict"][key]["qText"]
                for i in mainDict["qDict"][key]["extraMark"]:
                    if i != None:
                        xtraKey = i
                        for i in mainDict["xtraQDict"][key + "-" + xtraKey].keys():
                            excelColNameDict[key + "-" + xtraKey + "-" + i] = mainDict["xtraQDict"][key + "-" + xtraKey][i]["qText"]
                            doubleXtraKey = i
                            if "extraMark" in mainDict["xtraQDict"][key + "-" + xtraKey][doubleXtraKey].keys():
                                excelColNameDict[key + "-" + xtraKey + "-" + doubleXtraKey] = mainDict["xtraQDict"][key + "-" + xtraKey][doubleXtraKey]["qText"]
    
    
            elif mainDict["qDict"][key]["qType"] == "cQuest":
                excelColNameDict[key] = mainDict["qDict"][key]["qText"]
                
            elif mainDict["qDict"][key]["qType"] == "runFunc":
                excelColNameDict[key] = mainDict["qDict"][key]["qText"]
                
            elif mainDict["qDict"][key]["qType"] == "pullData":
                excelColNameDict[key] = mainDict["qDict"][key]["qText"]
                
            else:
                print("This shouldn't happen.")
                
        return excelColNameDict

    excelColNameDict = keyCounter(mainDict)
    printable_list, console_results = formFunc(mainDict, excelColNameDict)
    return printable_list, console_results, excelColNameDict
    

def mainMenu():
    #creates a list for the tests that were employed before, so it doesn't get used again
    menu_list = [testMmt, testMoca, test3ms, testGisd, testEcr, testSbst, testRkft, 
                         testTm, testStroop, testWisconsin, testVvt, testCct, testWechsler,
                         testWechslerSayi, testVf, testSf, testCd, testSdot, testMonths,
                         testVaNVC, testBNT
                         ]  
                #list of tests' function names

          
    test_name_list = [
    "(1)Mini Mental test ", "(2)Montreal Bilişsel Değerlendirme ", 
    "(3)3MS ", "(4)Görsel İşitsel Sayı Dizileri ", "(5)Artırılmış İpuçlu Hatırlama ",
    "(6)Öktem Sözel Bellek Süreçleri ", "(7)Rey Karmaşık Figür ", "(8)İz Sürme ", "(9)Stroop ",
    "(10)Yetişkin Wisconsin Kart Eşleme ", "(11)Görsel Sözel Test ", "(12)Renkli İz Sürme ",
    "(13)Wechsler Zeka Testi ", "(14)Wechsler Zeka Testi-Sadece Sayı Dizisi ",
    "(15)Sözel Akıcılık ", "(16)Semantik Akıcılık ", "(17)Saat Çizme ", "(18)Çizgi Yönünü Belirleme ",
    "(19)Ayları İleri-Geri Sayma ", "(20)İşaretleme ", "(21)Boston Adlandırma Testi "
    ]
        #verbal names of the tests      
    
    data_dict = {}
    done_tests_vanilla = []
    for i in range(len(menu_list)):
        done_tests_vanilla.append("(-)")       
                         
    menu_ui_layout = """
|>=============================================|=====================================================<|
| DİKKAT >>> EĞER ÇIKMADAN ÖNCE """ +  str(len(menu_list)+1) + """ GİRMEZSENİZ PROGRAM BİLGİLERİ KAYDETMEDEN KAPANACAKTIR. <<< DİKKAT |
|   DİKKAT >>>  ÇIKIŞ KOMUTU VERDİKTEN SONRA LÜTFEN PROGRAMIN KAYDETMESİNİ BEKLEYİNİZ.   <<< DİKKAT   |
|>=============================================|=====================================================<|    
|                                              |                                                      | 
|>=============================================|=====================================================<|
|A. Genel Bilişsel Tarama Testleri:            |B. Dikkat Testleri:                                   |
|(1)Mini Mental Test                  {0!s:^8} |(4)Görsel İşitsel Sayı Dizileri              {3!s:^8} |
|(2)Montreal Bilişsel Değerlendirme   {1!s:^8} |(14)Wechsler Zeka Testi-Sadece Sayı Dizisi  {13!s:^9} |
|(3)3MS                               {2!s:^8} |(19)Ayları İleri-Geri Sayma                 {18!s:^9} |
|                                              |                                                      |
|>=============================================|=====================================================<|
|C. Bellek Testleri:                           |D.Yönetici İşlev Testleri:                            |
|(5)Artırılmış İpuçlu Hatırlama       {4!s:^8} |(8)İz Sürme                                  {7!s:^8} |
|(6)Öktem Sözel Bellek Süreçleri      {5!s:^8} |(9)Stroop                                    {8!s:^8} |
|(7)Rey Karmaşık Figür                {6!s:^8} |(10)Yetişkin Wisconsin Kart Eşleme           {9!s:^8} |
|                                              |(11)Görsel Sözel Test                       {10!s:^9} |
|                                              |(12)Renkli İz Sürme                         {11!s:^9} |
|>=============================================|=====================================================<|
|E.Lisan Testleri:                             |F.Görsel-Uzaysal İşlev Testleri:                      |
|(15)Sözel Akıcılık                  {14!s:^9} |(7)Rey Karmaşık Figür                        {6!s:^8} |
|(16)Semantik Akıcılık               {15!s:^9} |(17)Saat Çizme                              {16!s:^9} |
|(21)Boston Adlandırma Testi         {20!s:^9} |(18)Çizgi Yönünü Belirleme                  {17!s:^9} |
|                                              |(20)İşaretleme                              {19!s:^9} |
|>=============================================|=====================================================<|
|G.WAIS:                                       |                                                      |
|(13)Wechsler Zeka Testi             {12!s:^9} |                                                      |
|>=============================================|=====================================================<|
| Not: Girilmemiş testlerin yanında "(-)", girilmiş testlerin yanında "(+)" işareti bulunmaktadır.    |
|      Eğer bir test iki defa girilmiş ise, yanında "(!+!)" işareti bulunmaktadır.                    |
|>=============================================|=====================================================<|
|                                              |                                                      |
|>=============================================|=====================================================<|
|                            Bilgi alma formları için '888' giriniz.                                  |
|>=============================================|=====================================================<|
| DİKKAT >>> EĞER ÇIKMADAN ÖNCE """ +  str(len(menu_list)+1) + """ GİRMEZSENİZ PROGRAM BİLGİLERİ KAYDETMEDEN KAPANACAKTIR. <<< DİKKAT |
|   DİKKAT >>>  ÇIKIŞ KOMUTU VERDİKTEN SONRA LÜTFEN PROGRAMIN KAYDETMESİNİ BEKLEYİNİZ.   <<< DİKKAT   |
|>=============================================|=====================================================<|    
"""

    menu_ui = menu_ui_layout.format(*tuple(done_tests_vanilla))
    done_tests_proper = done_tests_vanilla

    
    while True:
        try: 
            menu_done_tests_nums = sorted(list(data_dict.keys()))
            #ensures there are no duplicates on the done tests list
            
            menu_ui = menu_ui_layout.format(*tuple(done_tests_proper))
            print(menu_ui)
            #creates a user list from the done tests, without duplicates
             
            
            menu_input = numInput("Girmek istediğiniz testin numarasını giriniz. Kaydetmek için (" + str(len(menu_list)+1) + ") giriniz: " )   
            #input the test number you want to use, or press the calculated number (one higher than the test number) to exit
                          
            menu_input = int(menu_input)
            
            if menu_input in menu_done_tests_nums:
                x = input("Bu test zaten girilmiş, tekrar girmek istiyor musunuz? (e)vet/(h)ayır: ")
                if x in ["E", "e"]:
                    print("Test tekrar giriliyor.")                  
                else:
                    print("Bir önceki basamağa geri dönülüyor.")
                    continue
                    #returns to the previous step
                    
            if 0 < menu_input <= len(menu_list):
                print("\nLÜTFEN DİKKAT: Test içerisinde uygulamadığınız veya olmayan değerleri 999 olarak giriniz.")
                
                done_tests_proper[menu_input-1] = "(+)"
                if menu_input in menu_done_tests_nums:
                    done_tests_proper[menu_input-1] = "(!+!)"  

                data = menu_list[menu_input-1]()
                data_dict.update({menu_input : data})
                continue
                #using the number, calls the function user wanted, adds that to the done tests

            elif menu_input == (len(menu_list)+1):   
                try: 
                    for data_num in menu_done_tests_nums:
                                            
                        test_name = data_dict[data_num][0]
                        printable_list = data_dict[data_num][1]
                        console_results = data_dict[data_num][2]
                        #using the numbers from the used tests list, it creates data to feed into the writers
                
                        if settings("output_excel"):
                            excelWriter(settings("excel_path"), data_num, printable_list, "testData", None)
                            
                        if settings("output_csv"):
                            csvWriter(settings("csv_path"), test_name, printable_list)
                            #writes the printable_list in a CSV file
                        if settings("output_txt"):
                            txtWrite(settings("txt_path"), console_results)
                            #writes the console results on the txt file
                            
                    if settings("output_txt"):
                        done_tests_string = "\n----------------------------------------\nYapılan testler listesi: "
                        for i in menu_done_tests_nums:
                            done_tests_string = done_tests_string + "\n> " + test_name_list[i-1]
                            
                        done_tests_string = done_tests_string + "\n----------------------------------------"
                        
                        
                        txtWrite(settings("txt_path"), done_tests_string)      

                    if settings("output_txt"):
                        references = jsonLoader("info_data")["references_data"]
                        txtWrite(settings("txt_path"), references)
                        
                    print("\nSONUÇLAR KULLANICIYA AİT 'BELGELERİM' VEYA 'DOCUMENTS' KLASÖRÜ İÇERİSİNDE 'PsiNorm' İÇİNE KAYDEDİLMEKTEDİR.")
                    print("\nDOSYA İSİMLERİNDEKİ BOŞLUKLAR TEKNİK SEBEPLERDEN ÖTÜRÜ OTOMATİK OLARAK \"_\" İLE DEĞİŞTİRİLMEKTEDİR.")
                    print("\nŞu ana kadar yapılanlar kaydedildi.\nProgram baştan başlatılıyor.")    
                    break
                    #this is the exit command
                except SystemExit:
                    raise
                except:
                    print("Kaydedilirken bir hata oluştu.")
                    wait(2)
                    raise
                    
                    
            elif  menu_input == 888:
                mainDict = jsonLoader("form_data")
                
                form_list = []
                for i in mainDict["formNames"].keys():
                    form_list.append(i)
                    
                to_save_dict = {}
                                    
                while True:
                    print("|>===================================================================================================<|\n")
                    for i in mainDict["formNames"].keys():    
                        print(i + ") " + mainDict["formNames"][i])
                    print("\n|>===================================================================================================<|\n")
                    print("Tüm girilenleri kaydetip nöropsikolojik testlere dönmek için '888' giriniz. ")
                    userInput = numInput("Lütfen bir seçim yapınız: ")
                    if userInput in form_list:
                        printable_list, console_results, excelColNameDict = formMain(mainDict[userInput])
                        to_save_dict[userInput] = [printable_list, console_results, mainDict[userInput]["excelSheetName"], excelColNameDict]
                        continue
                    elif userInput == "888":
                        print("Nöropsikolojik testlere dönülüyor.\n")
                        if len(to_save_dict.keys()) > 0:
                            for i in to_save_dict.keys():
                                printable_list = to_save_dict[i][0]
                                console_results = to_save_dict[i][1]
                                excelSheetName = to_save_dict[i][2]
                                excelColNameDict = to_save_dict[i][3]
                                
                                if settings("output_excel"):
                                    excelWriter(settings("excel_path"), None, printable_list, excelSheetName, excelColNameDict)
                                   
                                if settings("output_csv"):
                                    csvWriter(settings("csv_path"), test_name, printable_list)
                                    #writes the printable_list in a CSV file
                                if settings("output_txt"):
                                    console_results = "\n===========================================\n" + console_results
                                    console_results = excelSheetName + "\n" + console_results
                                    console_results = console_results + "\n===========================================\n"
                                    txtWrite(settings("txt_path"), console_results)
                            #writes the console results on the txt file
                                
                        break
                        
                    else:
                        print("Lütfen varolan seçeneklerden birini seçiniz. ")
                        continue
                    
                continue
                    
                
            else:
                print("\nLütfen listede olan numaralardan giriniz.")
                continue
                #this is called if user enters a number higher than tests that are in the program
        except SystemExit:
            raise
        except:
            raise
            print("\nLütfen sadece sayı giriniz.")
            continue
            #DEATH PROTECTION TOME

def mainStartup(): #the mainStartup function
    print("""
================================================
 PsiNorm Persentil Hesaplayıcı - 1.6.11
================================================
""")
    timeGlobalization()
    guiPatientInfo()
    
    try:
        if None in [patient_admin, patient_ID, patient_name, patient_age, patient_sex, patient_edu]:
            print("Kayıt iptal edilmiş. Geri dönülüyor. ")
            return
        else:
            mainMenu()
    except NameError:
        print("Kayıt iptal edilmiş. Geri dönülüyor. ")
        return
    except:
        raise
        
    
#    while True:
#        
#        global patient_admin
#        patient_admin = hexaInput("Testi uygulayan kişi: ")
#             
#        global patient_ID
#        patient_ID = hexaInput("Hastanın kodu: ")
#    
#        global patient_name
#        patient_name = hexaInput("Hastanın ismi: ")
#        
#        global patient_age
#        patient_age = inputPatient_age()
#        global patient_sex
#        patient_sex = inputPatient_sex()
#        global patient_edu
#        patient_edu = inputPatient_edu()
#        #gets input for all the data program needs
#        
#        print("=======================================")    
#        user_input = input("Yukarıdaki bilgileri onaylıyor musunuz? (e)vet/(h)ayır: ")
#        #a step to ensure correct info was entered
#        if user_input == "E" or user_input == "e":
#            mainMenu()
#            #starts the program proper
#            break
#        else:
#            print("Program tekrar başlatılıyor.")
#            continue
#            #resets


def calcZscore(result_list, mean_list, sd_list):
    #finds out which SD interval patient result is in and orders it in a list
    z_score_list = []
    for i in range(len(result_list)):
        if sd_list[i] == 0:
            sd_list[i] = 0.00000001 #protects the program from failing
        z_score = ((result_list[i] - mean_list[i]) / sd_list[i])
        z_score_list.append(float("%.2f" % z_score))
    return z_score_list

def calcPercentile(z_score_list):
    #gets the Z score list, changes it into percentile list
    import math

    def percentile(z_score):
        return .5 * (math.erf(z_score / 2 ** .5) + 1)

    perc_list = []
    for i in range(len(z_score_list)):
        try:
            perc_temp = str("%.2f" % (100 * float(percentile(z_score_list[i]))))
            perc_list.append(perc_temp)
        except:
            print("Persentil hesaplanırken bir hata oluştu.")
            pass

    return perc_list
    
def outputPrintlist(result_list, z_score_list, z_score_verbal_list):
    #puts all the lists in their proper, more manageable order to print in CSV
    printable_list = []
    perc_list = calcPercentile(z_score_list)
    for y in range(len(result_list)):
        if not result_list[y] in ["999", 999, 999.0, "999.0"]:
            printable_list.append(result_list[y])
            if not float(z_score_list[y]) in [999.0, -999.0]:
                printable_list.append(z_score_list[y])
                printable_list.append(perc_list[y])
                printable_list.append(z_score_verbal_list[y])
            else:
                printable_list.append("N/A")
                printable_list.append("N/A")
                printable_list.append("N/A")
        else:
            printable_list.append("N/A")
            printable_list.append("N/A")
            printable_list.append("N/A")
            printable_list.append("N/A")
    return printable_list

def outputConsole_results(result_list, z_score_list, z_score_verbal_list):
    #gets the results ready to print onto the screen
    console_result = []
    perc_list = calcPercentile(z_score_list)
    
    for i in range(len(z_score_list)):
        if (result_list[i] != 999.00) and (999.00 != float(z_score_list[i])) and (-999.00 != float(z_score_list[i])):
            console_result.append("Hastanın puanı: " + str(result_list[i]) + " - " + 
            str(z_score_verbal_list[i]) + " Z skoru: " + str(z_score_list[i]) + " - Persentil: " + str(perc_list[i]))
            
        elif (result_list[i] != 999.00) and ((999.00 == float(z_score_list[i])) or (-999.00 == float(z_score_list[i]))):
            console_result.append("Hastanın puanı: " + str(result_list[i]) + " - Bu parametreye ait norm verisi yoktur.")
        
        else:
            console_result.append("Bu basamak uygulanmamış veya uygulanamamıştır.")

    return console_result


#def calcSD_less_better(z_score_list): 
#    #If lower score means better score, this is used to turn the SD floats into verbal results and lists it
#    z_score_verbal_list = []
#    
#    for i in range(len(z_score_list)):
#        if z_score_list[i] <= 1:
#            x = "Normal."
#        elif z_score_list[i] > 1 and z_score_list[i] <= 2:
#            x = "Hafif derecede bozulma."
#        elif z_score_list[i] > 2 and z_score_list[i] <= 3:
#            x = "Orta derecede bozulma."
#        elif z_score_list[i] > 3 and z_score_list[i] != 999:
#            x = "Ağır derecede bozulma." 
#        else:
#            x = "Bu grup için norm değeri bulunmamaktadır."
#        z_score_verbal_list.append(x)
#        
#    return z_score_verbal_list

def calcSD_less_better(z_score_list): #DUMMY VERSION UNTIL BETTER FIX, SAME WITH calcSD_more_better
    #If lower score means better score, this is used to turn the SD floats into verbal results and lists it
    z_score_verbal_list = []
    
    for i in range(len(z_score_list)):
        if z_score_list[i] >= -1:
            x = "Normal.*"
        elif -2 <= z_score_list[i] < -1:
            x = "Hafif derecede bozulma.*"
        elif -3 <= z_score_list[i] < -2:
            x = "Orta derecede bozulma.*"
        elif z_score_list[i] < -3 and z_score_list[i] != 999:
            x = "Ağır derecede bozulma.*" 
        else:
            x = "Bu grup için norm değeri bulunmamaktadır."
        z_score_verbal_list.append(x)
        
    return z_score_verbal_list



def calcSD_more_better(z_score_list):
#If higher score means better score, this is used to turn the z score floats into verbal results and lists it
    z_score_verbal_list = []
    for i in range(len(z_score_list)):
        if z_score_list[i] >= -1:
            x = "Normal."
        elif -2 <= z_score_list[i] < -1:
            x = "Hafif derecede bozulma."
        elif -3 <= z_score_list[i] < -2:
            x = "Orta derecede bozulma."
        elif z_score_list[i] < -3 and z_score_list[i] != 999:
            x = "Ağır derecede bozulma." 
        else:
            x = "Bu grup için norm değeri bulunmamaktadır."
        z_score_verbal_list.append(x)
    return z_score_verbal_list


def testWechsler():
    try:
        while True:
            try:        
                print("\n===================================\nWechsler zeka testi: ")
                            
                result_name = ["\nGenel bilgi: ", "\nYargılama: ", "\nAritmetik: ", "\nBenzerlik: ",
                                "\nSayı dizisi ", "\nKelime: ", "\nŞifre: ", "\nResim tamamlama: ",
                                "\nKüplerle desen: ", "\nResim düzenleme: ", "\nParça birleştirme: "]    
                               
                result_list = []               
                for i in range(len(result_name)):   
                    result_list.append(floInput(result_name[i]))
                #gets raw input from the user, these are test results and creates a list from them
                
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
            else:
                norm_exists = True
                break


        printable_list = []
        for y in range(len(result_list)):
            printable_list.append(result_list[y])
            
            
            valid_items_dict = {
                    0: 0, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0
                    }
               
            #For every scaled score, list of raw scores, min and max values. 
        raw_to_scaled_dict = {
            19: [[29], [27,28], [None], [26], [17], [78, 80], [87, 90], [None], [None], [None], [None]],
            18: [[28], [26], [None], [25], [None], [76, 77], [83, 86], [21], [None], [36], [44]],
            17: [[27], [25], [18], [24], [None], [74, 75], [79, 82], [None], [48], [35], [43]],
            16: [[26], [24], [17], [23], [16], [71, 73], [76, 78], [20], [47], [34], [42]],
            15: [[25], [23], [16], [22], [15], [67, 70], [72, 75], [None], [46], [33], [41]],
            14: [[23, 24], [22], [15], [21], [14], [63, 66], [69, 71], [19], [44, 45], [32], [40]],
            13: [[21, 22], [21], [14], [19, 20], [None], [59, 62], [66, 68], [18], [42, 43], [30, 31], [38, 39]],
            12: [[19, 20], [20], [13], [17, 18], [13], [54, 58], [62, 65], [17], [39, 41], [28, 29], [36, 37]],
            11: [[17, 18], [19], [12], [15, 16], [12], [47, 53], [58, 61], [15, 16], [35, 38], [26, 27], [34, 35]],
            10: [[15, 16], [17, 18], [11], [13, 14], [11], [40, 46], [52, 57], [14], [31, 34], [23, 25], [31, 33]],
            9:  [[13, 14], [15, 16], [10], [11, 12], [10], [32, 39], [47, 51], [12, 13], [28, 30], [20, 22], [28, 30]],
            8:  [[11, 12], [14], [9], [9, 10], [None], [26, 31], [41, 46], [10, 11], [25, 27], [18, 19], [25, 27]],
            7:  [[9, 10], [12, 13], [7, 8], [7, 8], [9], [22, 25], [35, 40], [8, 9], [21, 24], [15, 17], [22, 24]],
            6:  [[7, 8], [10, 11], [6], [5, 6], [8], [18, 21], [29, 34], [6, 7], [17, 20], [12, 14], [19, 21]],
            5:  [[5, 6], [8, 9], [5], [4], [None], [14, 17], [23, 28], [5], [13, 16], [9, 11], [15, 18]],
            4:  [[4], [6, 7], [4], [3], [7], [11, 13], [18, 22], [4], [10, 12], [8], [11, 14]],
            3:  [[3], [5], [3], [2], [None], [10], [15, 17], [3], [6, 9], [7], [8, 10]],
            2:  [[2], [4], [2], [1], [6], [9], [13, 14], [2], [3, 5], [6], [5, 7]],
            1:  [[1], [3], [1], [None], [4, 5], [8], [12], [1], [2], [5], [3, 4]],
            0:  [[0], [0, 2], [0], [0], [0, 3], [0, 7], [0, 11], [0], [0, 1], [0, 4], [0, 2]]         
         }
        
        scaled_dict = {}
        
        for raw_score_index in range(len(printable_list)):
            raw_score = printable_list[raw_score_index]
            for dictKey in range(20):
                to_check_list = raw_to_scaled_dict[dictKey][raw_score_index] 
                if len(to_check_list) == 2:
                    if to_check_list[0] <= raw_score <= to_check_list[1]:
                        scaled_dict[raw_score_index] = dictKey
                        valid_items_dict[raw_score_index] = 1
                        
                if len(to_check_list) == 1:
                    if to_check_list[0] != None:
                        if to_check_list[0] == raw_score:
                            scaled_dict[raw_score_index] = dictKey
                            valid_items_dict[raw_score_index] = 1
                            
        
        
        for i in scaled_dict.keys():
            printable_list[i] = scaled_dict[i]
            
        verb_items_list = [0, 1, 2, 3, 4, 5]
        perf_items_list = [6, 7, 8, 9, 10]
        verb_score = 0
        perf_score = 0
        verb_items = 0
        perf_items = 0
                
        for i in verb_items_list:
            if valid_items_dict[i] == 1:
                verb_score += printable_list[i]
                verb_items += 1
        
        if verb_items != 0:
            verb_score = (verb_score) * (6 / verb_items)
        else:
            verb_score = 0
        
        for i in perf_items_list:
            if valid_items_dict[i] == 1:
                perf_score += printable_list[i]
                perf_items += 1
                
        if perf_items != 0:
            perf_score = (perf_score) * (5 / perf_items)
        else:
            perf_score = 0
        
        total_score = verb_score + perf_score
        
        IQ_dict_list = [
                        {
                        "age": [16, 17], 
                        "120": [75, 64, 137], 
                        "110": [65, 57, 120],
                        "90": [44, 41, 86],
                        "80": [34, 34, 69],
                        "70": [24, 26, 53],
                        },
                        
                        {
                        "age": [18, 19], 
                        "120": [78, 65, 141], 
                        "110": [67, 57, 123],
                        "90": [47, 42, 90],
                        "80": [37, 34, 73],
                        "70": [27, 27, 56],
                        },
                                
                        {
                        "age": [20, 24], 
                        "120": [80, 66, 143], 
                        "110": [70, 58, 127],
                        "90": [49, 43, 93],
                        "80": [39, 35, 76],
                        "70": [29, 28, 59],
                        },
                                
                        {
                        "age": [25, 34], 
                        "120": [81, 65, 144], 
                        "110": [71, 57, 127],
                        "90": [51, 42, 93],
                        "80": [41, 34, 76],
                        "70": [30, 27, 60],
                        },
                            
                        {
                        "age": [35, 44], 
                        "120": [81, 61, 140], 
                        "110": [70, 54, 123],
                        "90": [50, 39, 89],
                        "80": [40, 31, 72],
                        "70": [30, 23, 56],
                        },
                                
                        {
                        "age": [45, 54], 
                        "120": [78, 56, 132], 
                        "110": [68, 49, 116],
                        "90": [48, 34, 82],
                        "80": [38, 26, 65],
                        "70": [28, 18, 48],
                        },
                                
                        {
                        "age": [55, 64],
                        "120": [76, 53, 126],
                        "110": [66, 45, 109],
                        "90": [46, 30, 76],
                        "80": [35, 22, 59],
                        "70": [25, 14, 42],
                        },
                                
                        {
                        "age": [65, 69],
                        "120": [74, 50, 121],
                        "110": [64, 42, 105],
                        "90": [44, 27, 71],
                        "80": [33, 19, 54],
                        "70": [23, 12, 37],
                        },
                                
                        {
                        "age": [70, 74],
                        "120": [68, 45, 110],
                        "110": [58, 37, 94],
                        "90": [38, 22, 60],
                        "80": [27, 14, 43],
                        "70": [17, 7, 26],
                        },
                                
                        {
                        "age": [75, 150],
                        "120": [64, 40, 102],
                        "110": [54, 32, 85],
                        "90": [34, 17, 52],
                        "80": [24, 9, 35],
                        "70": [14, 2, 18],
                        }
                        ]
            
        try:
            result_values = ["N/A", "N/A", "N/A"]
            testing_values = [verb_score, perf_score, total_score]
            #result_values = [None, None, None]
            for IQ_dict in IQ_dict_list:
                if IQ_dict["age"][0] <= patient_age <= IQ_dict["age"][1]:
                    for i in range(3):
                        if IQ_dict["120"][i] <= testing_values[i]:
                            result_values[i] = "Üstün zeka - (120 ve üzeri IQ) "
                        elif IQ_dict["110"][i] <= testing_values[i]:
                            result_values[i] = "Parlak zeka - (110-119 IQ) "
                        elif IQ_dict["90"][i] <= testing_values[i]:
                            result_values[i] = "Normal zeka - (90-109 IQ) "
                        elif IQ_dict["80"][i] <= testing_values[i]:
                            result_values[i] = "Donuk normal zeka - (80-89 IQ)"
                        elif IQ_dict["70"][i] <= testing_values[i]:
                            result_values[i] = "Sınırda zeka geriliği - (70-79 IQ)"
                        elif testing_values[i] < IQ_dict["70"][i]:
                            result_values[i] = "Mental retardasyon - (69 ve altı IQ)"
                            
                        else:
                            print("Bir şeyler yanlış gitti. Lütfen programcı ile bağlantıya geçiniz. ")
                            result_values = ["N/A", "N/A", "N/A"]
                            
    
        except:
             raise
                
                
        for i in range(3):
            printable_list.append(testing_values[i])
            printable_list.append(result_values[i])
            
        
        test_name = 'wechsler_data.csv' #test datasının toplanacağı csv dosyasını belirtiyor
        #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)
        
        
        outputConsole_results = []
        for i in range(len(result_list)):
            outputConsole_results.append("Hastanın standart puanı: " + str(printable_list[i]))
       
        console_results = "==================================\nWechsler zeka testinin sonuçları: "
        
        
        for i in range(len(result_name)):
            console_results = console_results + (result_name[i] + str(outputConsole_results[i]))
            
        console_results = console_results + "\nSözel standart puan: " + str(verb_score) + " - " + str(result_values[0])
        console_results = console_results + "\nPerformans standart puan: " + str(perf_score) + " - " + str(result_values[1])
        console_results = console_results + "\nToplam standart puan: " + str(total_score) + " - " + str(result_values[2])
        

        console_results = console_results + ("\n==================================")
        
        if norm_exists:
            print(console_results)
            return [test_name, printable_list, console_results]
        else:
            return [test_name, printable_list, console_results]
        
    except:
        print("Wechsler zeka testini değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return
    
"""    
def testWechslerTESTVERSION():
    try:
        while True:
            try:        
                print("\n===================================\nWechsler zeka testi:\n")
                            
                result_name = ["\nGenel bilgi: ", "\nYargılama: ", "\nAritmetik: ", "\nBenzerlik: ",
                                "\nSayı dizisi toplam: ", "\nSayı dizisi ileri: ", "\nSayı dizisi geri: ",
                               "\nKelime: ", "\nŞifre: ", "\nResim tamamlama: ", "\nKüplerle desen: ",
                               "\nResim düzenleme: ", "\nParça birleştirme: "]    
                               
                result_list = []               
                for i in range(len(result_name)):   
                    result_list.append(floInput(result_name[i])))
                #gets raw input from the user, these are test results and creates a list from them
                

                
                verbal_IQ = 0
                perf_IQ = 0
                for i in range(5):
                    if result_list[i] != 999:
                        verbal_IQ = verbal_IQ + result_list[i]

                existingValues = 0
                for i in range(6):
                    
                verbal_IQ = (verbal_IQ / 5) * 6
                    
                for i in range(3):
                    if result_list[i+8] != 999:
                        perf_IQ = perf_IQ + result_list[i+8]
                        
                if result_list[12] != 999:
                    perf_IQ = perf_IQ + result_list[12]
                    
                perf_IQ = (perf_IQ / 4) * 5
                
                sum_IQ = verbal_IQ + perf_IQ
                
                if sum_IQ > 110:
                    sum_IQ_res = "Normalüstü"
                elif 90 <= sum_IQ <= 110:
                    sum_IQ_res = "Normal"
                elif 80 <= sum_IQ <= 89:
                    sum_IQ_res = "Donuk normal"
                elif 70 <= sum_IQ <= 79:
                    sum_IQ_res = "Sınırda zeka geriliği"
                elif 50 <= sum_IQ <= 69:
                    sum_IQ_res = "Hafif zeka geriliği"
                elif 35 <= sum_IQ <= 49:
                    sum_IQ_res = "Orta zeka geriliği"
                elif sum_IQ <= 34:
                    sum_IQ_res = "Ağır zeka geriliği"
                    
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
            else:
                norm_exists = True
                break

#Bütün SD'lerin üst sınır değerleri burada sıralı wech_1'den wech_11'e
        wech_m2_list = [4, 6, 4, 3, 7, 2, 0, 11, 18, 4, 10, 8, 11]
        wech_m1_list = [9, 12, 7, 7, 9, 4, 2, 22, 35, 8, 21, 15, 22] 
        wech_0_list = [15.5, 17.5, 11, 13.5, 11, 6, 4, 43, 54.5, 14, 32.5, 24, 32]
        wech_1_list = [23, 22, 15, 21, 14, 8, 6, 63, 69, 19, 44, 32, 40]
        wech_2_list = [27, 25, 18, 24, 17, 10, 8, 74, 79, 21, 48, 35, 43]        
        z_score_verbal_list = []          
        
        sd_m2 = "-2SD'den küçük, ağır bozulma"
        sd_m1 = "-1SD ve -2SD aralığında, hafif bozulma"
        sd_m0 = "0SD ve -1SD aralığında, normal"
        sd_0 = "0SD ve 1SD aralığında, normal"
        sd_1 = "1SD ve 2SD aralığında, normal"
        sd_2 = "2SD üzerisinde, normal"
 
#Girilen sonuçlar hangi SD aralığında diye sırayla deniyor tüm aralıkları
        for i in range(len(result_list)):
            if result_list[i] < wech_m2_list[i]:
                z_score_verbal_list.append(sd_m2)
            elif wech_m2_list[i] <= result_list[i] < wech_m1_list[i]:
                z_score_verbal_list.append(sd_m1)
            elif wech_m1_list[i] <= result_list[i] < wech_0_list[i]:
                z_score_verbal_list.append(sd_m0)
            elif wech_0_list[i] <= result_list[i] < wech_1_list[i]:
                z_score_verbal_list.append(sd_0)
            elif wech_1_list[i] <= result_list[i] < wech_2_list[i]:
                z_score_verbal_list.append(sd_1)
            elif wech_2_list[i] <= result_list[i]:
                z_score_verbal_list.append(sd_2)

        printable_list = []
        for y in range(len(result_list)):
            printable_list.append(result_list[y])
            printable_list.append(z_score_verbal_list[y])
            
        printable_list = printable_list + [verbal_IQ, perf_IQ, sum_IQ, sum_IQ_res]
        
                
        test_name = 'wechsler_data.csv' #test datasının toplanacağı csv dosyasını belirtiyor
        #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)
        
        
        outputConsole_results = []
        for i in range(len(result_list)):
            outputConsole_results.append("Hastanın puanı: " + str(result_list[i]) + ", " + str(z_score_verbal_list[i]))
       
        console_results = "\n===================================\nWechsler zeka testinin sonuçları: "
        
        result_name = result_name + ["\nSözel IQ: ", "\nPerformans IQ: ", "\nToplam IQ: "]
        outputConsole_results = outputConsole_results + ([verbal_IQ, perf_IQ, str(sum_IQ) + ", " + sum_IQ_res])
        
        for i in range(len(result_name)):
            console_results = console_results + (result_name[i] + str(outputConsole_results[i]))    
        

        console_results = console_results + ("\n===================================\n")
        
        if norm_exists:
            #txtWrite(patient_admin, patient_ID, console_results)
            print(console_results)
            return [test_name, printable_list, console_results]
        else:
            #txtWrite(patient_admin, patient_ID, ("Wechsler: Bu grup için norm mevcut değildir.\n" + console_results))
            return [test_name, printable_list, console_results]
        
    except:
        print("Wechsler zeka testini değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return
    
"""
          
def testWechslerSayi():
    try:
        while True:
            try:        
                print("\n===================================\nWechsler zeka testi:\n")

                result_name = ["\nSayı dizisi toplam: ", "\nSayı dizisi ileri: ", "\nSayı dizisi geri: "]    
                               
                result_list = []               
                for i in range(len(result_name)):   
                    result_list.append(floInput(result_name[i]))
                #gets raw input from the user, these are test results and creates a list from them

            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
            else:
                norm_exists = True
                break

#Bütün SD'lerin üst sınır değerleri burada sıralı wech_1'den wech_11'e

        wech_m2_list = [7, 2, 0]
        wech_m1_list = [ 9, 4, 2] 
        wech_0_list = [11, 6, 4]
        wech_1_list = [14, 8, 6]
        wech_2_list = [17, 10, 8]
        
        z_score_verbal_list = []          
        
        sd_m2 = "-2SD'den küçük, ağır bozulma"
        sd_m1 = "-1SD ve -2SD aralığında, hafif bozulma"
        sd_m0 = "0SD ve -1SD aralığında, normal"
        sd_0 = "0SD ve 1SD aralığında, normal"
        sd_1 = "1SD ve 2SD aralığında, normal"
        sd_2 = "2SD üzerisinde, normal"
        
#Girilen sonuçlar hangi SD aralığında diye sırayla deniyor tüm aralıkları
        for i in range(len(result_list)):
            if result_list[i] < wech_m2_list[i]:
                z_score_verbal_list.append(sd_m2)
            elif wech_m2_list[i] <= result_list[i] < wech_m1_list[i]:
                z_score_verbal_list.append(sd_m1)
            elif wech_m1_list[i] <= result_list[i] < wech_0_list[i]:
                z_score_verbal_list.append(sd_m0)
            elif wech_0_list[i] <= result_list[i] < wech_1_list[i]:
                z_score_verbal_list.append(sd_0)
            elif wech_1_list[i] <= result_list[i] < wech_2_list[i]:
                z_score_verbal_list.append(sd_1)
            elif wech_2_list[i] <= result_list[i]:
                z_score_verbal_list.append(sd_2)
                
        printable_list = []
        for y in range(len(result_list)):
            printable_list.append(result_list[y])
            printable_list.append(z_score_verbal_list[y])
                
        test_name = 'wechsler_onlynumber_data.csv' #test datasının toplanacağı csv dosyasını belirtiyor
        #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)
        
        outputConsole_results = []
        for i in range(len(result_list)):
            outputConsole_results.append("Hastanın puanı: " + str(result_list[i]) + ", " + str(z_score_verbal_list[i]))
       
        console_results = "==================================\nWechsler zeka testinin sonuçları: "
        
        for i in range(len(result_name)):
            console_results = console_results + (result_name[i] + str(outputConsole_results[i]))    
        
        console_results = console_results + ("\n==================================")
        
        if norm_exists:
            #txtWrite(patient_admin, patient_ID, console_results)
            print(console_results)
            return [test_name, printable_list, console_results]
        else:
            #txtWrite(patient_admin, patient_ID, ("Wechsler: Bu grup için norm mevcut değildir.\n" + console_results))
            return [test_name, printable_list, console_results]
        
    except:
        print("Wechsler zeka testini değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return


def testSf():
    try:
        while True:
            try:
                print("\n===================================\nSemantik akıcılık: \n")
                result_name = ["\nHayvan için: ", "\nİnsan için: ", "\nİnsan-Hayvan için: "]    
                
                result_list = []               
                for i in range(len(result_name)):   
                    result_list.append(floInput(result_name[i]))
                #gets raw input from the user, these are test results and creates a list from them
                 
                mean_list = []
                sd_list = []
                for i in range(len(result_list)):
                    mean_list.append(result_list[i]-999)
                    sd_list.append(1) 
         
                norm_exists = True
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
            else:
                break

        if patient_sex == "Kadın": 
            if patient_edu <= 8:
                if 15 <= patient_age <= 24:
                    mean_list = [14.20, 22.30, 16.50] #Hayvan, İnsan, Hayvan-insan için ortalamalar sıralı liste
                    sd_list = [3.82, 5.21, 3.17] #Hayvan, İnsan, Hayvan-insan için standart sapma sıralı liste        
                    
                elif 25 <= patient_age <= 34:
                    mean_list = [15.40, 24.30, 16.30]
                    sd_list = [3.13, 6.67, 3.16]
        
                elif 35 <= patient_age <= 44:
                    mean_list = [16.83, 25.42, 16.83]
                    sd_list = [3.27, 4.32, 4.00]
        
                elif 45 <= patient_age <= 54:
                    mean_list = [16.83, 25.92, 17.50]
                    sd_list = [3.19, 5.16, 4.70]
        
                elif 55 <= patient_age <= 64:
                    mean_list = [16.91, 22.09, 14.10]
                    sd_list = [5.63, 5.05, 4.32]
                
                elif 65 <= patient_age:
                    mean_list = [17.36, 21.91, 14.91]
                    sd_list = [2.50, 6.27, 4.41]
                else:
                    print("Bu yaş grubu için norm mevcut değildir.")
                    norm_exists = False
                    
            elif 9 <= patient_edu <= 11:
                if 15 <= patient_age <= 24:
                    mean_list = [21.70, 32.70, 21.60]
                    sd_list = [4.92, 5.88, 4.40]  
                    
                elif 25 <= patient_age <= 34:
                    mean_list = [29.64, 20.45, 21.45]
                    sd_list = [4.80, 7.10, 7.27]
        
                elif 35 <= patient_age <= 44:
                    mean_list = [20.00, 27.50, 19.20]
                    sd_list = [3.92, 5.15, 3.36]
        
                elif 45 <= patient_age <= 54:
                    mean_list = [20.10, 27.00, 20.10]
                    sd_list = [4.43, 6.16, 5.11]
        
                elif 55 <= patient_age <= 64:
                    mean_list = [21.83, 29.83, 20.33]
                    sd_list = [6.28, 9.81, 6.44]
                
                elif 65 <= patient_age:
                    mean_list = [17.90, 23.22, 15.44]
                    sd_list = [4.57, 7.69, 4.82]
                else:
                    print("Bu yaş grubu için norm mevcut değildir.")
                    norm_exists = False
                    
            elif 12 <= patient_edu:
                if 15 <= patient_age <= 24:
                    mean_list = [21.82, 28.63, 21.27]
                    sd_list = [3.90, 5.20, 3.10]  
                    
                elif 25 <= patient_age <= 34:
                    mean_list = [21.50, 32.83, 21.83]
                    sd_list = [4.68, 8.10, 3.41]
        
                elif 35 <= patient_age <= 44:
                    mean_list = [24.00, 31.21, 22.43]
                    sd_list = [4.24, 5.49, 5.15]
        
                elif 45 <= patient_age <= 54:
                    mean_list = [21.90, 30.50, 19.80]
                    sd_list = [3.80, 8.20, 5.92]
        
                elif 55 <= patient_age <= 64:
                    mean_list = [21.82, 27.82, 19.18]
                    sd_list = [5.55, 6.42, 4.02]
                
                elif 65 <= patient_age:
                    mean_list = [21.30, 27.80, 20.40]
                    sd_list = [8.00, 9.85, 5.15]
                else:
                    print("Bu yaş grubu için norm mevcut değildir.")
                    norm_exists = False
            else:
                print("Bu eğitim grubu için norm mevcut değildir.")
                norm_exists = False
                
        elif patient_sex == "Erkek": 
            if patient_edu <= 8:
                if 15 <= patient_age <= 24:
                    mean_list = [17.20, 24.60, 18.60]
                    sd_list = [2.49, 3.34, 1.65]
                    
                elif 25 <= patient_age <= 34:
                    mean_list = [19.18, 24.18, 18.27]
                    sd_list = [5.53, 6.24, 4.08]
        
                elif 35 <= patient_age <= 44:
                    mean_list = [17.00, 23.30, 17.80]
                    sd_list = [4.62, 6.46, 2.97]
        
                elif 45 <= patient_age <= 54:
                    mean_list = [18.60, 22.70, 17.70]
                    sd_list = [3.10, 5.21, 3.10]
        
                elif 55 <= patient_age <= 64:
                    mean_list = [15.10, 19.80, 13.30]
                    sd_list = [5.76, 5.88, 4.55]
                
                elif 65 <= patient_age:
                    mean_list = [16.50, 17.10, 13.20]
                    sd_list = [5.30, 5.11, 3.12]
                else:
                    print("Bu yaş grubu için norm mevcut değildir.")
                    norm_exists = False
                    
            elif 9 <= patient_edu <= 11:
                if 15 <= patient_age <= 24:
                    mean_list = [19.85, 28.46, 20.92]
                    sd_list = [6.04, 8.55, 3.62]  
                    
                elif 25 <= patient_age <= 34:
                    mean_list = [21.82, 28.00, 20.36]
                    sd_list = [4.77, 5.80, 4.37]
        
                elif 35 <= patient_age <= 44:
                    mean_list = [20.30, 27.10, 18.50]
                    sd_list = [5.25, 4.33, 4.22]
        
                elif 45 <= patient_age <= 54:
                    mean_list = [21.10, 23.27, 17.64]
                    sd_list = [4.44, 3.82, 3.35]
        
                elif 55 <= patient_age <= 64:
                    mean_list = [18.45, 22.55, 18.36]
                    sd_list = [4.61, 6.74, 4.95]
                
                elif 65 <= patient_age:
                    mean_list = [16.80, 19.40, 15.80]
                    sd_list = [4.66, 3.20, 2.62]
                else:
                    print("Bu yaş grubu için norm mevcut değildir.")
                    norm_exists = False
            
            elif 12 <= patient_edu:
                if 15 <= patient_age <= 24:
                    mean_list = [23.04, 31.91, 23.35]
                    sd_list = [3.62, 7.60, 5.79]
                    
                elif 25 <= patient_age <= 34:
                    mean_list = [20.79, 27.50, 20.57]
                    sd_list = [3.29, 6.47, 4.52]
        
                elif 35 <= patient_age <= 44:
                    mean_list = [20.40, 26.20, 20.70]
                    sd_list = [3.89, 4.44, 3.34]
        
                elif 45 <= patient_age <= 54:
                    mean_list = [21.30, 26.57, 21.86]
                    sd_list = [7.10, 6.81, 5.05]
        
                elif 55 <= patient_age <= 64:
                    mean_list = [20.60, 24.10, 17.80]
                    sd_list = [3.31, 3.90, 2.49]
                
                elif 65 <= patient_age:
                    mean_list = [19.60, 21.90, 15.80]
                    sd_list = [4.97, 8.54, 5.31]
                else:
                    print("Bu yaş grubu için norm mevcut değildir.")
                    norm_exists = False
            
            else:
                print ("Bu eğitim grubu için norm mevcut değildir.")
                norm_exists = False
        else:
            print("Bu cinsiyet için norm mevcut değildir.")
            norm_exists = False
        
        z_score_list = calcZscore(result_list, mean_list, sd_list)
        z_score_verbal_list = calcSD_more_better(z_score_list)
        printable_list = outputPrintlist(result_list, z_score_list, z_score_verbal_list)

        test_name = 'semantic_fluency_data.csv' #test datasının toplanacağı csv dosyasını belirtiyor
        #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)
            
        console_results = "==================================\nSemantik akıcılık testinin sonuçları: "
        for i in range(len(mean_list)):
            console_results = console_results + (result_name[i] + str(outputConsole_results(result_list, z_score_list, z_score_verbal_list)[i]))            
        console_results = console_results + ("\n==================================")
        
        if norm_exists:
            #txtWrite(patient_admin, patient_ID, console_results)
            print(console_results)
            return [test_name, printable_list, console_results]
        else:
            #txtWrite(patient_admin, patient_ID, ("Semantik akıcılık: Bu grup için norm mevcut değildir.\n" + console_results))
            return [test_name, printable_list, console_results]
        
    except:
        print("Semantik akıcılık testini değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return
        
def testMmt():
    while True:
        try:
            score = int(numInput("\nMini mental test puanı: "))
            result_name = "MMT test: "
            if score < 0:
                print("Lütfen doğru aralıkta değer giriniz.")
                continue
              
            if patient_edu < 5:
                if score < 18:
                    verbal_result = "Eşik altı değer"   
                    
                elif 18 <= score <= 19:
                    verbal_result = "Sınır değer"
                    
                elif 19 < score:
                    verbal_result = "Normal"
                    
            elif 5 <= patient_edu:
                if score < 22:
                    verbal_result =  "Eşik altı değer"
                    
                elif 22 <= score <= 23:
                    verbal_result =  "Sınır değer"
                    
                elif 23 < score:
                    verbal_result = "Normal"
                    
            if score == 999:
                verbal_result = "Bu basamak uygulanmamış veya uygulanamamıştır."
                
        except SystemExit:
            raise
        except:
            print("Lütfen yalnızca sayı giriniz")
            continue
        
        else:
            break
            
    console_results = "==================================\n"
    console_results = console_results + (result_name + str(score) + ", " +verbal_result)        
    printable_list = [score, verbal_result]            
    console_results = console_results + "\n=================================="
    test_name = 'testMmt.csv' #test datasının toplanacağı csv dosyasını belirtiyor
    #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)
    #txtWrite(patient_admin, patient_ID, console_results)
    print(console_results)        
    return [test_name, printable_list, console_results]
        
        

def testEcr(): #enhanced cued recall/Artırılmış ipuçlu hatırlama    
    while True:
        try:
            score = int(numInput("\nArtırılmış ipuçlu hatırlama: "))
            result_name = "Artırılmış ipuçlu hatırlama: "
            
            if score < 0:
                print("Lütfen doğru aralıkta değer giriniz.")
                continue
            
            if score < 41:
                verbal_result = "Eşik altı değer"
                
            elif 41 <= score <= 42:
                verbal_result = "Sınır değer"
                
            elif 42 < score and score != 999:
                verbal_result = "Normal"
                
            else:
                verbal_result = "Bu basamak uygulanmamış veya uygulanamamıştır."
                
        except SystemExit:
            raise
        except:
            print("Lütfen yalnızca sayı giriniz")
            continue
        
        else:
            break         
        
    printable_list = [score, verbal_result]
    console_results = "==================================\n"
    
    console_results = console_results + (result_name + str(score) + ", " +verbal_result)      
    console_results = console_results + "\n=================================="
    test_name = 'testEcr.csv' #test datasının toplanacağı csv dosyasını belirtiyor
    #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)
    #txtWrite(patient_admin, patient_ID, console_results)
    print(console_results)        
    return [test_name, printable_list, console_results]

                
def testMoca():
    while True:
        try:
            score = int(numInput("\nMOCA: "))
            result_name = "MOCA: "
            if score < 0:
                print("Lütfen doğru aralıkta değer giriniz.")
                continue
            
            if score < 21:
                verbal_result = "Eşik altı değer"
                
            elif 21 <= score and score != 999:
                verbal_result = "Normal"
                
            elif score == 999:
                verbal_result = "Bu basamak uygulanmamış veya uygulanamamıştır."
                
        except SystemExit:
            raise
        except:
            print("Lütfen yalnızca sayı giriniz")
            continue
        
        else:
            break    
                    
    printable_list = [score, verbal_result]
    console_results = "==================================\n"
    
    console_results = console_results + (result_name + str(score) + ", " +verbal_result)
    console_results = console_results + "\n=================================="
    test_name = 'testMoca.csv' #test datasının toplanacağı csv dosyasını belirtiyor
    #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)
    #txtWrite(patient_admin, patient_ID, console_results)
    print(console_results)
    return [test_name, printable_list, console_results]

        
def testCd(): #Clock drawing test
    try:
        while True:
            try:
                print("\n===================================\nSaat çizme: ")
                result_name = ["\nSkor: "]  
                #prints user interface
                
                result_list = []               
                for i in range(len(result_name)):
                    while True:
                        try:
                            x = floInput(result_name[i])
                        except:
                            print("Lütfen yalnızca sayı giriniz.")
                            continue
                        if 0 <= x <= 4:
                            break
                        elif x == 999:
                            break
                        else:
                            print("Lütfen 0-4 arasında bir değer giriniz.")
                            continue
                    result_list.append(x)
                #gets raw input from the user, these are test results
                    
                mean_list = []
                sd_list = []
                #creates a list for the test results, and empty ones 
                #for the means and standard deviations
                for i in range(len(result_list)):
                    mean_list.append(result_list[i]-999)
                    sd_list.append(1)            
                    #makes it so that it prints out 999 if there's 
                    #no norm calculated for that group
                norm_exists = True
                #It assumes that there's a norm for every group
                #if none exists, it changes to False
                
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
                #"Only enter numbers", and then resets the function
            
            else:
                break
            
            #tries to get user input, makes sure it's correct input
            
            """Following are lists of means and standard deviations,
            of the specified age, education and sex, where it applies.
            It's in a way that it corresponds to order of the result_list."""
            
        if patient_edu <= 5:
            if 50 <= patient_age <= 54:
                mean_list = [4.00]
                sd_list = [0]
                
            elif 55 <= patient_age <= 59:
                mean_list = [3.70]
                sd_list = [0.65]
            
            elif 60 <= patient_age <= 64:
                mean_list = [3.53]
                sd_list = [0.61]
            
            elif 65 <= patient_age <= 69:
                mean_list = [3.70]
                sd_list = [0.47]
            
            elif 70 <= patient_age <= 74:
                mean_list = [3.15]
                sd_list = [0.75]
            
            elif 75 <= patient_age <= 79:
                mean_list = [3.00]
                sd_list = [0.35]         
            
            elif 80 <= patient_age:
                mean_list = [2.44]
                sd_list = [0.51]
                
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False     
                #"No norm exists for the group", sets it to False
                
        elif 6 <= patient_edu <= 11:
            if 50 <= patient_age <= 54:
                mean_list = [4.00]
                sd_list = [0]
                
            elif 55 <= patient_age <= 59:
                mean_list = [4.00]
                sd_list = [0] 
            
            elif 60 <= patient_age <= 64:
                mean_list = [3.79]
                sd_list = [0.42]
            
            elif 65 <= patient_age <= 69:
                mean_list = [3.59]
                sd_list = [0.51]
            
            elif 70 <= patient_age <= 74:
                mean_list = [3.56]
                sd_list = [0.51]
            
            elif 75 <= patient_age <= 79:
                mean_list = [3.50]
                sd_list = [0.52]         
            
            elif 80 <= patient_age:
                mean_list = [2.75]
                sd_list = [0.45]
                
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False     
                #"No norm exists for the group", sets it to False
                
                
        elif 12 <= patient_edu:
            if 50 <= patient_age <= 54:
                mean_list = [4.00]
                sd_list = [0]
                
            elif 55 <= patient_age <= 59:
                mean_list = [3.95]
                sd_list = [0.22]
            
            elif 60 <= patient_age <= 64:
                mean_list = [4.00]
                sd_list = [0]
            
            elif 65 <= patient_age <= 69:
                mean_list = [3.85]
                sd_list = [0.37]
            
            elif 70 <= patient_age <= 74:
                mean_list = [3.75]
                sd_list = [0.45]
            
            elif 75 <= patient_age <= 79:
                mean_list = [3.53]
                sd_list = [0.51]    
            
            elif 80 <= patient_age:
                mean_list = [3.00]
                sd_list = [0.22]
                
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False     
                #"No norm exists for the group", sets it to False
                
        else:
            print("Bu eğitim grubu için norm mevcut değildir.")
            norm_exists = False  
            #"No norm exists for the group", sets it to False

        z_score_list = calcZscore(result_list, mean_list, sd_list)
        #it calculated the patient's SD interval as a float using the results, means and the SD
        z_score_verbal_list = calcSD_more_better(z_score_list)
        #Makes a verbal list to input to the users
        printable_list = outputPrintlist(result_list, z_score_list, z_score_verbal_list)
        #creates a list to be put into a CSV file

        test_name = 'clock_drawing_data.csv' #declares name of the CSV file to save the data in
        #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)
        #writes the printable_list in a CSV file
        
        console_results = "==================================\nSaat Çizme Testi sonuçları:"
        for i in range(len(mean_list)):
            console_results = console_results + (result_name[i] + str(outputConsole_results(result_list, z_score_list, z_score_verbal_list)[i]))            
        console_results = console_results + ("\n==================================")
        #prints the list using the range of the list and makes it more appealing for the user (probably should've used a decorator here)
        
        if norm_exists:
            #txtWrite(patient_admin, patient_ID, console_results)
            print(console_results)
            #creates a patient report for the physician and prints it out for the user
            return [test_name, printable_list, console_results]
        else:
            #txtWrite(patient_admin, patient_ID, ("Saat çizme: Bu grup için norm mevcut değildir.\n"+console_results))
            return [test_name, printable_list, console_results]
    
    except:
        print("Saat çizmeyi değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return
        #saves the program from fiery death

   
def testVf():
    try:
        while True:
            try:
                print("\n===================================\nSözel akıcılık: \n")
                result_name = ["\n'S' harfi için: " , "\n'A' harfi için: ", "\n'Z' harfi için: "]          
                
                result_list = []               
                for i in range(len(result_name)):   
                    result_list.append(floInput(result_name[i]))
                #gets raw input from the user, these are test results and creates a list from them
                  
                mean_list = []
                sd_list = []
                for i in range(len(result_list)):
                    mean_list.append(result_list[i]-999)
                    sd_list.append(1)      
        
                norm_exists = True
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
            else:
                break
            
        if patient_edu <= 8:
            if 15 <= patient_age <= 24:
                mean_list = [7, 6.45, 4.25] #[S, A, Z] için ortalamalar sıralı liste
                sd_list = [3.49, 3.15, 2.24] #[S, A, Z] için standart sapma sıralı liste        
                
            elif 25 <= patient_age <= 34:
                mean_list = [6.81, 6.38, 4.00]
                sd_list = [4.15, 2.48, 2.45]
    
            elif 35 <= patient_age <= 44:
                mean_list = [8.73, 7.23, 4.95]
                sd_list = [5.37, 4.12, 2.36]
    
            elif 45 <= patient_age <= 54:
                mean_list = [7.73, 7.05, 6.23]
                sd_list = [2.95, 3.08, 2.51]
    
            elif 55 <= patient_age <= 64:
                mean_list = [7.71, 6.29, 5.24]
                sd_list = [3.72, 3.02, 2.53]
            
            elif 65 <= patient_age:
                mean_list = [8.95, 7.00, 6.19]
                sd_list = [3.58, 4.07, 2.82]
            else:
                print("Bu yaş grubu için norm mevcut değildir.")
                norm_exists = False
                
        elif 9 <= patient_edu <= 11:
            if 15 <= patient_age <= 24:
                mean_list = [11.13, 11.43, 6.04]
                sd_list = [3.45, 5.77, 3.08]        
                
            elif 25 <= patient_age <= 34:
                mean_list = [11.45, 9.55, 7.73]
                sd_list = [4.94, 3.76, 3.21]
    
            elif 35 <= patient_age <= 44:
                mean_list = [11.50, 11.35, 7.35]
                sd_list = [4.97, 4.04, 3.20]
    
            elif 45 <= patient_age <= 54:
                mean_list = [11.38, 9.76, 7.67]
                sd_list = [3.99, 4.85, 4.36]
    
            elif 55 <= patient_age <= 64:
                mean_list = [11.70, 10.57, 7.96]
                sd_list = [6.01, 6.13, 4.61]
            
            elif 65 <= patient_age:
                mean_list = [9.68, 9.58, 6.89]
                sd_list = [3.67, 4.07, 2.88]
            else:
                print("Bu yaş grubu için norm mevcut değildir.")
                norm_exists = False
        
        elif 12 <= patient_edu:
            if 15 <= patient_age <= 24:
                mean_list = [15.82, 15.11, 10.87]
                sd_list = [5.10, 4.80, 3.49]
                
            elif 25 <= patient_age <= 34:
                mean_list = [15.85, 13.08, 8.81]
                sd_list = [4.03, 3.99, 2.40]
    
            elif 35 <= patient_age <= 44:
                mean_list = [15.05, 14.29, 10.00]
                sd_list = [4.75, 5.43, 3.22]
    
            elif 45 <= patient_age <= 54:
                mean_list = [13.13, 13.88, 9.54]
                sd_list = [4.38, 4.62, 3.89]
    
            elif 55 <= patient_age <= 64:
                mean_list = [13.10, 11.81, 8.33]
                sd_list = [3.52, 4.01, 2.42]
            
            elif 65 <= patient_age:
                mean_list = [13.75, 12.80, 9.25]
                sd_list = [5.39, 4.76, 3.54]
            else:
                print("Bu yaş grubu için norm mevcut değildir.")
                norm_exists = False
        
        else:
            print("Bu eğitim grubu için norm mevcut değildir.")
            norm_exists = False
        
        z_score_list = calcZscore(result_list, mean_list, sd_list)
        z_score_verbal_list = calcSD_more_better(z_score_list)
        printable_list = outputPrintlist(result_list, z_score_list, z_score_verbal_list)

        test_name = 'verbal_fluency_data.csv' #test datasının toplanacağı csv dosyasını belirtiyor
        #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)
        
        console_results = "==================================\nSözel akıcılık testinin sonuçları: "
        for i in range(len(mean_list)):
            console_results = console_results + (result_name[i] + str(outputConsole_results(result_list, z_score_list, z_score_verbal_list)[i]))            
        console_results = console_results + ("\n==================================")
        

        if norm_exists:      
            #txtWrite(patient_admin, patient_ID, console_results)
            print(console_results)
            return [test_name, printable_list, console_results]
            
        else:
            #txtWrite(patient_admin, patient_ID, ("Sözel akıcılık: Bu grup için norm mevcut değildir.\n" + console_results))
            return [test_name, printable_list, console_results]
    
    except:
        print("Sözel akıcılık testini değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return

def testWisconsin():
    try:
        while True:
            try:
                print("\n===================================\nYetişkin Wisconsin kart eşleme: \n")
                result_name = ["\nToplam tepki sayısı puanı: " ,
                               "\nTamamlanan kategori sayısı puanı: ",
                               "\nToplam perseveretif hata sayısı puanı: ",
                               "\nPerseveratif hata yüzdesi puanı: "]              
                               
                result_list = []               
                for i in range(len(result_name)):   
                    result_list.append(floInput(result_name[i]))
                #gets raw input from the user, these are test results and creates a list from them
                  
                mean_list = []
                sd_list = []
                for i in range(len(result_list)):
                    mean_list.append(result_list[i]-999)
                    sd_list.append(1)      
        
                norm_exists = True
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
            else:
                break
            
        if patient_edu <= 11:
            if 20 <= patient_age <= 54:
                mean_list = [121.70, 3.45, 29.85, 19.22] 
                sd_list = [13.27, 1.78, 13.55, 13.17] 
                
            elif 55 <= patient_age:
                mean_list = [122.04, 3.18, 32.71, 22.36]
                sd_list = [13.43, 1.63, 13.78, 14.93]

            else:
                print("Bu yaş grubu için norm mevcut değildir.")
                norm_exists = False
                
        elif 12 <= patient_edu:
            if 20 <= patient_age <= 54:
                mean_list = [106.25, 4.98, 17.69, 15.72]
                sd_list = [21.12, 1.59, 11.11, 8.34] 
                
            elif 55 <= patient_age:
                mean_list = [104.68, 4.77, 16.77, 18.14]
                sd_list = [20.49, 1.95, 10.34, 12.30]

            else:
                print("Bu yaş grubu için norm mevcut değildir.")
                norm_exists = False
  
        else:
            print("Bu eğitim grubu için norm mevcut değildir.")
            norm_exists = False
        
        z_score_list = calcZscore(result_list, mean_list, sd_list)
        z_score_verbal_list = calcSD_more_better(z_score_list)
        printable_list = outputPrintlist(result_list, z_score_list, z_score_verbal_list)

        
        test_name = 'wisconsin_data.csv' #test datasının toplanacağı csv dosyasını belirtiyor
        #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)
        
        console_results = "==================================\nYetişkin Wisconsin kart eşleme testinin sonuçları: "
        for i in range(len(result_list)):
            console_results = console_results + (result_name[i] + str(outputConsole_results(result_list, z_score_list, z_score_verbal_list)[i]))            
        console_results = console_results + ("\n==================================")
        
        if norm_exists:      
            #txtWrite(patient_admin, patient_ID, console_results)
            print(console_results)
            return [test_name, printable_list, console_results]
            
        else:
            #txtWrite(patient_admin, patient_ID, ("Yetişkin Wisconsin kart eşleme: Bu grup için norm mevcut değildir.\n" + console_results))
            return [test_name, printable_list, console_results]
    
    except:
        print("Yetişkin Wisconsin kart eşleme testini değerlendirirken bir hata oluştu," +
        "program kapatılacak.")
        raise
        return

        
def testTm():
    try:
        while True:
            try:
                print("\n===================================\nİz sürme testi: \n")
        
                result_name = ["\nA skoru: ","\nA formu düzeltme sayısı: ",
                               "\nB skoru: ","\nB formu düzeltme sayısı: ",
                               "\nB+A skoru: ", "\nB-A skoru: "]
                
                #prints user interface
                
                result_list = []         
                i = 0
                shouldsum = True
                while i < len(result_name):
                    if i in [0,2]:
                        x = floInput(result_name[i])
                        result_list.append(x)
                        if x == 999:
                            shouldsum = False
                    elif i in [1,3]:
                        result_list.append(floInput(result_name[i]))
                        
                    elif i == 4:
                        if shouldsum:
                            result_list.append(result_list[1] + result_list[0])
                        else:
                            result_list.append(999)
                    elif i == 5:
                        if shouldsum:
                            result_list.append(result_list[1] - result_list[0])
                        else:
                            result_list.append(999)
                    
                    i += 1
                #gets raw input from the user, these are test results and creates a list from them
                
                mean_list = []
                sd_list = []
                #creates empty lists for future use 
                #for the means and standard deviations
                for i in range(len(result_list)):
                    mean_list.append(result_list[i]-999)
                    sd_list.append(1)            
                    #makes it so that it prints out 999 if there's 
                    #no norm calculated for that group
                norm_exists = True
                #It assumes that there's a norm for every group
                #if none exists, it changes to False
                
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
                #"Only enter numbers", and then resets the function
            
            else:
                break
            
            #tries to get user input, makes sure it's correct input
            
        if patient_sex == "Kadın":
            if patient_edu <= 5:
                if patient_age <= 54:
                    mean_list = [69.67, 108.83, 178.50] 
                    sd_list = [21.59, 15.22, 32.29] 
                
                elif 55 <= patient_age <= 59:
                    mean_list = [72.77, 127.85, 200.62]
                    sd_list = [21.41, 26.90, 42.62]
                
                elif 60 <= patient_age <= 64:
                    mean_list = [82.23, 141.62, 223.85]
                    sd_list = [29.80, 49.28, 74.05]
                
                elif 65 <= patient_age <= 69:
                    mean_list = [100.00, 168.91, 268.91]
                    sd_list = [33.33, 66.00, 91.14]
                
                elif 70 <= patient_age <= 74:
                    mean_list = [110.20, 235.30, 345.50]
                    sd_list = [34.29, 27.84, 31.78]
                    
                elif 75 <= patient_age <= 79:
                    mean_list = [158.20, 236.10, 394.30]
                    sd_list = [57.48, 55.33, 100.11]
                    
                elif 80 <= patient_age:
                    mean_list = [238.30, 301.10, 539.40]
                    sd_list = [89.47, 78.71, 165.17]
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir")
                    norm_exists = False
            
            elif 6 <= patient_edu <= 11:  
                if patient_age <= 54:
                    mean_list = [60.31, 94.92, 155.23]
                    sd_list = [13.24, 16.91, 22.37]
                
                elif 55 <= patient_age <= 59:
                    mean_list = [61.69, 122.15, 200.62]
                    sd_list = [24.55, 27.12, 42.62]
                
                elif 60 <= patient_age <= 64:
                    mean_list = [77.00, 140.10, 223.85]
                    sd_list = [22.57, 30.45, 74.05]
                
                elif 65 <= patient_age <= 69:
                    mean_list = [63.70, 105.70, 268.91]
                    sd_list = [13.66, 28.14, 91.14]
                
                elif 70 <= patient_age <= 74:
                    mean_list = [93.83, 226.50, 320.33]
                    sd_list = [47.53, 105.29, 128.01]
                    
                elif 75 <= patient_age <= 79:
                    mean_list = [110.50, 234.50, 345.00]
                    sd_list = [44.24, 50.34, 58.93]
                    
                elif 80 <= patient_age:
                    mean_list = [195.90, 309.50, 505.40]
                    sd_list = [74.72, 22.07, 89.17]
                else:
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False
                
            elif 12 <= patient_edu:
                if patient_age <= 54:
                    mean_list = [53.53, 89.71, 143.24]
                    sd_list = [10.81, 18.03, 23.93]
                
                elif 55 <= patient_age <= 59:
                    mean_list = [53.14, 109.07, 162.21]
                    sd_list = [13.60, 41.46, 49.78]
                
                elif 60 <= patient_age <= 64:
                    mean_list = [58.00, 110.70, 168.70]
                    sd_list = [20.01, 32.51, 51.46]
                
                elif 65 <= patient_age <= 69:
                    mean_list = [78.60, 122.00, 200.60]
                    sd_list = [31.54, 30.34, 50.21]
                
                elif 70 <= patient_age <= 74:
                    mean_list = [107.60, 229.80, 337.40]
                    sd_list = [43.50, 43.04, 51.57]
                    
                elif 75 <= patient_age <= 79:
                    mean_list = [132.00, 186.60, 318.60]
                    sd_list = [23.88, 43.74, 39.40]
                    
                elif 80 <= patient_age:
                    mean_list = [183.10, 252.00, 435.10]
                    sd_list = [13.43, 46.54, 50.56]
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir")
                    norm_exists = False
            else:
                print("Bu eğitim grubu için norm mevcut değildir.")
                norm_exists = False
                
        elif patient_sex == "Erkek":
            if patient_edu <= 5:
                if patient_age <= 54:
                    mean_list = [88.00, 126.17, 214.17]
                    sd_list = [21.90, 27.61, 42.52]
                
                elif 55 <= patient_age <= 59:
                    mean_list = [69.70, 122.90, 192.60]
                    sd_list = [7.67, 15.65, 18.49]
                
                elif 60 <= patient_age <= 64:
                    mean_list = [78.90, 121.40, 200.30]
                    sd_list = [27.02, 20.50, 44.68]
                
                elif 65 <= patient_age <= 69:
                    mean_list = [97.40, 161.10, 258.50]
                    sd_list = [38.74, 37.86, 69.72]
                
                elif 70 <= patient_age <= 74:
                    mean_list = [89.20, 219.20, 308.40]
                    sd_list = [35.39, 61.58, 71.95]
                    
                elif 75 <= patient_age <= 79:
                    mean_list = [121.40, 228.50, 276.60]
                    sd_list = [30.89, 43.22, 51.51]
                    
                elif 80 <= patient_age:
                    mean_list = [286.40, 351.70, 355.00]
                    sd_list = [28.23, 26.05, 51.97]
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir")
                    norm_exists = False
            
            elif 6 <= patient_edu <= 11:
                if patient_age <= 54:
                    mean_list = [55.35, 89.47, 144.82]
                    sd_list = [14.27, 16.69, 26.61]
                
                elif 55 <= patient_age <= 59:
                    mean_list = [60.81, 123.91, 184.73]
                    sd_list = [8.33, 31.00, 35.17]
                
                elif 60 <= patient_age <= 64:
                    mean_list = [62.25, 117.08, 179.33]
                    sd_list = [8.02, 32.79, 34.03]
                
                elif 65 <= patient_age <= 69:
                    mean_list = [69.39, 130.15, 199.54]
                    sd_list = [12.68, 38.78, 48.05]
                
                elif 70 <= patient_age <= 74:
                    mean_list = [94.50, 206.80, 320.33]
                    sd_list = [29.50, 63.31, 128.01]
                    
                elif 75 <= patient_age <= 79:
                    mean_list = [90.20, 167.60, 257.80]
                    sd_list = [29.50, 49.80, 59.37]
                    
                elif 80 <= patient_age:
                    mean_list = [158.30, 2542.50, 412.80]
                    sd_list = [62.90, 51.58, 81.25]
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir")
                    norm_exists = False
            
            elif 12 <= patient_edu:
                if patient_age <= 54:
                    mean_list = [52.38, 86.25,138.53]
                    sd_list = [16.23, 23.02, 36.48]
                
                elif 55 <= patient_age <= 59:
                    mean_list = [51.16, 109.07, 174.68]
                    sd_list = [7.54, 41.46, 22.62]
                
                elif 60 <= patient_age <= 64:
                    mean_list = [71.60, 125.40, 197.00]
                    sd_list = [39.18, 56.23, 93.00]
                
                elif 65 <= patient_age <= 69:
                    mean_list = [54.27, 108.55, 162.82]
                    sd_list = [14.95, 31.40, 40.73]
                
                elif 70 <= patient_age <= 74:
                    mean_list = [83.40, 193.90, 277.30]
                    sd_list = [28.70, 71.35, 79.37]
                    
                elif 75 <= patient_age <= 79:
                    mean_list = [91.10, 185.50, 355.00]
                    sd_list = [43.29, 51.07, 51.97]
                    
                elif 80 <= patient_age:
                    mean_list = [128.90, 226.10, 3.00]
                    sd_list = [41.25, 53.64, 0.22]
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False              
            else:
                print("Bu eğitim grubu için norm mevcut değildir.")
                norm_exists = False
                
        else:
            print("Bu cinsiyet için norm mevcut değildir.")
            norm_exists = False
            
        if patient_age <= 54:
            mean_list = mean_list + [36.12, 0.02, 0.15] 
            sd_list = sd_list + [17.93, 0.21, 0.60] 
        
        elif 55 <= patient_age <= 59:
             mean_list = mean_list + [61.08, 0.00, 0.14]
             sd_list = sd_list + [26.32, 0, 0.65]

        elif 60 <= patient_age <= 64:
            mean_list = mean_list + [54.57, 0.00, 0.00]
            sd_list = sd_list + [26.86, 0.17, 0.38]
        
        elif 65 <= patient_age <= 69:
            mean_list = mean_list + [55.94, 0.00, 0.15]
            sd_list = sd_list + [26.86, 0.12, 0.62]
        
        elif 70 <= patient_age <= 74:
            mean_list = mean_list + [122.60, 0.11, 0.31]
            sd_list = sd_list + [74.46, 0.32, 0.80]
            
        elif 75 <= patient_age <= 79:
            mean_list = mean_list + [89.23, 0.00, 0.32]
            sd_list = sd_list + [64.05, 0.22, 0.77]
            
        elif 80 <= patient_age:
            mean_list = mean_list + [84.00, 0.15, 0.52]
            sd_list = sd_list + [60.04, 5.55, 0.98]
        
        #this is a temporary fix for the reordering of the data structure till
        #json files are implemented
        temp_mean_list = []
        temp_sd_list = []
        
        for i in [0, 4, 1, 5, 2, 3]:
            temp_mean_list.append(mean_list[i])
            temp_sd_list.append(sd_list[i])
    
        mean_list = temp_mean_list
        sd_list = temp_sd_list
        
        
        z_score_list = calcZscore(result_list, mean_list, sd_list)
        
        for i in range(len(z_score_list)):
            z_score_list[i] = -(float(z_score_list[i]))
            
        z_score_verbal_list = calcSD_less_better(z_score_list)
        printable_list = outputPrintlist(result_list, z_score_list, z_score_verbal_list)

        test_name = 'trail_making_data.csv' #test datasının toplanacağı csv dosyasını belirtiyor
        #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)

        console_results = "==================================\nİz sürme testinin sonuçları: "
        for i in range(len(mean_list)):
            console_results = console_results + (result_name[i] + str(outputConsole_results(result_list, z_score_list, z_score_verbal_list)[i]))            
        console_results = console_results + ("\n==================================")
        
        if norm_exists:
            #txtWrite(patient_admin, patient_ID, console_results)
            print(console_results)
            return [test_name, printable_list, console_results]
            
        else:
            #txtWrite(patient_admin, patient_ID, ("İz sürme testi: Bu grup için norm mevcut değildir.\n"+console_results))
            return [test_name, printable_list, console_results]
    
    except:
        print("İz sürme testini değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return
                     
                
def testStroop(): #Stroop testinin ana fonksiyonu
    try:
        while True:
            try:
                print("\n===================================\nStroop Testi: \n")
                result_name = ["\nBölüm 1: ","\nHata-1: ", "\nDüzeltme-1: ",
                               "\nBölüm 2: ","\nHata-2: ", "\nDüzeltme-2: ",
                               "\nBölüm 3: ","\nHata-3: ", "\nDüzeltme-3: ",
                               "\nBölüm 4: ","\nHata-4: ", "\nDüzeltme-4: ",
                               "\nBölüm 5: ","\nHata-5: ", "\nDüzeltme-5: "
                               ]
                
                result_list = []               
                for i in range(len(result_name)):   
                    result_list.append(floInput(result_name[i]))
                #gets raw input from the user, these are test results and creates a list from them
                         
                
                mean_list = []
                sd_list = []
                for i in range(len(result_list)):
                    mean_list.append(result_list[i]-999)
                    sd_list.append(1)      
                norm_exists = True
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
            else:
                break
            
        if patient_age <= 54 and patient_edu <= 8:
            mean_list = [12.13, result_list[1]-999, result_list[2]-999,
                        13.61, result_list[4]-999, result_list[5]-999,
                        17.46, result_list[7]-999, result_list[8]-999,
                        28.07, result_list[10]-999, result_list[11]-999,
                        40.57, result_list[13]-999, result_list[14]-999]
                        
            sd_list = [6.29, 1, 1,
                       7.41, 1, 1,
                       9.60, 1, 1,
                       13.85, 1, 1,
                       24.24, 1, 1]     
            
        elif 55 <= patient_age and patient_edu <= 8:
            mean_list = [13.51, result_list[1]-999, result_list[2]-999,
                         16.47, result_list[4]-999, result_list[5]-999,
                         24.45, result_list[7]-999, result_list[8]-999,
                         38.39, result_list[10]-999, result_list[11]-999,
                         47.93, result_list[13]-999, result_list[14]-999]

            sd_list = [5.49, 1, 1,
                       6.76, 1, 1,
                       13.36, 1, 1,
                       18.52, 1, 1,
                       20.82, 1, 1]
            
        elif patient_age <= 54 and  8 < patient_edu:
            mean_list = [8.81, result_list[1]-999, result_list[2]-999,
                         9.43, result_list[4]-999, result_list[5]-999,
                         12.32, result_list[7]-999, result_list[8]-999,
                         16.95, result_list[10]-999, result_list[11]-999,
                         26.38, result_list[13]-999, result_list[14]-999]

            sd_list = [1.76, 1, 1,
                       2.52, 1, 1,
                       2.71, 1, 1,
                       6.70, 1, 1,
                       12.29, 1, 1]
            
        elif 55 <= patient_age and 8 < patient_edu:
            mean_list = [10.09, result_list[1]-999, result_list[2]-999,
                         11.63, result_list[4]-999, result_list[5]-999,
                         15.93, result_list[7]-999, result_list[8]-999,
                         24.87, result_list[10]-999, result_list[11]-999,
                         35.96, result_list[13]-999, result_list[14]-999]

            sd_list = [3.71, 1, 1,
                       5.41, 1, 1,
                       4.06, 1, 1,
                       10.94, 1, 1,
                       16.23, 1, 1]
            
        else:
            print("Stroop: Bu yaş veya eğitim grubu için norm mevcut değildir.") 
            norm_exists = False
                
                
        z_score_list = calcZscore(result_list, mean_list, sd_list)
        for i in range(len(z_score_list)):
            z_score_list[i] = -(float(z_score_list[i]))
        z_score_verbal_list = calcSD_less_better(z_score_list)
        printable_list = outputPrintlist(result_list, z_score_list, z_score_verbal_list)
        
        
        test_name = 'stroop_data.csv' #test datasının toplanacağı csv dosyasını belirtiyor
                       
        console_results = "==================================\nStroop testinin sonuçları: "
        for i in range(len(mean_list)):
            console_results = console_results + (result_name[i] +
            str(outputConsole_results(result_list, z_score_list, z_score_verbal_list)[i]))            
        console_results = console_results + ("\n==================================")
        
        if norm_exists:
            print(console_results)
            return [test_name, printable_list, console_results]
            
        else:
            print(console_results)
            return [test_name, printable_list, console_results]
        
    except:
        print("Stroop testini işlerken bir hata oluştu, sistem yeniden başlatılacak.")
        raise
        return

                
def testSdot():
    try:
        while True:
            try:
                print("\n===================================\nYetişkinlerde çizgi yönünü belirleme: \n")
                sdot_1 = floInput("Puan: ")
                result_list = [sdot_1]
                
                result_name = ["\nÇizgi yönü: "]
                
                mean_list = []
                sd_list = []
                for i in range(len(result_list)):
                    mean_list.append(result_list[i]-999)
                    sd_list.append(1)      
                norm_exists = True
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
            else:
                break
            
        if patient_age <= 54 and patient_edu <= 8:
            mean_list = [20.03]
            sd_list = [4.58]     
            
        elif 55 <= patient_age and patient_edu <= 8:
            mean_list = [19.81]
            sd_list = [4.00]
            
        elif patient_age <= 54 and  9 <= patient_edu <= 11:
            mean_list = [22.46]
            sd_list = [4.52]
            
        elif 55 <= patient_age and 9 <= patient_edu <= 11:
            mean_list = [21.18]
            sd_list = [4.82]
            
        elif patient_age <= 54 and 12 <= patient_edu:
            mean_list = [25.37]
            sd_list = [3.51]
            
        elif 55 <= patient_age and 12 <= patient_edu:
            mean_list = [23.04]
            sd_list = [3.43]
            
        else:
            print("Çizgi yönü: Bu yaş veya eğitim grubu için norm mevcut değildir.") 
            norm_exists = False
        
        z_score_list = calcZscore(result_list, mean_list, sd_list)
        z_score_verbal_list = calcSD_more_better(z_score_list)
        printable_list = outputPrintlist(result_list, z_score_list, z_score_verbal_list)
        
        test_name = 'sdot_data.csv' #test datasının toplanacağı csv dosyasını belirtiyor
        #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)
        
        console_results = "==================================\nÇizgi yönünü belirleme testinin sonuçları: "
        for i in range(len(mean_list)):
            console_results = console_results + (result_name[i] + str(outputConsole_results(result_list, z_score_list, z_score_verbal_list)[i]))            
        console_results = console_results + ("\n==================================")
        
        if norm_exists:
            #txtWrite(patient_admin, patient_ID, console_results)
            print(console_results)
            return [test_name, printable_list, console_results]
            
        else:
            #txtWrite(patient_admin, patient_ID, ("Çizgi yönü: Bu grup için norm mevcut değildir\n"+console_results))
            return [test_name, printable_list, console_results]
        
    except:
        print("Çizgi yönü testini işlerken bir hata oluştu, sistem yeniden başlatılacak.")
        raise
        return


#these variables were got before
def testSbst(): 
    try:
        while True:
            try:
                print("\n===================================\nSBST: ")
                result_name = ["\nKendiliğinden hatırlama boyutu: ", "\nTanıma boyutu: ", 
                "\nToplam hatırlama boyutu: ",  "\nHatırlama yanlışı boyutu: ", 
                "\nYanlış tanıma boyutu: ", "\nA listesi anlık bellek boyutu: ", 
                "\nA listesi toplam öğrenme boyutu: "]  
                #prints user interface

                result_list = []               
                for i in range(len(result_name)):   
                    result_list.append(floInput(result_name[i]))
                    #gets raw input from the user, these are test results and creates a list from them
              
                mean_list = []
                sd_list = []
                #creates empty lists for future use 
                #for the means and standard deviations
                for i in range(len(result_list)):
                    mean_list.append(result_list[i]-999)
                    sd_list.append(1)            
                    #makes it so that it prints out 999 if there's 
                    #no norm calculated for that group
                norm_exists = True
                #It assumes that there's a norm for every group
                #if none exists, it changes to False
                
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
                #"Only enter numbers", and then resets the function
            
            else:
                break
            
            #tries to get user input, makes sure it's correct input
            
            
            """
            Following are lists of means and standard deviations,
            of the specified age, education and sex, where it applies.
            It's in a way that it corresponds to order of the
            result_list.
            """
        if patient_edu <= 7:
            if patient_age <= 39:
                mean_list = [13.64, 1.21, 14.86, 0.24, 0.04, 5.50, 116.27]
                sd_list = [1.20, 1.16, 0.35, 0.51, 0.19, 1.64, 10.74]
                
            elif 40 <= patient_age <= 49:
                mean_list = [13.27, 1.69, 14.97, 0.15, 0.08, 5.32, 109.11]
                sd_list = [1.24, 1.21, 0.18, 0.40, 0.27, 1.35, 11.94]
                
            elif 50 <= patient_age <= 59:
                mean_list = [13.24, 1.74, 14.98, 0.15, 0.00, 5.08, 113.50] 
                sd_list = [1.38, 1.37, 0.12, 0.40, 0, 1.03, 12.00]
            
            elif 60 <= patient_age <= 69:
                mean_list = [12.90, 2.27, 14.94, 0.20, 0.02, 5.10, 107.80]
                sd_list = [1.36, 2.04, 0.24, 0.46, 0.14, 1.37, 11.16]
            
            elif 70 <= patient_age <= 79:
                mean_list = [12.93, 2.02, 14.96, 0.24, 0.02, 5.00, 103.69]
                sd_list = [1.44, 1.39, 0.21, 0.43, 0.15, 1.09, 14.21]
                
            elif 80 <= patient_age:
                mean_list = [10.35, 4.35, 14.70, 0.26, 0.30, 3.96, 88.96]
                sd_list = [1.72, 1.58, 0.64, 0.45, 0.56, 1.66, 14.02]
                
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False     
                #"No norm exists for the group", sets it to False
                
        elif 8 <= patient_edu:
            if patient_age <= 39:
                mean_list = [14.00, 0.97, 14.97, 0.13, 0.00, 7.87, 131.21] 
                sd_list = [1.05, 1.04, 0.16, 0.45, 0, 2.05, 10.61]
                
            elif 40 <= patient_age <= 49:
                mean_list = [13.62, 1.29, 14.91, 0.15, 0.03, 6.98, 126.12]
                sd_list = [1.39, 1.27, 0.29, 0.40, 0.17, 1.96, 11.25]
                
            elif 50 <= patient_age <= 59:
                mean_list = [13.45, 1.54, 14.97, 0.13, 0.03, 6.61, 124.22]
                sd_list = [1.29, 1.28, 0.17, 0.34, 0.17, 1.50, 10.48]
            
            elif 60 <= patient_age <= 69:
                mean_list = [13.64, 1.22, 14.88, 0.12, 0.00, 5.94, 116.88]
                sd_list = [1.10, 0.91, 0.39, 0.39, 0, 1.33, 11.15]
            
            elif 70 <= patient_age <= 79:
                mean_list = [12.29, 2.63, 14.94, 0.13, 0.04, 5.27, 109.75]
                sd_list = [1.80, 1.75, 0.24, 0.34, 0.20, 1.45, 13.98]
                
            elif 80 <= patient_age:
                mean_list = [11.88, 3.02, 14.89, 0.18, 0.09, 5.19, 109.68]
                sd_list = [1.72, 1.63, 0.36, 0.47, 0.34, 1.64, 16.80]
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False  
        else:
            print("Bu eğitim grubu için norm mevcut değildir.")
            norm_exists = False  
                
        z_score_list = calcZscore(result_list, mean_list, sd_list)
        #it calculated the patient's SD interval as a float using the results, means and the SD
        z_score_verbal_list = calcSD_more_better([z_score_list[0], z_score_list[1], z_score_list[2]])
        z_score_list[3] = -z_score_list[3]
        z_score_list[4] = -z_score_list[4]
        z_score_verbal_list = z_score_verbal_list + calcSD_less_better([(z_score_list[3]),(z_score_list[4])])        
        z_score_verbal_list = z_score_verbal_list + calcSD_more_better([z_score_list[5], z_score_list[6]])        
        #Makes a verbal list to input to the users
        printable_list = outputPrintlist(result_list, z_score_list, z_score_verbal_list)
        #creates a list to be put into a CSV file

        test_name = 'SBST_data.csv' #test datasının toplanacağı csv dosyasını belirtiyor
        #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)
        #writes the printable_list in a CSV file
        
        console_results = "==================================\nSBST testinin sonuçları:"
        for i in range(len(mean_list)):
            console_results = console_results + (result_name[i] + str(outputConsole_results(result_list, z_score_list, z_score_verbal_list)[i]))            
        console_results = console_results + ("\n==================================")
        #prints the list using the range of the list and makes it more appealing for the user (probably should've used a decorator here)
        
        if norm_exists:
            #txtWrite(patient_admin, patient_ID, console_results)
            print(console_results)
            return [test_name, printable_list, console_results]
            #creates a patient report for the physician and prints it out for the user
        else:
            #txtWrite(patient_admin, patient_ID, ("SBST: Bu grup için norm mevcut değildir.\n"+console_results))
            return [test_name, printable_list, console_results]
            #if there's no norm, it puts "there's no norm for this group" in the patient report
    
    except:
        print("SBST testini değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return
        #saves the program from fiery death
        
#these variables were got before
def testVvt(): #visual verbal test
    try:
        while True:
            try:
                print("\n===================================\nGörsel Sözel Test: ")
                result_name = ["\nToplam 'sort' sayısı: ", "\nToplam 'shift' sayısı: "]  
                #prints user interface
                
                vvt_sort = floInput(result_name[0])
                vvt_shift = floInput(result_name[1])
                
                #gets raw input from the user, these are test results
                
                result_list = [vvt_sort, vvt_shift]
                mean_list = []
                sd_list = []
                #creates a list for the test results, and empty ones 
                #for the means and standard deviations
                for i in range(len(result_list)):
                    mean_list.append(result_list[i]-999)
                    sd_list.append(1)            
                    #makes it so that it prints out 999 if there's 
                    #no norm calculated for that group
                norm_exists = True
                #It assumes that there's a norm for every group
                #if none exists, it changes to False
                
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
                #"Only enter numbers", and then resets the function
            
            else:
                break
            
            #tries to get user input, makes sure it's correct input
            
            
            """
            Following are lists of means and standard deviations,
            of the specified age, education and sex, where it applies.
            It's in a way that it corresponds to order of the
            result_list.
            """
        if patient_edu <= 5:
            if patient_age <= 49:
                mean_list = [15.72, 6.53]
                sd_list = [3.40, 2.55]
                
            elif 50 <= patient_age <= 59:
                mean_list = [13.74, 5.36]
                sd_list = [4.81, 2.85]
            
            elif 60 <= patient_age <= 69:
                mean_list = [12.64, 4.46]
                sd_list = [4.22, 2.76]
            
            elif 70 <= patient_age:
                mean_list = [9.27, 2.82]
                sd_list = [5.90, 2.87]
                
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False     
                #"No norm exists for the group", sets it to False
                
        elif 6 <= patient_edu <= 11:
            if patient_age <= 49:
                mean_list = [16.45, 7.11]
                sd_list = [3.47, 2.43]
                
            elif 50 <= patient_age <= 59:
                mean_list = [15.43, 6.93]
                sd_list = [4.76, 3.22]
            
            elif 60 <= patient_age <= 69:
                mean_list = [14.30, 5.25]
                sd_list = [4.10,  2.95]
            
            elif 70 <= patient_age:
                mean_list = [13.80, 4.95]
                sd_list = [4.07, 3.15]
                
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False     
                #"No norm exists for the group", sets it to False
                
        elif 12 <= patient_edu:
            if patient_age <= 49:
                mean_list = [18.41, 8.60]
                sd_list = [1.75, 1.43]
                
            elif 50 <= patient_age <= 59:
                mean_list = [17.71, 7.97]
                sd_list = [2.52, 2.10]
            
            elif 60 <= patient_age <= 69:
                mean_list = [16.75, 7.18]
                sd_list = [2.94, 2.18]
            
            elif 70 <= patient_age:
                mean_list = [15.00, 5.83]
                sd_list = [4.28, 3.01]
                
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False     
                #"No norm exists for the group", sets it to False
        else:
            print("Bu eğitim grubu için norm mevcut değildir.")
            norm_exists = False  

        z_score_list = calcZscore(result_list, mean_list, sd_list)
        #it calculated the patient's SD interval as a float using the results, means and the SD
        z_score_verbal_list = calcSD_more_better(z_score_list)
        #Makes a verbal list to input to the users
        printable_list = outputPrintlist(result_list, z_score_list, z_score_verbal_list)
        #creates a list to be put into a CSV file

        test_name = 'VVT_data.csv' #declares name of the CSV file to save the data in
        #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)
        #writes the printable_list in a CSV file
        
        console_results = "==================================\nGörsel Sözel Test sonuçları:"
        for i in range(len(mean_list)):
            console_results = console_results + (result_name[i] + str(outputConsole_results(result_list, z_score_list, z_score_verbal_list)[i]))            
        console_results = console_results + ("\n==================================")
        #prints the list using the range of the list and makes it more appealing for the user (probably should've used a decorator here)
        
        if norm_exists:
            #txtWrite(patient_admin, patient_ID, console_results)
            print(console_results)
            return [test_name, printable_list, console_results]
            #creates a patient report for the physician and prints it out for the user
        else:
            #txtWrite(patient_admin, patient_ID, ("VVT: Bu grup için norm mevcut değildir.\n"+console_results))
            return [test_name, printable_list, console_results]
            #if there's no norm, it puts "there's no norm for this group" in the patient report
    
    except:
        print("VVT'yi değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return
        #saves the program from fiery death


def testCct_1(): #color trails test
    try:
        while True:
            try:
                print("\n===================================\nCTT1: ")
                result_name = ["\nTime: ", "\nErrors: ", "\nNear-Misses: ", "\nPrompts: "]  
                #prints user interface
                
                result_list = []               
                
                for i in range(len(result_name)):
                    result_list.append(floInput(result_name[i]))
                #gets raw input from the user, these are test results
                    
                mean_list = []
                sd_list = []
                
                #creates a list for the test results, and empty ones 
                #for the means and standard deviations
                for i in range(len(result_list)):
                    mean_list.append(result_list[i]-999)
                    sd_list.append(1)            
                    #makes it so that it prints out 999 if there's 
                    #no norm calculated for that group
                    
                norm_exists = True
                #It assumes that there's a norm for every group
                #if none exists, it changes to False
                
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                raise
                continue
                #"Only enter numbers", and then resets the function
            
            else:
                break
            
            #tries to get user input, makes sure it's correct input
            """
            Following are lists of means and standard deviations,
            of the specified age, education and sex, where it applies.
            It's in a way that it corresponds to order of the
            result_list.
            """

        if patient_edu <= 5:
            if patient_age <= 49:
                mean_list = [6.65, 0.15, 0.46, 0.20]
                sd_list = [29.90, 0.74, 1.30, 0.40]
                
            elif 50 <= patient_age <= 59:
                mean_list = [90.53, 0.23, 1.38, 0.07]
                sd_list = [36.14, 0.61, 3.17, 0.27]
            
            elif 60 <= patient_age <= 69:
                mean_list = [91.68, 0.33, 1.24, 0.17]
                sd_list = [35.17, 0.63, 1.28, 0.44]
            
            elif 70 <= patient_age:
                mean_list = [152.36, 0.58, 2.21, 0.34]
                sd_list = [68.13, 1.46, 2.25, 0.76]
                
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False     
                #"No norm exists for the group", sets it to False
                
        elif 6 <= patient_edu <= 11:
            if patient_age <= 49:
                mean_list = [55.45, 0.04, 0.26, 0.08]
                sd_list = [21.06, 0.26, 0.67, 0.27]
                
            elif 50 <= patient_age <= 59:
                mean_list = [65.97, 0.10, 0.28, 0.02]
                sd_list = [23.43, 0.31, 0.50, 0.14]
            
            elif 60 <= patient_age <= 69:
                mean_list = [87.72, 0.22, 0.72, result_list[3]]
                sd_list = [32.62, 0.73, 1.39, 1]
            
            elif 70 <= patient_age:
                mean_list = [110.58, 0.25, 1.65, 0.04]
                sd_list = [47.84, 0.58, 2.84, 0.21]
                
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False     
                #"No norm exists for the group", sets it to False
                
        elif 12 <= patient_edu:
            if patient_age <= 49:
                mean_list = [44.91, 0.01, 0.01, 0.01]
                sd_list = [14.66, 0.10, 0.10, 0.10]
                
            elif 50 <= patient_age <= 59:
                mean_list = [61.20, 0.02, 0.26, 0.04]
                sd_list = [25.72, 0.14, 0.70, 0.19]
            
            elif 60 <= patient_age <= 69:
                mean_list = [73.42, 0.09, 0.21, 0.12]
                sd_list = [36.88, 0.29, 0.48, 0.41]
            
            elif 70 <= patient_age:
                mean_list = [106.77, 0.27, 1.05, 0.11]
                sd_list = [49.88, 0.66, 1.58, 0.47]
                
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False     
                #"No norm exists for the group", sets it to False
        else:
            print("Bu eğitim grubu için norm mevcut değildir.")
            norm_exists = False  
            #"No norm exists for the group", sets it to False
                
        z_score_list = calcZscore(result_list, mean_list, sd_list)
        #it calculated the patient's SD interval as a float using the results, means and the SD
        for i in range(len(z_score_list)):
            z_score_list[i] = -(float(z_score_list[i]))
        z_score_verbal_list = calcSD_less_better(z_score_list)
        #Makes a verbal list to input to the users
        printable_list = outputPrintlist(result_list, z_score_list, z_score_verbal_list)
        #creates a list to be put into a CSV file

        test_name = 'CCT1_data.csv' #declares name of the CSV file to save the data in
        #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)
        #writes the printable_list in a CSV file        
              
        console_results = "==================================\nColor Trails Test 1 sonuçları:"
        for i in range(len(mean_list)):
            console_results = console_results + (result_name[i] + str(outputConsole_results(result_list, z_score_list, z_score_verbal_list)[i]))            
        console_results = console_results + ("\n==================================")
        #prints the list using the range of the list and makes it more appealing for the user (probably should've used a decorator here)
        
        if norm_exists:
            #txtWrite(patient_admin, patient_ID, console_results)
            print(console_results)
            return [test_name, printable_list, console_results]
            #creates a patient report for the physician and prints it out for the user
        else:
            #txtWrite(patient_admin, patient_ID, ("CCT1: Bu grup için norm mevcut değildir.\n"+console_results))
            return [test_name, printable_list, console_results]
            #if there's no norm, it puts "there's no norm for this group" in the patient report
    
    except:
        print("CCT 1'i değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        #saves the program from fiery death

        
def testCct_2(): #color trails test
    try:
        while True:
            try:
                print("\n===================================\nCTT2: ")
                result_name = ["\nTime: ", "\nNumber Errors: ", "\nColor Errors: ", "\nNear-Misses: ", "\nPrompts: "]  
                #prints user interface
                
                result_list = []               
                for i in range(len(result_name)):
                    result_list.append(floInput(result_name[i]))
                #gets raw input from the user, these are test results
                    
                mean_list = []
                sd_list = []
                #creates a list for the test results, and empty ones 
                #for the means and standard deviations
                for i in range(len(result_list)):
                    mean_list.append(result_list[i]-999)
                    sd_list.append(1)            
                    #makes it so that it prints out 999 if there's 
                    #no norm calculated for that group
                norm_exists = True
                #It assumes that there's a norm for every group
                #if none exists, it changes to False
                
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
                #prints "Only enter numbers", and then resets the function
            
            else:
                break
            
            #tries to get user input, makes sure it's correct input
            
            #Following are lists of means and standard deviations,
            #of the specified age, education and sex, where it applies.
            #It's in a way that it corresponds to order of the
            #result_list.
            
        if patient_edu <= 5:
            if patient_age <= 39:
                mean_list = [114.97, 0.62, 0.06, 1.20, 0.23]
                sd_list = [46.15, 0.87, 0.25, 1.37, 0.42]
            
            elif 40 <= patient_age <= 49:
                mean_list = [138.85, 1.29, 0.41, 2.70, 0.26]
                sd_list = [67.24, 1.89, 1.15, 3.15, 0.51]
                
            elif 50 <= patient_age <= 59:
                mean_list = [163.68, 0.44, 0.26, 2.34, 0.22]
                sd_list = [63.03, 0.73, 0.78, 2.01, 0.49]
            
            elif 60 <= patient_age <= 69:
                mean_list = [186.11, 0.82, 0.44, 3.77, 0.55]
                sd_list = [62.50, 0.96, 0.86, 2.77, 0.81]
            
            elif 70 <= patient_age:
                mean_list = [250.17, 1.51, 0.21, 4.12, 0.24]
                sd_list = [93.83, 1.89, 0.57, 2.96, 0.58]
                
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False     
                #"No norm exists for the group", sets it to False
                
        elif 6 <= patient_edu <= 11:
            if 20 <= patient_age <= 39:
                mean_list = [100.82, 0.29, 0.12, 1.10, 0.37]
                sd_list = [35.16, 0.58, 0.48, 1.25, 0.67]
            
            elif 40 <= patient_age <= 49:
                mean_list = [108.83, 0.50, result_list[2], 0.75, 0.25]
                sd_list = [31.04, 0.88, 1, 0.79, 0.44]
                
            elif 50 <= patient_age <= 59:
                mean_list = [117.23, 0.34, 0.08, 1.04, 0.13]
                sd_list = [37.81, 0.64, 0.35, 1.31, 0.34]
            
            elif 60 <= patient_age <= 69:
                mean_list = [159.65, 0.57, 0.17, 1.92, 0.23]
                sd_list = [55.94, 0.90, 0.50, 2.10, 0.61]
            
            elif 70 <= patient_age <= 100:
                mean_list = [197.44, 0.65, 0.11, 3.58, 0.34]
                sd_list = [58.89, 0.75, 0.32, 3.56, 0.78]
                
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False     
                #"No norm exists for the group", sets it to False
        
                
        elif 12 <= patient_edu:
            
            if patient_age <= 39:
                mean_list = [89.72, 0.20, 0.03, 0.66, 0.22]
                sd_list = [34.86, 0.51, 0.25, 0.93, 0.58]
            
            elif 40 <= patient_age <= 49:
                mean_list = [101.09, 0.45, 0.04, 1.00, 0.18]
                sd_list = [36.72, 0.85, 0.21, 1.23, 0.39]
                
            elif 50 <= patient_age <= 59:
                mean_list = [112.24, 0.24, 0.02, 0.81, 0.14]
                sd_list = [36.28, 0.56, 0.14, 1.13, 0.35]
            
            elif 60 <= patient_age <= 69:
                mean_list = [144.12, 0.45, 0.15, 0.15, 0.15]
                sd_list = [58.16, 0.86, 0.44, 0.44, 0.44]
            
            elif 70 <= patient_age:
                mean_list = [183.77, 0.55, 0.27, 2.72, 0.16]
                sd_list = [69.47, 1.04, 0.75, 2.44, 0.38]
                
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False     
                #"No norm exists for the group", sets it to False
        else:
            print("Bu eğitim grubu için norm mevcut değildir.")
            norm_exists = False  
            #"No norm exists for the group", sets it to False

        
        z_score_list = calcZscore(result_list, mean_list, sd_list)
        #it calculated the patient's SD interval as a float using the results, means and the SD
        for i in range(len(z_score_list)):
            z_score_list[i] = -(float(z_score_list[i]))
        z_score_verbal_list = calcSD_less_better(z_score_list)
        #Makes a verbal list to input to the users
        printable_list = outputPrintlist(result_list, z_score_list, z_score_verbal_list)
        #creates a list to be put into a CSV file

        test_name = 'CCT2_data.csv' #declares name of the CSV file to save the data in
        #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)
        #writes the printable_list in a CSV file
        
              
        console_results = "===================================\nColor Trails Test 2 sonuçları:"
        for i in range(len(mean_list)):
            console_results = console_results + (result_name[i] + str(outputConsole_results(result_list, z_score_list, z_score_verbal_list)[i]))            
        console_results = console_results + ("\n==================================")
            #prints the list using the range of the list and makes it more appealing for the user (probably should've used a decorator here)
        
        if norm_exists:
            #txtWrite(patient_admin, patient_ID, console_results)
            print(console_results)
            return [test_name, printable_list, console_results]
            #creates a patient report for the physician and prints it out for the user
        else:
            #txtWrite(patient_admin, patient_ID, ("CCT2: Bu grup için norm mevcut değildir.\n"+console_results))
            return [test_name, printable_list, console_results]
            #if there's no norm, it puts "there's no norm for this group" in the patient report
    
    except:
        print("CCT'yi değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return
        #saves the program from fiery death

def testCct():
    while True:
        try:
            dataCct = ['CCT_data.csv', "", ""]
            
            dataCct1 = testCct_1()            
            dataCct2 = testCct_2()
            
            dataCct[1] = dataCct1[1] + dataCct2[1]
            dataCct[2] = dataCct1[2] + dataCct2[2]
            
        except:
            print("CCT'yi değerlendirirken bir hata oluştu, program kapatılacak.")
            continue
        
        else:
            break
    
    return dataCct


def testRkft():
    try:
        while True:
            try: 
                print("\n===================================\nRey karmaşık figür testi: ")
                result_name = ["\nKopyalama: ", "\nAnlık hatırlama: ", 
                "\nGecikmeli hatırlama: ",  "\nTanıma doğru pozitif puanı: ", 
                "\nTanıma yanlış pozitif puanı: "]  
                
                result_list = []               
                for i in range(len(result_name)):
                    result_list.append(floInput(result_name[i]))
                #gets raw input from the user, these are test results
    
                mean_list = []
                sd_list = []
                for i in range(len(result_list)):
                    mean_list.append(result_list[i]-999)
                    sd_list.append(1)            
                norm_exists = True
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
            else:
                break
        
        if patient_edu < 8:
            if patient_age <= 49:
                mean_list = [29.90, 14.45, 14.35, 10.37, 2.63]
                sd_list = [5.277, 5.604, 5.992, 1.71, 1.56]
                
            elif 50 <= patient_age <= 60:
                mean_list = [27.3, 13.3, 13.5, 10, 2.50]
                sd_list = [6.391, 4.461, 4.577, 1.05, 1.65]
                
            elif 61 <= patient_age <= 71:
                mean_list = [27.05, 10.3, 8.85, 9.7, 2.4]
                sd_list = [4.764, 4.656, 4.308, 1.7, 1.78]
            
            elif 72 <= patient_age:
                mean_list = [24.3, 8.3, 7.55, 10.3, 3.2]
                sd_list = [5.673, 5.089, 4.919, 1.7, 1.99]
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False     
                
        elif 8 <= patient_edu:
            
            if patient_age <= 49:
                mean_list = [33.096, 19.75, 18.84, 9.73, 1.27]
                sd_list = [2.548, 4.993, 5.062, 1.74, 1.17]
                
            elif 50 <= patient_age <= 60:
                mean_list = [32.75, 16.7, 16.25, 10.2, 1.55]
                sd_list = [3.263, 4.998, 5.175, 1.4, 1.39]
                
            elif 61 <= patient_age <= 71:
                mean_list = [31.125, 14.525, 14.35, 9.45, 1.8]
                sd_list = [3.947, 6.315, 6.796, 1.9, 1.54]
            
            elif 72 <= patient_age:
                mean_list = [28.6, 10.975, 11.025, 9.1, 2.2]
                sd_list = [3.192, 4.805, 5.466, 1.83, 1.67]
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False  
        else:
            print("Bu eğitim grubu için norm mevcut değildir.")
            norm_exists = False  
                                
        z_score_list = calcZscore(result_list, mean_list, sd_list)
        z_score_verbal_list = calcSD_more_better(z_score_list)
        printable_list = outputPrintlist(result_list, z_score_list, z_score_verbal_list)

        test_name = 'rey_data.csv' #test datasının toplanacağı csv dosyasını belirtiyor
        #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)
        
        console_results = "==================================\nRey karmaşık figür testinin sonuçları:"
        for i in range(len(mean_list)):
            console_results = console_results + (result_name[i] + str(outputConsole_results(result_list, z_score_list, z_score_verbal_list)[i]))            
        console_results = console_results + ("\n==================================")
                
        if norm_exists:
            #txtWrite(patient_admin, patient_ID, console_results)
            print(console_results)
            return [test_name, printable_list, console_results]
            
        else:
            #txtWrite(patient_admin, patient_ID, ("Rey: Bu grup için norm mevcut değildir.\n"+console_results))
            return [test_name, printable_list, console_results]
    
    except:
        print("Rey testini değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return


def testGisd(): #görsel işitsel sayı dizileri testi
    try:
        while True:
            try:
                print("\n===================================\nGörsel İşitsel Sayı Dizileri: ")
                result_name = ["\nSonuç: "]  
                #prints user interface
                
                result_list = []               
                for i in range(len(result_name)):
                    result_list.append(floInput(result_name[i]))
                #gets raw input from the user, these are test results
                    
                mean_list = []
                sd_list = []
                #creates a list for the test results, and empty ones 
                #for the means and standard deviations
                for i in range(len(result_list)):
                    mean_list.append(result_list[i]-999)
                    sd_list.append(1)            
                    #makes it so that it prints out 999 if there's 
                    #no norm calculated for that group
                norm_exists = True
                #It assumes that there's a norm for every group
                #if none exists, it changes to False
                
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
                #"Only enter numbers", and then resets the function
            
            else:
                break
            
            #tries to get user input, makes sure it's correct input
            
            #Following are lists of means and standard deviations,
            #of the specified age, education and sex, where it applies.
            #It's in a way that it corresponds to order of the
            #result_list.
        if patient_sex == "Erkek":
            if patient_edu <= 5:
                if patient_age <= 15:
                    norm_exists = False
                    
                elif 16 <= patient_age <= 19:
                    norm_exists = False
                    
                elif 20 <= patient_age <= 34:
                    mean_list = [4.93]
                    sd_list = [1.21]
                
                elif 35 <= patient_age <= 54:
                    mean_list = [4.58]
                    sd_list = [0.98]
                
                elif 55 <= patient_age:
                    mean_list = [4.13]
                    sd_list = [0.97]
                    
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False
                    
            elif 6 <= patient_edu <= 11:
                if patient_age <= 15:
                    mean_list = [6.17]
                    sd_list = [1.21]
                    
                elif 16 <= patient_age <= 19:
                    mean_list = [6.40]
                    sd_list = [1.02]
                
                elif 20 <= patient_age <= 34:
                    mean_list = [6.20]
                    sd_list = [1.02]
                
                elif 35 <= patient_age <= 54:
                    mean_list = [5.57]
                    sd_list = [1.20]
                
                elif 55 <= patient_age:
                    mean_list = [5.00]
                    sd_list = [1.20]
                
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False
                    
            elif 12 <= patient_edu:
                if patient_age <= 15:
                    norm_exists = False
                    
                elif 16 <= patient_age <= 19:
                    norm_exists = False
                
                elif 20 <= patient_age <= 34:
                    mean_list = [5.97]
                    sd_list = [1.00]
                
                elif 35 <= patient_age <= 54:
                    mean_list = [6.00]
                    sd_list = [1.03]
                
                elif 55 <= patient_age:
                    mean_list = [5.13]
                    sd_list = [1.07]
                
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False
                    
            else:
                print("Bu eğitim grubu için norm mevcut değildir.")
                norm_exists = False  
                #"No norm exists for the group", sets it to False
                
        elif patient_sex == "Kadın":
            
            if patient_edu <= 5:
                if patient_age <= 15:
                    norm_exists = False
                    
                elif 16 <= patient_age <= 19:
                    norm_exists = False
                
                elif 20 <= patient_age <= 34:
                    mean_list = [5.09]
                    sd_list = [1.06]
                
                elif 35 <= patient_age <= 54:
                    mean_list = [4.32]
                    sd_list = [0.85]
                
                elif 55 <= patient_age:
                    mean_list = [3.92]
                    sd_list = [0.73]
                    
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False
                    
            elif 6 <= patient_edu <= 11:
                if patient_age <= 15:
                    mean_list = [5.81]
                    sd_list = [1.18]
                    
                elif 16 <= patient_age <= 19:
                    mean_list = [6.30]
                    sd_list = [1.24]
                
                elif 20 <= patient_age <= 34:
                    mean_list = [5.81]
                    sd_list = [1.12]
                
                elif 35 <= patient_age <= 54:
                    mean_list = [5.47]
                    sd_list = [0.93]
                
                elif 55 <= patient_age:
                    mean_list = [5.17]
                    sd_list = [0.99]
                    
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False

            elif 12 <= patient_edu:
                if patient_age <= 15:
                    norm_exists = False
                    
                elif 16 <= patient_age <= 19:
                    norm_exists = False
                
                elif 20 <= patient_age <= 34:
                    mean_list = [6.06]
                    sd_list = [1.14]
                
                elif 35 <= patient_age <= 54:
                    mean_list = [5.69]
                    sd_list = [1.19]
                
                elif 55 <= patient_age:
                    mean_list = [5.47]
                    sd_list = [1.04]
                    
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False
                    
            else:
                print("Bu eğitim grubu için norm mevcut değildir.")
                norm_exists = False  
                #"No norm exists for the group", sets it to False        
        else:
            print("Bu cinsiyet için norm mevcut değildir.")
            norm_exists = False
        
        z_score_list = calcZscore(result_list, mean_list, sd_list)
        #it calculated the patient's SD interval as a float using the results, means and the SD
        z_score_verbal_list = calcSD_more_better(z_score_list)
        #Makes a verbal list to input to the users
        printable_list = outputPrintlist(result_list, z_score_list, z_score_verbal_list)
        #creates a list to be put into a CSV file

        test_name = 'GISD_data.csv' #declares name of the CSV file to save the data in
        #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, printable_list)
        #writes the printable_list in a CSV file
        
            
        console_results = "==================================\nGörsel İşitsel Sayı Dizileri sonuçları:"
        for i in range(len(mean_list)):
            console_results = console_results + (result_name[i] + str(outputConsole_results(result_list, z_score_list, z_score_verbal_list)[i]))            
        console_results = console_results + ("\n==================================")
        #prints the list using the range of the list and makes it more appealing for the user (probably should've used a decorator here)
        
        if norm_exists:
            #txtWrite(patient_admin, patient_ID, console_results)
            print(console_results)
            return [test_name, printable_list, console_results]
            #creates a patient report for the physician and prints it out for the user
        else:
            #txtWrite(patient_admin, patient_ID, ("GISD: Bu grup için norm mevcut değildir.\n"+console_results))
            return [test_name, printable_list, console_results]
            #if there's no norm, it puts "there's no norm for this group" in the patient report
    
    except:
        print("GISD'i değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return
        #saves the program from fiery death

        



def test3ms(): #3MS testi
    try:
        while True:
            try:
                print("\n===================================\n3MS: ")
                result_name = ["\nSonuç: "]  
                #prints user interface
                
                result_list = []               
                result_list.append(floInput(result_name[0]))
                #gets raw input from the user, these are test results
                
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
                #"Only enter numbers", and then resets the function
            
            else:
                break
            
            #tries to get user input, makes sure it's correct input

        printable_list = result_list 
        test_name = '3MS_data.csv' #declares name of the CSV file to save the data in
        #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, result_list)
        #writes the printable_list in a CSV file
        
        console_results = "==================================\n3MS sonuçları:"
        console_results = console_results + (result_name[0] + str(result_list[0]))          
        console_results = console_results + ("\n==================================")
        #txtWrite(patient_admin, patient_ID, console_results)
        
        return [test_name, printable_list, console_results]
    
    except:
        print("3MS'i değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return
        #saves the program from fiery death


def testMonths(): 
    try:
        while True:
            try:
                print("\n===================================\nAyları ileri-geri sayma: ")
                result_name = ["\nİleri yorum: ","\nGeri yorum: "]  
                #prints user interface
                
                result_list = []               
                for i in range(len(result_name)):
                    result_list.append(numInput(result_name[i]))
                #gets raw input from the user, these are test results
                
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
                #"Only enter numbers", and then resets the function
            
            else:
                break
            
            #tries to get user input, makes sure it's correct input

        printable_list = result_list 
        test_name = 'Months_data.csv' #declares name of the CSV file to save the data in
        #csvWriter(patient_admin, patient_ID, patient_age, patient_sex, patient_edu, test_name, result_list)
        #writes the printable_list in a CSV file
        
        console_results = "==================================\nAyları ileri-geri sayma sonuçları:"
        for i in range(len(result_list)):
            console_results = console_results + result_name[i] + str(result_list[i])
        console_results = console_results + ("\n==================================")
        #txtWrite(patient_admin, patient_ID, console_results)
        
        return [test_name, printable_list, console_results]
    
    except:
        print("Ayları ileri saymayı değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return
        #saves the program from fiery death

def testVaNVC(): #Verbal and Nonverbal Cancellation Test, İşaretleme Testi
    try:
        while True:
            try:
                print("\n===================================\nİşaretleme Testi: ")
                result_name = ["\nDüzenli Harfler Formu\nİşaretlenen hedef sayısı puanı: ",
                               "\nAtlanan hedef sayısı puanı: ",
                               "\nİşaretlenen yanlış half sayısı puanı: ",
                               "\nToplam hata puanı: ", "\nTarama süresi puanı: ",
                               "\nDüzenli Şekiller Formu\nİşaretlenen hedef sayısı puanı: ",
                               "\nAtlanan hedef sayısı puanı: ", 
                               "\nİşaretlenen yanlış half sayısı puanı: ", 
                               "\nToplam hata puanı: ", "\nTarama süresi puanı: ",
                               "\nDüzensiz Harfler Formu\nİşaretlenen hedef sayısı puanı: ",
                               "\nAtlanan hedef sayısı puanı: ",
                               "\nİşaretlenen yanlış half sayısı puanı: ", "\nToplam hata puanı: ",
                               "\nTarama süresi puanı: ", 
                               "\nDüzensiz Şekiller Formu\nİşaretlenen hedef sayısı puanı: ",
                               "\nAtlanan hedef sayısı puanı: ", 
                               "\nİşaretlenen yanlış half sayısı puanı: ",
                               "\nToplam hata puanı: ", "\nTarama süresi puanı: "]
                #prints user interface
                
                result_list = []               
                for i in range(len(result_name)):
                    result_list.append(floInput(result_name[i]))
                #gets raw input from the user, these are test results
                    
                mean_list = []
                sd_list = []
                #creates a list for the test results, and empty ones 
                #for the means and standard deviations
                for i in range(len(result_list)):
                    mean_list.append(result_list[i]-999)
                    sd_list.append(1)            
                    #makes it so that it prints out 999 if there's 
                    #no norm calculated for that group
                norm_exists = True
                #It assumes that there's a norm for every group
                #if none exists, it changes to False
                
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
                #"Only enter numbers", and then resets the function
            
            else:
                break
            
            #tries to get user input, makes sure it's correct input
            
            
            
            #Following are lists of means and standard deviations,
            #of the specified age, education and sex, where it applies.
            #It's in a way that it corresponds to order of the
            #result_list.
            
        if 5 <= patient_edu <= 11:
            if 20 <= patient_age <= 54:
                mean_list = [58.87, 1.12, 0.01, 1.13, 120.07, 58.05, 1.94, 0.22,
                             2.16, 113.21, 59.05, 0.96, 0.0, 0.95, 123.18, 59.32,
                             0.68, 0.04, 0.72, 102.86]
                sd_list = [1.64, 1.64, 0.11, 1.64, 44.2, 2.21, 2.23, 0.67, 2.47,
                           43.07, 1.21, 1.21, 0.0, 1.21, 46.62, 1.1, 1.1, 0.23,
                           1.14, 41.81]
                
            elif 55 <= patient_age <= 89:
                mean_list = [58.77, 1.23, 0.05, 1.27, 153.59, 57.86, 2.14, 0.86,
                             3.0, 144.45, 58.68, 1.32, 0.0, 1.32, 160.5, 58.64, 
                             1.36, 0.14, 1.5, 132.45]
                sd_list = [1.77, 1.77, 0.21, 1.8, 42.04, 2.46, 2.46, 1.36, 2.81,
                           44.57, 1.52, 1.52, 0, 1.52, 46.02, 1.56, 1.56, 0.47,
                           1.63, 36.43]
                
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False     
                #"No norm exists for the group", sets it to False
                               
                
        elif 12 <= patient_edu:
            if 20 <= patient_age <= 54:
                mean_list = [59.38, 0.62, 0.01, 0.63, 93.44, 58.14, 1.85, 0.06,
                             1.9, 89.1, 59.16, 0.84, 0.0, 0.84, 97.65, 59.59, 0.41,
                             0.03, 0.45, 78.81]
                sd_list = [1.05, 1.05, 0.11, 1.08, 25.71, 2.02, 2.03, 0.23, 2.1,
                           23.44, 1.19, 1.19, 0, 1.19, 29.62, 0.84, 0.84, 0.18,
                           0.85, 25.33]
                
            elif 55 <= patient_age <= 85:
                mean_list = [58.25, 1.75, 0.0, 1.75, 134.38, 56.94, 3.06, 0.0,
                             3.06, 133.19, 57.88, 2.13, 0.0, 2.13, 148.69, 58.38,
                             1.63, 0.0, 1.63, 123.94]
                sd_list = [2.46, 2.46, 0, 2.46, 38.94, 2.86, 2.86, 0, 2.86,
                           38.21, 2.16, 2.16, 0, 2.16, 48.84, 2.03, 2.03, 0,
                           2.03, 40.38]
            
            else: 
                print("Bu yaş aralığı için norm mevcut değildir.")
                norm_exists = False     
                #"No norm exists for the group", sets it to False
                
        else:
            print("Bu eğitim grubu için norm mevcut değildir.")
            norm_exists = False  
            #"No norm exists for the group", sets it to False

                
        
        z_score_list = calcZscore(result_list, mean_list, sd_list)
        #it calculated the patient's SD interval as a float using the results, means and the SD
        z_score_verbal_list = []
        for i in range(20):
            if i in [0, 5, 10, 15]:
                z_score_verbal_list = z_score_verbal_list + calcSD_more_better([z_score_list[i]])
                
            else:
                z_score_list[i] = -z_score_list[i]
                z_score_verbal_list = z_score_verbal_list + calcSD_less_better([z_score_list[i]])
        #Makes a verbal list to input to the users
        printable_list = outputPrintlist(result_list, z_score_list, z_score_verbal_list)
        #creates a list to be put into a CSV file

        test_name = 'VaNVC_data.csv' #declares name of the CSV file to save the data in
        
              
        console_results = "==================================\nİşaretleme Testi sonuçları:"
        for i in range(len(mean_list)):
            console_results = console_results + (result_name[i] + str(outputConsole_results(result_list, z_score_list, z_score_verbal_list)[i]))            
        console_results = console_results + ("\n==================================")
        #prints the list using the range of the list and makes it more appealing for the user (probably should've used a decorator here)
        
        
        if norm_exists:
            #txtWrite(patient_admin, patient_ID, console_results)
            print(console_results)
            return [test_name, printable_list, console_results]
            #creates a patient report for the physician and prints it out for the user
        else:
            #txtWrite(patient_admin, patient_ID, ("GISD: Bu grup için norm mevcut değildir.\n"+console_results))
            return [test_name, printable_list, console_results]
            #if there's no norm, it puts "there's no norm for this group" in the patient report
    
    except:
        print("İşaretleme testini değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return
        #saves the program from fiery death


def testBNT(): #Boston Naming Test, Boston Adlandırma Testi
    try:
        while True:
            try:
                print("\n===================================\nBoston Adlandırma Testi: ")
                result_name = ["\nKendiliğinden Adlandırma (KA) puanı: ",
                               "\nKendiliğinden Adlandırma ve Anlamsal İpuçları ile Adlandırma (KA+AİA) puanı: ",
                               "\nKendiliğinden Adlandırma, Anlamsal İpuçları ile Adlandırma ve Sessel İpuçları ile Adlandırma (KA+AİA+SİA) puanı: "
                               
                               ]
                #prints user interface
                
                result_list = []               
                for i in range(len(result_name)):
                    result_list.append(floInput(result_name[i]))
                #gets raw input from the user, these are test results
                    
                mean_list = []
                sd_list = []
                #creates a list for the test results, and empty ones 
                #for the means and standard deviations
                for i in range(len(result_list)):
                    mean_list.append(result_list[i]-999)
                    sd_list.append(1)            
                    #makes it so that it prints out 999 if there's 
                    #no norm calculated for that group
                norm_exists = True
                #It assumes that there's a norm for every group
                #if none exists, it changes to False
                
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
                #"Only enter numbers", and then resets the function
            
            else:
                break
            
            #tries to get user input, makes sure it's correct input
            
            
            
            #Following are lists of means and standard deviations,
            #of the specified age, education and sex, where it applies.
            #It's in a way that it corresponds to order of the
            #result_list.
        
        if patient_sex == "Erkek":
            if patient_edu <= 8:
                if 20 <= patient_age <= 39:
                    mean_list = [28.14, 31.36, 36.64]
                    sd_list = [6.22, 5.66, 6.28]
                    
                elif 40 <= patient_age <= 49:
                    mean_list = [30.7, 34.4, 39.4]
                    sd_list = [10.56, 10.32, 9.89]
                    
                elif 50 <= patient_age <= 59:
                    mean_list = [29, 32.8, 38.5]
                    sd_list = [7.8, 8.05, 6.87]
                    
                elif 60 <= patient_age <= 69:
                    mean_list = [25.5, 31.2, 35.3]
                    sd_list = [8.42, 7.45, 7.12]
                    
                elif 70 <= patient_age <= 79:
                    mean_list = [24.2, 28.6, 33.6]
                    sd_list = [9.19, 8.71, 9.25]
                    
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False
            
            elif 8 < patient_edu <= 11:
                if 20 <= patient_age <= 39:
                    mean_list = [41.5, 44.7, 48.7]
                    sd_list = [3.87, 3.65, 2.11]
                    
                elif 40 <= patient_age <= 49:
                    mean_list = [33.9, 37.7, 42.9]
                    sd_list = [6.97, 6.57, 4.46]
                    
                elif 50 <= patient_age <= 59:
                    mean_list = [35.18, 39.36, 42.64]
                    sd_list = [6.52, 5.02, 4.74]
                    
                elif 60 <= patient_age <= 69:
                    mean_list = [35.67, 39.89, 43.55]
                    sd_list = [7.3, 5.86, 4.22]
                    
                elif 70 <= patient_age <= 79:
                    mean_list = [28.2, 32.9, 39.5]
                    sd_list = [9.2, 7.32, 5.7]
                    
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False
                                   
                    
            elif 12 <= patient_edu:
                if 20 <= patient_age <= 39:
                    mean_list = [38.27, 41.82, 46.19]
                    sd_list = [4.65, 4.45, 3.63]
                    
                elif 40 <= patient_age <= 49:
                    mean_list = [42.8, 44.8, 49.1]
                    sd_list = [5.41, 4.45, 3.11]
                    
                elif 50 <= patient_age <= 59:
                    mean_list = [35.82, 39.64, 44.18]
                    sd_list = [8.51, 7.14, 6.91]
                    
                elif 60 <= patient_age <= 69:
                    mean_list = [38.4, 40.54, 45]
                    sd_list = [4.62, 4.41, 3.03]
                    
                elif 70 <= patient_age <= 79:
                    mean_list = [35.27, 38.64, 43.36]
                    sd_list = [5.22, 4.67, 5.94]
                    
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False
                    
            else:
                print("Bu eğitim grubu için norm mevcut değildir.")
                norm_exists = False  
                #"No norm exists for the group", sets it to False
                    
        elif patient_sex == "Kadın":
            if patient_edu <= 8:
                if 20 <= patient_age <= 39:
                    mean_list = [28.4, 31.5, 36.9]
                    sd_list = [7.59, 7.34, 6.39]
                    
                elif 40 <= patient_age <= 49:
                    mean_list = [22.73, 27, 32.09]
                    sd_list = [7.43, 7.39, 8]
                    
                elif 50 <= patient_age <= 59:
                    mean_list = [26, 30.7, 35.7]
                    sd_list = [8.14, 7.56, 7.27]
                    
                elif 60 <= patient_age <= 69:
                    mean_list = [26.91, 30.82, 35.82]
                    sd_list = [9.29, 8.93, 9.94]
                    
                elif 70 <= patient_age <= 79:
                    mean_list = [28.3, 33.2, 37.4]
                    sd_list = [6.22, 6.37, 6.55]
                    
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False
            
            elif 8 < patient_edu <= 11:
                if 20 <= patient_age <= 39:
                    mean_list = [31.91, 35.64, 40.91]

                    sd_list = [7.34, 6.87, 6.79]

                    
                elif 40 <= patient_age <= 49:
                    mean_list = [29.5, 32.9, 38.2]

                    sd_list = [7.23, 7, 5.92]

                    
                elif 50 <= patient_age <= 59:
                    mean_list = [30.9, 35, 41.5]

                    sd_list = [4.91, 4.4, 4.43]

                    
                elif 60 <= patient_age <= 69:
                    mean_list = [35.7, 41.3, 44.2]

                    sd_list = [5.19, 3.09, 2.86]

                    
                elif 70 <= patient_age <= 79:
                    mean_list = [29.1, 34.3, 39.7]

                    sd_list = [7.4, 6.96, 5.81]

                    
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False
                                         
                    
            elif 12 <= patient_edu:
                if 20 <= patient_age <= 39:
                    mean_list = [40.25, 43.33, 47.67]
                    sd_list = [5.4, 4.29, 3.97]
                    
                elif 40 <= patient_age <= 49:
                    mean_list = [40, 43.17, 46.83]
                    sd_list = [4.49, 3.51, 3.24]
                    
                elif 50 <= patient_age <= 59:
                    mean_list = [36.8, 40.1, 46]
                    sd_list = [3.26, 3.79, 3.86]
                    
                elif 60 <= patient_age <= 69:
                    mean_list = [38.4, 42.4, 46.5]
                    sd_list = [4.62, 4.03, 2.32]
                    
                elif 70 <= patient_age <= 79:
                    mean_list = [34.9, 39.6, 43.8]
                    sd_list = [6.37, 5.08, 4.24]
                    
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False
                
            else:
                print("Bu eğitim grubu için norm mevcut değildir.")
                norm_exists = False  
                #"No norm exists for the group", sets it to False
        
        else:
            print("Bir şeyler bir yerde yanlış gitti. Lütfen yazılımcıya ulaşınız.")
            norm_exists = False  
            #"No norm exists for the group", sets it to False

                
        
        z_score_list = calcZscore(result_list, mean_list, sd_list)
        #it calculated the patient's SD interval as a float using the results, means and the SD
        z_score_verbal_list = calcSD_more_better(z_score_list)
        #Makes a verbal list to input to the users
        printable_list = outputPrintlist(result_list, z_score_list, z_score_verbal_list)
        #creates a list to be put into a CSV file

        test_name = 'BNT_data.csv' #declares name of the CSV file to save the data in
        
              
        console_results = "==================================\nBoston Adlandırma Testi sonuçları:"
        for i in range(len(mean_list)):
            console_results = console_results + (result_name[i] + str(outputConsole_results(result_list, z_score_list, z_score_verbal_list)[i]))            
        console_results = console_results + ("\n==================================")
        #prints the list using the range of the list and makes it more appealing for the user (probably should've used a decorator here)
        
        
        if norm_exists:
            #txtWrite(patient_admin, patient_ID, console_results)
            print(console_results)
            return [test_name, printable_list, console_results]
            #creates a patient report for the physician and prints it out for the user
        else:
            #txtWrite(patient_admin, patient_ID, ("GISD: Bu grup için norm mevcut değildir.\n"+console_results))
            return [test_name, printable_list, console_results]
            #if there's no norm, it puts "there's no norm for this group" in the patient report
    
    except:
        print("İşaretleme testini değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return
        #saves the program from fiery death

"""
#TEMPLATE FOR LATER USE
def testBNT(): #Boston Naming Test, Boston Adlandırma Testi
    try:
        while True:
            try:
                print("\n===================================\nBoston Adlandırma Testi: ")
                result_name = ["\nKendiliğinden Adlandırma (KA) puanı: ",
                               "\nKendiliğinden Adlandırma ve Anlamsal İpuçları ile Adlandırma (KA+AİA) puanı: ",
                               "\nKendiliğinden Adlandırma, Anlamsal İpuçları ile Adlandırma ve Sessel İpuçları ile Adlandırma (KA+AİA+SİA) puanı: "
                               
                               ]
                #prints user interface
                
                result_list = []               
                for i in range(len(result_name)):
                    result_list.append(floInput(result_name[i]))
                #gets raw input from the user, these are test results
                    
                mean_list = []
                sd_list = []
                #creates a list for the test results, and empty ones 
                #for the means and standard deviations
                for i in range(len(result_list)):
                    mean_list.append(result_list[i]-999)
                    sd_list.append(1)            
                    #makes it so that it prints out 999 if there's 
                    #no norm calculated for that group
                norm_exists = True
                #It assumes that there's a norm for every group
                #if none exists, it changes to False
                
            except SystemExit:
                raise
            except:
                print("Lütfen sadece sayı giriniz.")
                continue
                #"Only enter numbers", and then resets the function
            
            else:
                break
            
            #tries to get user input, makes sure it's correct input
            
            
            
            #Following are lists of means and standard deviations,
            #of the specified age, education and sex, where it applies.
            #It's in a way that it corresponds to order of the
            #result_list.
        
        if patient_sex == "Erkek":
            if patient_edu <= 8:
                if 20 <= patient_age <= 39:
                    mean_list = []
                    sd_list = []
                    
                elif 40 <= patient_age <= 49:
                    mean_list = []
                    sd_list = []
                    
                elif 50 <= patient_age <= 59:
                    mean_list = []
                    sd_list = []
                    
                elif 60 <= patient_age <= 69:
                    mean_list = []
                    sd_list = []
                    
                elif 70 <= patient_age <= 79:
                    mean_list = []
                    sd_list = []
                    
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False
            
            elif 8 < patient_edu <= 11:
                if 20 <= patient_age <= 39:
                    mean_list = []
                    sd_list = []
                    
                elif 40 <= patient_age <= 49:
                    mean_list = []
                    sd_list = []
                    
                elif 50 <= patient_age <= 59:
                    mean_list = []
                    sd_list = []
                    
                elif 60 <= patient_age <= 69:
                    mean_list = []
                    sd_list = []
                    
                elif 70 <= patient_age <= 79:
                    mean_list = []
                    sd_list = []
                    
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False
                                   
                    
            elif 12 <= patient_edu:
                if 20 <= patient_age <= 39:
                    mean_list = []
                    sd_list = []
                    
                elif 40 <= patient_age <= 49:
                    mean_list = []
                    sd_list = []
                    
                elif 50 <= patient_age <= 59:
                    mean_list = []
                    sd_list = []
                    
                elif 60 <= patient_age <= 69:
                    mean_list = []
                    sd_list = []
                    
                elif 70 <= patient_age <= 79:
                    mean_list = []
                    sd_list = []
                    
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False
                    
            else:
                print("Bu eğitim grubu için norm mevcut değildir.")
                norm_exists = False  
                #"No norm exists for the group", sets it to False
                    
        elif patient_sex == "Kadın":
            if patient_edu <= 8:
                if 20 <= patient_age <= 39:
                    mean_list = []
                    sd_list = []
                    
                elif 40 <= patient_age <= 49:
                    mean_list = []
                    sd_list = []
                    
                elif 50 <= patient_age <= 59:
                    mean_list = []
                    sd_list = []
                    
                elif 60 <= patient_age <= 69:
                    mean_list = []
                    sd_list = []
                    
                elif 70 <= patient_age <= 79:
                    mean_list = []
                    sd_list = []
                    
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False
            
            elif 8 < patient_edu <= 11:
                if 20 <= patient_age <= 39:
                    mean_list = []
                    sd_list = []
                    
                elif 40 <= patient_age <= 49:
                    mean_list = []
                    sd_list = []
                    
                elif 50 <= patient_age <= 59:
                    mean_list = []
                    sd_list = []
                    
                elif 60 <= patient_age <= 69:
                    mean_list = []
                    sd_list = []
                    
                elif 70 <= patient_age <= 79:
                    mean_list = []
                    sd_list = []
                    
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False
                                         
                    
            elif 12 <= patient_edu:
                if 20 <= patient_age <= 39:
                    mean_list = []
                    sd_list = []
                    
                elif 40 <= patient_age <= 49:
                    mean_list = []
                    sd_list = []
                    
                elif 50 <= patient_age <= 59:
                    mean_list = []
                    sd_list = []
                    
                elif 60 <= patient_age <= 69:
                    mean_list = []
                    sd_list = []
                    
                elif 70 <= patient_age <= 79:
                    mean_list = []
                    sd_list = []
                    
                else: 
                    print("Bu yaş aralığı için norm mevcut değildir.")
                    norm_exists = False     
                    #"No norm exists for the group", sets it to False
                
            else:
                print("Bu eğitim grubu için norm mevcut değildir.")
                norm_exists = False  
                #"No norm exists for the group", sets it to False
        
        else:
            print("Bir şeyler bir yerde yanlış gitti. Lütfen yazılımcıya ulaşınız.")
            norm_exists = False  
            #"No norm exists for the group", sets it to False

                
        
        z_score_list = calcZscore(result_list, mean_list, sd_list)
        #it calculated the patient's SD interval as a float using the results, means and the SD
        z_score_verbal_list = []
        for i in range(20):
            if i in [4, 9, 14, 19]:
                z_score_verbal_list = z_score_verbal_list + calcSD_less_better([z_score_list[i]])
                
            else:
                z_score_verbal_list = z_score_verbal_list + calcSD_more_better([z_score_list[i]])
        #Makes a verbal list to input to the users
        printable_list = outputPrintlist(result_list, z_score_list, z_score_verbal_list)
        #creates a list to be put into a CSV file

        test_name = 'VaNVC_data.csv' #declares name of the CSV file to save the data in
        
              
        console_results = "==================================\nİşaretleme Testi sonuçları:"
        for i in range(len(mean_list)):
            console_results = console_results + (result_name[i] + str(outputConsole_results(result_list, z_score_list, z_score_verbal_list)[i]))            
        console_results = console_results + ("\n===================================\n")
        #prints the list using the range of the list and makes it more appealing for the user (probably should've used a decorator here)
        
        
        if norm_exists:
            #txtWrite(patient_admin, patient_ID, console_results)
            print(console_results)
            return [test_name, printable_list, console_results]
            #creates a patient report for the physician and prints it out for the user
        else:
            #txtWrite(patient_admin, patient_ID, ("GISD: Bu grup için norm mevcut değildir.\n"+console_results))
            return [test_name, printable_list, console_results]
            #if there's no norm, it puts "there's no norm for this group" in the patient report
    
    except:
        print("İşaretleme testini değerlendirirken bir hata oluştu, program kapatılacak.")
        raise
        return
        #saves the program from fiery death


        
"""

def progStructure():
    first_run = True
    if first_run:
        import os
        import sys
        os.system("mode con: cols=160 lines=50")
        print("""
================================================
 PsiNorm Persentil Hesaplayıcı - 1.6.11
================================================
 Copyright (C) 2017 Bilal Bahadır Akbulut & Yavuz Ayhan
 
 Bu programın KESİNLİKLE HİÇBİR TEMİNATI YOKTUR; 
 ayrıntılar için programın içinde bulunduğu klasördeki lisans dosyalarına başvurunuz.

 Bu bir özgür yazılımdır, ve bazı koşullar altında yeniden dağıtmakta serbestsiniz;
 ayrıntılar için programın içinde bulunduğu klasördeki lisans dosyalarına başvurunuz.
================================================
          BU EKRANI KAPATMAYINIZ
================================================
""")
        if settings('first_ever_run'):
            if guiAgreeTerms():
                configFilePath = getUserDocumentsPsiPath() + "settings.ini"
                from configobj import ConfigObj
                config = ConfigObj(configFilePath, encoding='UTF8')
                config['General']['first_ever_run'] = "False"
                config.write()
                
                
        guiStartupMenu()
        mainStartup()
        first_run = False
    
    while True:
        guiStartupMenu()
        mainStartup()
        
    sys.exit()
        #informs the user data has been saved then restarts
 
    
    
import logging
try:
    progStructure()
    
except SystemExit:
    raise
    
except Exception as e:
    crashlogPath = getUserDocumentsPsiPath() + ("psinorm.log")
    
    guiSimplePopup("Kritik hata!", "Bir hata oluştu ve program kapatılacak. \n\n'Documents' veya 'Belgelerim' içerisinde, 'PsiNorm' klasörünün içerisinde bulunan 'psinorm.log' dosyasını b.bahadirakbulut@gmail.com'a yönlendiriniz.")
    logging.basicConfig(level=logging.DEBUG, filename= crashlogPath, filemode='w')
    logging.exception("Bu dosyayı email ile b.bahadirakbulut@gmail.com adresine gönderiniz.")   
    logging.shutdown()
